"""
build_value_capture_workbook.py
Builds value_capture_builder.xlsx from scratch. Run this file to (re)generate the workbook --
do not hand-edit the generated xlsx's formulas, edit this script instead.

Sheets:
  INPUT  -- paste SQL block output here (contract columns + arm_role + success_pick)
  PAIRED -- pools cohort_months, pairs test/control per (mne, test_desc, stratum)
  TESTS  -- one row per test contrast; stratified two-proportion z-test via Excel formulas
  REPORT -- partner's 12-column layout, formula-linked from TESTS

See value_capture/README.md for the full contract/formula writeup.
"""
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT_PATH = os.path.join(os.path.dirname(__file__), "value_capture_builder.xlsx")

# ---- shared styles ----
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
EXAMPLE_FILL = PatternFill("solid", fgColor="F2F2F2")
EXAMPLE_FONT = Font(italic=True, color="808080")
SEPARATOR_FILL = PatternFill("solid", fgColor="D9D9D9")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COUNT_FMT = "#,##0"
LIFT_FMT = '0.00"pp"'
PVAL_FMT = "0.0000"
PCT_FMT = "0.00%"
Z_FMT = "0.000"

INPUT_LAST_ROW = 1000  # generous fixed range so SUMIFS keep working as rows are pasted in


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate
    ws.row_dimensions[row].height = 30


def style_example_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = EXAMPLE_FILL
        cell.font = EXAMPLE_FONT


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================================
# INPUT
# ============================================================================
wb = Workbook()
ws_in = wb.active
ws_in.title = "INPUT"

input_headers = [
    "mne", "test_desc", "trt_start_dt", "trt_end_dt", "success_name", "stratum",
    "cohort_month", "test_clients", "test_successes", "control_clients", "control_successes",
    "arm_role", "success_pick",
]
ws_in.append(input_headers)
style_header(ws_in, len(input_headers))
set_widths(ws_in, [14, 42, 13, 13, 26, 9, 12, 12, 13, 13, 15, 10, 40])

# --- EXAMPLE rows (illustrative format only -- distinct test_desc so they never pool with real data) ---
example_rows = [
    # PCL illustrative
    ["PCL (EXAMPLE FORMAT)", "EXAMPLE - illustrative PCL row, DELETE before pasting real data",
     date(2026, 5, 1), date(2026, 6, 30), "Credit limit increase accepted", "overall", "2026-05",
     50000, 2500, 100, 4, "both", "responder_cli (illustrative numbers only)"],
    # PCQ illustrative -- test arm row
    ["PCQ (EXAMPLE FORMAT)", "EXAMPLE - illustrative PCQ decile row, DELETE before pasting real data",
     date(2026, 6, 1), date(2026, 6, 30), "App approved", "D1", "2026-06",
     1200, 90, 0, 0, "test", "approved_asc (illustrative - this row = NG3_CHLN challenger code)"],
    # PCQ illustrative -- control arm row (same stratum/cohort, opposite arm)
    ["PCQ (EXAMPLE FORMAT)", "EXAMPLE - illustrative PCQ decile row, DELETE before pasting real data",
     date(2026, 6, 1), date(2026, 6, 30), "App approved", "D1", "2026-06",
     0, 0, 1150, 40, "control", "approved_asc (illustrative - this row = NG3_CHMP champion code)"],
    # EXAMPLE worked-example -- FUNCTIONAL, feeds PAIRED/TESTS/REPORT for the checkable worked example
    ["EXAMPLE", "Worked Example (two-proportion z-test check)",
     date(2026, 1, 1), date(2026, 1, 31), "Worked-example success (hardcoded check)", "overall", "2026-01",
     1000, 60, 1000, 40, "both", "hardcoded verification row - do not delete, formulas check against this"],
]
for row in example_rows:
    ws_in.append(row)

first_data_row = 2
last_example_row = 1 + len(example_rows)
for r in range(first_data_row, last_example_row + 1):
    style_example_row(ws_in, r, len(input_headers))

# blank pre-formatted rows for pasting real SQL output
BLANK_ROWS_END = 100
for r in range(last_example_row + 1, BLANK_ROWS_END + 1):
    for c in (8, 9, 10, 11):
        ws_in.cell(row=r, column=c).number_format = COUNT_FMT

# date format on trt_start_dt / trt_end_dt columns
for r in range(2, BLANK_ROWS_END + 1):
    ws_in.cell(row=r, column=3).number_format = "yyyy-mm-dd"
    ws_in.cell(row=r, column=4).number_format = "yyyy-mm-dd"

# arm_role dropdown
dv = DataValidation(type="list", formula1='"test,control,both"', allow_blank=True)
ws_in.add_data_validation(dv)
dv.add(f"L2:L{BLANK_ROWS_END}")

# ============================================================================
# PAIRED
# ============================================================================
ws_p = wb.create_sheet("PAIRED")
paired_headers = ["mne", "test_desc", "stratum", "n1", "x1", "n0", "x0",
                   "p1", "p0", "d", "w", "pbar", "var", "wd", "w2var"]
ws_p.append(paired_headers)
style_header(ws_p, len(paired_headers))
set_widths(ws_p, [10, 46, 9, 10, 10, 10, 10, 9, 9, 9, 10, 9, 11, 11, 12])

PAIRED_LAST_ROW = 13  # header(1) + PCL(1) + PCQ deciles(10) + EXAMPLE(1)

paired_rows = [("PCL", "Sales Modal (served) vs BAU (not served)", "overall")]
for i in range(1, 11):
    paired_rows.append(("PCQ", "Modal Sales assignment (challenger) vs champion", f"D{i}"))
paired_rows.append(("EXAMPLE", "Worked Example (two-proportion z-test check)", "overall"))

IN = "INPUT"
for i, (mne, test_desc, stratum) in enumerate(paired_rows):
    r = 2 + i
    ws_p.cell(row=r, column=1, value=mne)
    ws_p.cell(row=r, column=2, value=test_desc)
    ws_p.cell(row=r, column=3, value=stratum)

    def sumifs(col_letter, arm_val):
        return (f'SUMIFS({IN}!${col_letter}$2:${col_letter}${INPUT_LAST_ROW},'
                f'{IN}!$A$2:$A${INPUT_LAST_ROW},$A{r},'
                f'{IN}!$B$2:$B${INPUT_LAST_ROW},$B{r},'
                f'{IN}!$F$2:$F${INPUT_LAST_ROW},$C{r},'
                f'{IN}!$L$2:$L${INPUT_LAST_ROW},"{arm_val}")')

    ws_p.cell(row=r, column=4, value=f'={sumifs("H", "test")}+{sumifs("H", "both")}')       # n1
    ws_p.cell(row=r, column=5, value=f'={sumifs("I", "test")}+{sumifs("I", "both")}')       # x1
    ws_p.cell(row=r, column=6, value=f'={sumifs("J", "control")}+{sumifs("J", "both")}')    # n0
    ws_p.cell(row=r, column=7, value=f'={sumifs("K", "control")}+{sumifs("K", "both")}')    # x0

    ws_p.cell(row=r, column=8, value=f'=IF(D{r}=0,0,E{r}/D{r})')                            # p1
    ws_p.cell(row=r, column=9, value=f'=IF(F{r}=0,0,G{r}/F{r})')                             # p0
    ws_p.cell(row=r, column=10, value=f'=H{r}-I{r}')                                         # d
    ws_p.cell(row=r, column=11, value=f'=IF((D{r}+F{r})=0,0,(D{r}*F{r})/(D{r}+F{r}))')       # w
    ws_p.cell(row=r, column=12, value=f'=IF((D{r}+F{r})=0,0,(E{r}+G{r})/(D{r}+F{r}))')       # pbar
    ws_p.cell(row=r, column=13,
              value=f'=L{r}*(1-L{r})*(IF(D{r}=0,0,1/D{r})+IF(F{r}=0,0,1/F{r}))')             # var
    ws_p.cell(row=r, column=14, value=f'=K{r}*J{r}')                                         # wd
    ws_p.cell(row=r, column=15, value=f'=(K{r}^2)*M{r}')                                     # w2var

    for col, fmt in ((4, COUNT_FMT), (5, COUNT_FMT), (6, COUNT_FMT), (7, COUNT_FMT),
                      (8, PCT_FMT), (9, PCT_FMT), (10, PCT_FMT),
                      (11, "0.00"), (12, PCT_FMT), (13, "0.000000"),
                      (14, "0.000000"), (15, "0.000000")):
        ws_p.cell(row=r, column=col).number_format = fmt

    if mne == "EXAMPLE":
        style_example_row(ws_p, r, len(paired_headers))

# ============================================================================
# TESTS
# ============================================================================
ws_t = wb.create_sheet("TESTS")
tests_headers = ["mne", "DESC", "Type", "test_desc", "success_name", "trt_start_dt", "trt_end_dt",
                  "reference_document", "notes",
                  "p1", "p0", "se", "z", "p_value", "",
                  "Leads", "Lift_pp", "Significance"]
ws_t.append(tests_headers)
style_header(ws_t, len(tests_headers))
set_widths(ws_t, [10, 22, 22, 46, 26, 13, 13, 20, 50,
                   9, 9, 10, 8, 9, 3,
                   11, 10, 12])
ws_t.cell(row=1, column=15).fill = SEPARATOR_FILL
for r in range(2, 6):
    ws_t.cell(row=r, column=15).fill = SEPARATOR_FILL

PA = "PAIRED"


def tests_row(r, mne, desc, ttype, test_desc, success_name, ref_doc, notes):
    ws_t.cell(row=r, column=1, value=mne)
    ws_t.cell(row=r, column=2, value=desc)
    ws_t.cell(row=r, column=3, value=ttype)
    ws_t.cell(row=r, column=4, value=test_desc)
    ws_t.cell(row=r, column=5, value=success_name)
    ws_t.cell(row=r, column=6,
              value=f'=_xlfn.MINIFS({IN}!$C$2:$C${INPUT_LAST_ROW},{IN}!$A$2:$A${INPUT_LAST_ROW},$A{r},'
                    f'{IN}!$B$2:$B${INPUT_LAST_ROW},$D{r})')
    ws_t.cell(row=r, column=7,
              value=f'=_xlfn.MAXIFS({IN}!$D$2:$D${INPUT_LAST_ROW},{IN}!$A$2:$A${INPUT_LAST_ROW},$A{r},'
                    f'{IN}!$B$2:$B${INPUT_LAST_ROW},$D{r})')
    ws_t.cell(row=r, column=6).number_format = "yyyy-mm-dd"
    ws_t.cell(row=r, column=7).number_format = "yyyy-mm-dd"
    ws_t.cell(row=r, column=8, value=ref_doc)
    ws_t.cell(row=r, column=9, value=notes)

    def psum(col_letter):
        return (f'SUMIFS({PA}!${col_letter}$2:${col_letter}${PAIRED_LAST_ROW},'
                f'{PA}!$A$2:$A${PAIRED_LAST_ROW},$A{r},{PA}!$B$2:$B${PAIRED_LAST_ROW},$D{r})')

    n1s, x1s, n0s, x0s = psum("D"), psum("E"), psum("F"), psum("G")
    ws_t.cell(row=r, column=10, value=f'=IF({n1s}=0,0,{x1s}/{n1s})')                          # p1
    ws_t.cell(row=r, column=11, value=f'=IF({n0s}=0,0,{x0s}/{n0s})')                          # p0
    w_sum, wd_sum, w2var_sum = psum("K"), psum("N"), psum("O")
    ws_t.cell(row=r, column=12, value=f'=IF({w_sum}=0,0,SQRT({w2var_sum})/{w_sum})')          # se
    ws_t.cell(row=r, column=13,
              value=f'=IF(L{r}=0,0,({wd_sum}/{w_sum})/L{r})')                                 # z (guards on se)
    ws_t.cell(row=r, column=14,
              value=f'=IF(M{r}=0,1,2*(1-_xlfn.NORM.S.DIST(ABS(M{r}),TRUE)))')                 # p_value

    ws_t.cell(row=r, column=16,
              value=(f'=SUMIFS({IN}!$H$2:$H${INPUT_LAST_ROW},{IN}!$A$2:$A${INPUT_LAST_ROW},$A{r},'
                      f'{IN}!$B$2:$B${INPUT_LAST_ROW},$D{r},{IN}!$L$2:$L${INPUT_LAST_ROW},"test")'
                      f'+SUMIFS({IN}!$H$2:$H${INPUT_LAST_ROW},{IN}!$A$2:$A${INPUT_LAST_ROW},$A{r},'
                      f'{IN}!$B$2:$B${INPUT_LAST_ROW},$D{r},{IN}!$L$2:$L${INPUT_LAST_ROW},"both")'))  # Leads
    ws_t.cell(row=r, column=17, value=f'=IF({w_sum}=0,0,({wd_sum}/{w_sum})*100)')             # Lift_pp
    ws_t.cell(row=r, column=18, value=f'=IF(N{r}<0.05,"Y","N")')                              # Significance

    fmt_map = {10: PCT_FMT, 11: PCT_FMT, 12: "0.000000", 13: Z_FMT, 14: PVAL_FMT,
               16: COUNT_FMT, 17: LIFT_FMT}
    for col, fmt in fmt_map.items():
        ws_t.cell(row=r, column=col).number_format = fmt


tests_row(2, "PCL", "PLI Sales Modal", "Champion/Challenger",
          "Sales Modal (served) vs BAU (not served)", "Credit limit increase accepted",
          "", "1 stratum (overall) -> reduces to a plain two-proportion z-test.")
tests_row(3, "PCQ", "PCQ Modal Sales", "Champion/Challenger (assignment)",
          "Modal Sales assignment (challenger) vs champion", "App approved (Period-ASC) - CONFIRM",
          "",
          "Decile strata D1..D10 in PAIRED rows 3-12. Arm codes (reverify before mapping): "
          "champion=NG3_CHMP, challenger=NG3_CHLN/NG3_CHLG.")
tests_row(4, "EXAMPLE", "Worked-example verification row - DELETE before submitting", "N/A",
          "Worked Example (two-proportion z-test check)", "Worked-example success (hardcoded check)",
          "N/A",
          "Hardcoded n1=1000,x1=60,n0=1000,x0=40 -> expect lift=2.00pp, z~2.052, p~0.0402, Sig=Y. "
          "Delete this row (and its PAIRED/INPUT counterparts) before submitting real numbers.")

for r in (2, 3):
    ws_t.cell(row=r, column=8).fill = PatternFill("solid", fgColor="FFF2CC")  # flag: fill in manually
style_example_row(ws_t, 4, len(tests_headers))

# ============================================================================
# REPORT
# ============================================================================
ws_r = wb.create_sheet("REPORT")
report_headers = ["MNE", "DESC", "Type", "Test Desc", "Treatment Start Date", "Treatment End Date",
                   "Success", "Leads/Unique Clients", "Lift", "P-value/Significance",
                   "Reference Document", "Notes"]
ws_r.append(report_headers)
style_header(ws_r, len(report_headers))
set_widths(ws_r, [10, 22, 22, 46, 16, 16, 26, 16, 10, 18, 20, 50])

TS = "TESTS"
for i, tr in enumerate((2, 3, 4)):
    r = 2 + i
    ws_r.cell(row=r, column=1, value=f'={TS}!A{tr}')
    ws_r.cell(row=r, column=2, value=f'={TS}!B{tr}')
    ws_r.cell(row=r, column=3, value=f'={TS}!C{tr}')
    ws_r.cell(row=r, column=4, value=f'={TS}!D{tr}')
    ws_r.cell(row=r, column=5, value=f'={TS}!F{tr}')
    ws_r.cell(row=r, column=6, value=f'={TS}!G{tr}')
    ws_r.cell(row=r, column=7, value=f'={TS}!E{tr}')
    ws_r.cell(row=r, column=8, value=f'={TS}!P{tr}')
    ws_r.cell(row=r, column=9, value=f'={TS}!Q{tr}')
    ws_r.cell(row=r, column=10, value=f'=TEXT({TS}!N{tr},"0.0000")&" ("&{TS}!R{tr}&")"')
    ws_r.cell(row=r, column=11, value=f'={TS}!H{tr}')
    ws_r.cell(row=r, column=12, value=f'={TS}!I{tr}')

    ws_r.cell(row=r, column=5).number_format = "yyyy-mm-dd"
    ws_r.cell(row=r, column=6).number_format = "yyyy-mm-dd"
    ws_r.cell(row=r, column=8).number_format = COUNT_FMT
    ws_r.cell(row=r, column=9).number_format = LIFT_FMT

    if tr == 4:
        style_example_row(ws_r, r, len(report_headers))

# borders on all populated cells, all sheets
for ws in (ws_in, ws_p, ws_t, ws_r):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = BORDER

wb.save(OUT_PATH)
print(f"wrote {OUT_PATH}")
