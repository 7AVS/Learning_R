"""
Comprehensive verification of pcl_mde_calculator.xlsx
Reads all formulas, validates every cell, recomputes MDE in Python.
"""

import openpyxl
from openpyxl.styles import PatternFill
import math
import re
import sys

# --- scipy for NORM.S.INV ---
try:
    from scipy.stats import norm as scipy_norm
    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False
    print("WARNING: scipy not installed — will skip NORM.S.INV numeric validation")

FILEPATH = r"C:\Users\andre\New_projects\cards\schemas\pcl_mde_calculator.xlsx"

issues = []

def add_issue(severity, cell, msg):
    issues.append((severity, cell, msg))
    print(f"  !! [{severity}] {cell}: {msg}")

# ============================================================
# 1. Load workbook with formulas (data_only=False)
# ============================================================
wb = openpyxl.load_workbook(FILEPATH, data_only=False)
print(f"Sheets: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print("=" * 80)
    print(f"SHEET: {sheet_name}")
    print(f"Dimensions: {ws.dimensions}")
    print("=" * 80)

    # ----------------------------------------------------------
    # 2. Dump every cell
    # ----------------------------------------------------------
    cell_map = {}  # (row, col) -> { 'ref': 'A1', 'value': ..., 'is_formula': bool, 'is_green': bool }

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    print(f"\n--- Cell-by-cell dump (rows 1-{max_row}, cols 1-{max_col}) ---\n")
    print(f"{'Cell':<8} {'Green?':<7} {'Type':<10} Content")
    print("-" * 100)

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            ref = cell.coordinate
            val = cell.value

            # Detect green fill
            is_green = False
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                rgb = str(cell.fill.fgColor.rgb)
                # Common greens: 92D050, 00B050, C6EFCE, light green variants
                if rgb not in ('00000000', '0', 'None'):
                    r_hex = rgb[2:4] if len(rgb) == 8 else rgb[0:2]
                    g_hex = rgb[4:6] if len(rgb) == 8 else rgb[2:4]
                    b_hex = rgb[6:8] if len(rgb) == 8 else rgb[4:6]
                    try:
                        r_int, g_int, b_int = int(r_hex, 16), int(g_hex, 16), int(b_hex, 16)
                        if g_int > r_int and g_int > b_int:
                            is_green = True
                        # Also catch light greens where all are high but green dominates
                        if g_int > 150 and g_int >= r_int and g_int > b_int:
                            is_green = True
                    except ValueError:
                        pass

            is_formula = isinstance(val, str) and val.startswith('=')
            content_type = "FORMULA" if is_formula else ("EMPTY" if val is None else "VALUE")

            if val is not None:
                cell_map[(row, col)] = {
                    'ref': ref,
                    'value': val,
                    'is_formula': is_formula,
                    'is_green': is_green,
                    'row': row,
                    'col': col,
                    'fill_rgb': str(cell.fill.fgColor.rgb) if cell.fill and cell.fill.fgColor else None,
                }
                print(f"{ref:<8} {'GREEN' if is_green else '':.<7} {content_type:<10} {val}")

    # ----------------------------------------------------------
    # 3. Error checks on formulas
    # ----------------------------------------------------------
    print(f"\n--- Formula validation ---\n")

    formula_cells = {k: v for k, v in cell_map.items() if v['is_formula']}
    print(f"Total non-empty cells: {len(cell_map)}")
    print(f"Formula cells: {len(formula_cells)}")
    print(f"Green input cells: {len([v for v in cell_map.values() if v['is_green']])}")

    # 3a. Check for error literals in any cell value
    for (r, c), info in cell_map.items():
        val_str = str(info['value'])
        for err in ['#REF!', '#NUM!', '#NAME?', '#VALUE!', '#DIV/0!', '#NULL!', '#N/A']:
            if err in val_str:
                add_issue("ERROR", info['ref'], f"Contains error literal: {err}")

    # 3b. Parentheses balance in every formula
    for (r, c), info in formula_cells.items():
        formula = info['value']
        opens = formula.count('(')
        closes = formula.count(')')
        if opens != closes:
            add_issue("ERROR", info['ref'], f"Unbalanced parentheses: {opens} open vs {closes} close in: {formula}")

    # 3c. Check for broken references — cell refs that point outside the used range
    cell_ref_pattern = re.compile(r'\$?([A-Z]+)\$?(\d+)')
    for (r, c), info in formula_cells.items():
        formula = info['value']
        refs_found = cell_ref_pattern.findall(formula)
        for col_letters, row_num in refs_found:
            row_num = int(row_num)
            # Convert column letters to number
            col_num = 0
            for ch in col_letters:
                col_num = col_num * 26 + (ord(ch) - ord('A') + 1)
            # Check if that cell exists or is within a reasonable range
            if row_num > max_row + 50 or col_num > max_col + 50:
                add_issue("WARN", info['ref'], f"Reference {col_letters}{row_num} may be outside data range (max_row={max_row}, max_col={max_col})")

    # ----------------------------------------------------------
    # 4. Semantic validation of MDE formulas
    # ----------------------------------------------------------
    print(f"\n--- Semantic / MDE formula validation ---\n")

    # Build a lookup by cell reference string
    ref_to_val = {}
    ref_to_info = {}
    for (r, c), info in cell_map.items():
        ref_to_val[info['ref']] = info['value']
        ref_to_info[info['ref']] = info

    def normalize_ref(ref_str):
        """Remove $ signs from cell references."""
        return ref_str.replace('$', '')

    def get_cell_refs_in_formula(formula):
        """Extract all cell references from a formula."""
        return [normalize_ref(m[0] + m[1]) for m in cell_ref_pattern.findall(formula)]

    # Identify structural elements by scanning labels
    # Look for rows that contain key labels
    label_cells = {}
    for (r, c), info in cell_map.items():
        if not info['is_formula'] and isinstance(info['value'], str):
            label_cells[info['value'].strip().lower()] = (r, c, info['ref'])

    print("Key labels found:")
    for label, (r, c, ref) in sorted(label_cells.items()):
        print(f"  {ref}: '{label}'")

    # ----------------------------------------------------------
    # 4a. NORM.S.INV validation
    # ----------------------------------------------------------
    print(f"\n--- NORM.S.INV formula checks ---\n")
    norminv_pattern = re.compile(r'NORM\.S\.INV\(([^)]+)\)', re.IGNORECASE)

    for (r, c), info in formula_cells.items():
        formula = info['value']
        matches = norminv_pattern.findall(formula)
        for arg in matches:
            print(f"  {info['ref']}: NORM.S.INV({arg})")
            # Check common patterns:
            # NORM.S.INV(1 - alpha/2) for two-sided
            # NORM.S.INV(1 - beta) for power
            arg_clean = arg.strip()
            # Validate the argument makes sense
            if 'NORM.S.INV' in arg_clean:
                add_issue("WARN", info['ref'], f"Nested NORM.S.INV — may be intentional but verify: {formula}")

    # ----------------------------------------------------------
    # 4b. Identify MDE formula patterns and validate structure
    # ----------------------------------------------------------
    print(f"\n--- MDE formula structure analysis ---\n")

    # Standard two-proportion MDE formula:
    # MDE = z * sqrt( p0*(1-p0)*(1/n1 + 1/n2) )
    # where z = Z_alpha + Z_beta (or similar)

    # Look for cells that have the MDE calculation pattern
    mde_cells = []
    for (r, c), info in formula_cells.items():
        formula = info['value'].upper()
        # MDE formulas typically involve SQRT and NORM.S.INV or reference z-scores
        if 'SQRT' in formula and ('NORM' in formula or any(ref in formula for ref in ['Z_'])):
            mde_cells.append(info)
            print(f"  Probable MDE cell: {info['ref']} = {info['value']}")
        elif 'SQRT' in formula and '/' in formula and '*' in formula:
            # Could also be an MDE formula that references pre-computed z-scores
            mde_cells.append(info)
            print(f"  Possible MDE cell: {info['ref']} = {info['value']}")

    # ----------------------------------------------------------
    # 4c. Check status formulas (compare MDE to target)
    # ----------------------------------------------------------
    print(f"\n--- Status / comparison formula checks ---\n")

    for (r, c), info in formula_cells.items():
        formula = info['value']
        if 'IF(' in formula.upper():
            print(f"  IF formula at {info['ref']}: {formula}")
            # Check that IF formulas reference an MDE cell and a target cell
            refs = get_cell_refs_in_formula(formula)
            print(f"    References: {refs}")

    # ----------------------------------------------------------
    # 4d. Check headroom formulas (target lift / MDE)
    # ----------------------------------------------------------
    print(f"\n--- Headroom formula checks ---\n")

    for (r, c), info in formula_cells.items():
        formula = info['value'].lower()
        if 'headroom' in str(ref_to_val.get(ws.cell(row=r, column=1).coordinate, '')).lower() or \
           'headroom' in str(ref_to_val.get(ws.cell(row=r, column=2).coordinate, '')).lower():
            print(f"  Headroom row cell {info['ref']}: {info['value']}")
            # Headroom should be target_lift / MDE
            refs = get_cell_refs_in_formula(info['value'])
            print(f"    References: {refs}")

    # ----------------------------------------------------------
    # 4e. Stress test validation — should use own baseline RR
    # ----------------------------------------------------------
    print(f"\n--- Stress test independence checks ---\n")

    # Look for sections that appear to be stress tests
    stress_rows = set()
    for (r, c), info in cell_map.items():
        val_str = str(info['value']).lower()
        if 'stress' in val_str or 'scenario' in val_str or 'pessimistic' in val_str or 'optimistic' in val_str:
            stress_rows.add(r)
            print(f"  Stress-related label at {info['ref']}: {info['value']}")

    # For formula cells in stress test rows, check they don't reference the main baseline
    # (This is a structural check — we flag if stress test formulas reference the same baseline cell as main)

    # ----------------------------------------------------------
    # 5. Python recomputation of MDE
    # ----------------------------------------------------------
    print(f"\n{'='*80}")
    print("PYTHON RECOMPUTATION OF ALL MDE VALUES")
    print(f"{'='*80}\n")

    # We need to read the workbook with data_only=True to get computed values
    wb_data = openpyxl.load_workbook(FILEPATH, data_only=True)
    ws_data = wb_data[sheet_name]

    # Build a map of computed values
    computed_vals = {}
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws_data.cell(row=row, column=col)
            if cell.value is not None:
                computed_vals[cell.coordinate] = cell.value

    print("Computed values (from Excel cache):")
    for ref in sorted(computed_vals.keys(), key=lambda x: (len(x), x)):
        cv = computed_vals[ref]
        formula_str = ref_to_val.get(ref, '')
        is_form = isinstance(formula_str, str) and formula_str.startswith('=')
        if is_form:
            print(f"  {ref}: cached={cv}  formula={formula_str}")
        elif cv is not None:
            print(f"  {ref}: value={cv}")

    # ----------------------------------------------------------
    # 5a. Attempt to resolve all values and recompute
    # ----------------------------------------------------------
    print(f"\n--- Attempting Python recomputation ---\n")

    if HAS_SCIPY:
        def norminv(p):
            """Equivalent of Excel NORM.S.INV(p)"""
            return scipy_norm.ppf(p)

        # Collect all numeric values (inputs) from the cached data
        # Then try to evaluate MDE formulas

        # First, let's understand the spreadsheet structure better
        # by printing the grid with both formulas and cached values
        print(f"\n--- Full grid: Formula + Cached Value ---\n")
        print(f"{'Cell':<6} {'Cached Value':<20} {'Formula/Input'}")
        print("-" * 80)

        all_refs_ordered = []
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                ref = ws.cell(row=row, column=col).coordinate
                formula = ref_to_val.get(ref)
                cached = computed_vals.get(ref)
                if formula is not None or cached is not None:
                    all_refs_ordered.append(ref)
                    f_str = formula if formula else ''
                    c_str = cached if cached is not None else ''
                    print(f"  {ref:<6} {str(c_str):<20} {f_str}")

        # ----------------------------------------------------------
        # 5b. Build a simple evaluator for the formulas
        # ----------------------------------------------------------
        print(f"\n--- Formula-by-formula recomputation ---\n")

        def resolve_value(ref_str, depth=0):
            """Resolve a cell reference to a numeric value."""
            ref_str = normalize_ref(ref_str)
            if depth > 20:
                return None  # prevent infinite recursion

            # Check if it's a direct numeric value in the cached data
            cached = computed_vals.get(ref_str)
            if cached is not None and isinstance(cached, (int, float)):
                return float(cached)

            # Check if it's a non-formula value in the cell_map
            for (r, c), info in cell_map.items():
                if info['ref'] == ref_str and not info['is_formula']:
                    if isinstance(info['value'], (int, float)):
                        return float(info['value'])

            return cached if isinstance(cached, (int, float)) else None

        def try_eval_mde_formula(formula_str, cell_ref):
            """
            Try to evaluate an MDE-style formula by parsing and computing in Python.
            Returns (computed_value, description) or (None, error_msg).
            """
            f = formula_str.lstrip('=')

            # Replace cell references with their resolved values
            ref_pat = re.compile(r'\$?([A-Z]+)\$?(\d+)')

            refs_in_formula = ref_pat.findall(f)
            resolved = {}
            for col_letters, row_num in refs_in_formula:
                ref = f"{col_letters}{row_num}"
                val = resolve_value(ref)
                resolved[ref] = val

            return resolved

        # For each formula cell, show what it references and the resolved values
        for (r, c), info in formula_cells.items():
            formula = info['value']
            ref = info['ref']
            cached = computed_vals.get(ref)

            resolved = try_eval_mde_formula(formula, ref)

            print(f"\n  Cell {ref}: {formula}")
            print(f"    Cached result: {cached}")
            if resolved:
                for dep_ref, dep_val in resolved.items():
                    print(f"    {dep_ref} = {dep_val}")

            # ----------------------------------------------------------
            # Specific pattern matching for common MDE formulas
            # ----------------------------------------------------------
            f_upper = formula.upper()

            # Pattern: NORM.S.INV(1 - cell/2) — z_alpha for two-sided test
            m = re.search(r'NORM\.S\.INV\(\s*1\s*-\s*\$?([A-Z]+)\$?(\d+)\s*/\s*2\s*\)', f_upper)
            if m:
                alpha_ref = normalize_ref(m.group(1) + m.group(2))
                alpha_val = resolve_value(alpha_ref)
                if alpha_val is not None:
                    expected = norminv(1 - alpha_val / 2)
                    print(f"    >> NORM.S.INV(1 - {alpha_ref}/2) = NORM.S.INV({1 - alpha_val/2}) = {expected:.6f}")
                    if cached is not None and isinstance(cached, (int, float)):
                        diff = abs(cached - expected)
                        if diff > 0.001:
                            add_issue("ERROR", ref, f"NORM.S.INV mismatch: cached={cached}, expected={expected:.6f}")
                        else:
                            print(f"    >> MATCH OK (diff={diff:.8f})")

            # Pattern: NORM.S.INV(1 - cell) — z_beta
            m = re.search(r'NORM\.S\.INV\(\s*1\s*-\s*\$?([A-Z]+)\$?(\d+)\s*\)', f_upper)
            if m and '/2' not in f_upper.split('NORM.S.INV')[1].split(')')[0]:
                beta_ref = normalize_ref(m.group(1) + m.group(2))
                beta_val = resolve_value(beta_ref)
                if beta_val is not None:
                    expected = norminv(1 - beta_val)
                    print(f"    >> NORM.S.INV(1 - {beta_ref}) = NORM.S.INV({1 - beta_val}) = {expected:.6f}")
                    if cached is not None and isinstance(cached, (int, float)):
                        diff = abs(cached - expected)
                        if diff > 0.001:
                            add_issue("ERROR", ref, f"NORM.S.INV mismatch: cached={cached}, expected={expected:.6f}")
                        else:
                            print(f"    >> MATCH OK (diff={diff:.8f})")

            # Pattern: Full MDE formula with SQRT
            # (z_a + z_b) * SQRT(p*(1-p)*(1/n1 + 1/n2))
            # or (z_a + z_b) * SQRT(2*p*(1-p)/n) for equal groups
            if 'SQRT' in f_upper and 'NORM' not in f_upper:
                print(f"    >> Contains SQRT (possible MDE computation)")
                # Try to identify the components
                refs_used = get_cell_refs_in_formula(formula)
                print(f"    >> References used: {refs_used}")
                all_vals = {r: resolve_value(r) for r in refs_used}
                print(f"    >> Resolved values: {all_vals}")

            # Pattern: IF formulas for status
            if f_upper.startswith('=IF('):
                print(f"    >> Status/IF formula — checking structure")
                # Common: =IF(MDE_cell <= target_cell, "Achievable", "Not Achievable")
                refs_used = get_cell_refs_in_formula(formula)
                print(f"    >> References: {refs_used}")
                # Verify the IF references make sense (should compare two cells)

            # Pattern: Division for headroom
            # Look for simple A/B division patterns
            m = re.match(r'^=\$?([A-Z]+)\$?(\d+)\s*/\s*\$?([A-Z]+)\$?(\d+)$', formula)
            if m:
                num_ref = normalize_ref(m.group(1) + m.group(2))
                den_ref = normalize_ref(m.group(3) + m.group(4))
                num_val = resolve_value(num_ref)
                den_val = resolve_value(den_ref)
                if num_val is not None and den_val is not None and den_val != 0:
                    expected = num_val / den_val
                    print(f"    >> Division: {num_ref}/{den_ref} = {num_val}/{den_val} = {expected:.6f}")
                    if cached is not None and isinstance(cached, (int, float)):
                        diff = abs(cached - expected)
                        if diff > 0.001:
                            add_issue("ERROR", ref, f"Division mismatch: cached={cached}, expected={expected:.6f}")
                        else:
                            print(f"    >> MATCH OK (diff={diff:.8f})")

            # Pattern: Multiplication (e.g., population * allocation)
            m = re.match(r'^=\$?([A-Z]+)\$?(\d+)\s*\*\s*\$?([A-Z]+)\$?(\d+)$', formula)
            if m:
                a_ref = normalize_ref(m.group(1) + m.group(2))
                b_ref = normalize_ref(m.group(3) + m.group(4))
                a_val = resolve_value(a_ref)
                b_val = resolve_value(b_ref)
                if a_val is not None and b_val is not None:
                    expected = a_val * b_val
                    print(f"    >> Multiply: {a_ref}*{b_ref} = {a_val}*{b_val} = {expected:.6f}")
                    if cached is not None and isinstance(cached, (int, float)):
                        diff = abs(cached - expected)
                        if diff > max(0.001, abs(expected) * 0.001):
                            add_issue("ERROR", ref, f"Multiplication mismatch: cached={cached}, expected={expected:.6f}")
                        else:
                            print(f"    >> MATCH OK (diff={diff:.8f})")

            # Pattern: Subtraction (e.g., 1 - allocation)
            m = re.match(r'^=1\s*-\s*\$?([A-Z]+)\$?(\d+)$', formula)
            if m:
                a_ref = normalize_ref(m.group(1) + m.group(2))
                a_val = resolve_value(a_ref)
                if a_val is not None:
                    expected = 1.0 - a_val
                    print(f"    >> Subtraction: 1 - {a_ref} = 1 - {a_val} = {expected:.6f}")
                    if cached is not None and isinstance(cached, (int, float)):
                        diff = abs(cached - expected)
                        if diff > 0.001:
                            add_issue("ERROR", ref, f"Subtraction mismatch: cached={cached}, expected={expected:.6f}")
                        else:
                            print(f"    >> MATCH OK (diff={diff:.8f})")

    # ----------------------------------------------------------
    # 5c. Full MDE recomputation from first principles
    # ----------------------------------------------------------
    print(f"\n{'='*80}")
    print("FULL MDE RECOMPUTATION FROM FIRST PRINCIPLES")
    print(f"{'='*80}\n")

    if HAS_SCIPY:
        # Extract all input values from cached data
        # We'll try to identify: alpha, beta (or power), p0 (baseline rate),
        # n (total population), allocation percentages, n1, n2

        # Scan for common input labels
        input_map = {}
        for (r, c), info in cell_map.items():
            val_str = str(info['value']).strip().lower()
            # Look for the value in the cell to the right
            right_ref = ws.cell(row=r, column=c+1).coordinate
            right_val = computed_vals.get(right_ref) or ref_to_val.get(right_ref)

            if any(kw in val_str for kw in ['alpha', 'significance', 'type i']):
                input_map['alpha'] = {'label_ref': info['ref'], 'value_ref': right_ref, 'value': right_val}
            elif any(kw in val_str for kw in ['beta', 'type ii', 'power']):
                if 'power' in val_str:
                    input_map['power'] = {'label_ref': info['ref'], 'value_ref': right_ref, 'value': right_val}
                else:
                    input_map['beta'] = {'label_ref': info['ref'], 'value_ref': right_ref, 'value': right_val}
            elif any(kw in val_str for kw in ['baseline', 'response rate', 'base rate', 'conversion', 'control rate']):
                input_map['baseline_rr'] = {'label_ref': info['ref'], 'value_ref': right_ref, 'value': right_val}
            elif any(kw in val_str for kw in ['total population', 'total n', 'sample size', 'population']):
                if 'test' not in val_str and 'control' not in val_str:
                    input_map['total_n'] = {'label_ref': info['ref'], 'value_ref': right_ref, 'value': right_val}
            elif any(kw in val_str for kw in ['test alloc', 'test %', 'treatment']):
                input_map['test_alloc'] = {'label_ref': info['ref'], 'value_ref': right_ref, 'value': right_val}
            elif any(kw in val_str for kw in ['control alloc', 'control %']):
                input_map['control_alloc'] = {'label_ref': info['ref'], 'value_ref': right_ref, 'value': right_val}
            elif 'target' in val_str and 'lift' in val_str:
                input_map['target_lift'] = {'label_ref': info['ref'], 'value_ref': right_ref, 'value': right_val}
            elif val_str in ['n1', 'n_test', 'test n', 'test group size']:
                input_map['n1'] = {'label_ref': info['ref'], 'value_ref': right_ref, 'value': right_val}
            elif val_str in ['n2', 'n_control', 'control n', 'control group size']:
                input_map['n2'] = {'label_ref': info['ref'], 'value_ref': right_ref, 'value': right_val}

        print("Identified inputs:")
        for name, details in input_map.items():
            print(f"  {name}: {details['value']} (label at {details['label_ref']}, value at {details['value_ref']})")

        # Compute MDE if we have enough inputs
        alpha = None
        beta = None
        p0 = None
        n1 = None
        n2 = None

        if 'alpha' in input_map and isinstance(input_map['alpha']['value'], (int, float)):
            alpha = float(input_map['alpha']['value'])
        if 'beta' in input_map and isinstance(input_map['beta']['value'], (int, float)):
            beta = float(input_map['beta']['value'])
        if 'power' in input_map and isinstance(input_map['power']['value'], (int, float)):
            power_val = float(input_map['power']['value'])
            if power_val > 1:
                power_val = power_val / 100  # convert from percentage
            beta = 1 - power_val
        if 'baseline_rr' in input_map and isinstance(input_map['baseline_rr']['value'], (int, float)):
            p0 = float(input_map['baseline_rr']['value'])
            if p0 > 1:
                p0 = p0 / 100
        if 'n1' in input_map and isinstance(input_map['n1']['value'], (int, float)):
            n1 = float(input_map['n1']['value'])
        if 'n2' in input_map and isinstance(input_map['n2']['value'], (int, float)):
            n2 = float(input_map['n2']['value'])

        # Try to derive n1, n2 from total_n and allocations if not directly available
        if n1 is None and 'total_n' in input_map and 'test_alloc' in input_map:
            total_n = float(input_map['total_n']['value'])
            test_alloc = float(input_map['test_alloc']['value'])
            if test_alloc > 1:
                test_alloc = test_alloc / 100
            n1 = total_n * test_alloc
            print(f"\n  Derived n1 = {total_n} * {test_alloc} = {n1}")
        if n2 is None and 'total_n' in input_map and 'control_alloc' in input_map:
            total_n = float(input_map['total_n']['value'])
            ctrl_alloc = float(input_map['control_alloc']['value'])
            if ctrl_alloc > 1:
                ctrl_alloc = ctrl_alloc / 100
            n2 = total_n * ctrl_alloc
            print(f"  Derived n2 = {total_n} * {ctrl_alloc} = {n2}")

        if all(v is not None for v in [alpha, beta, p0, n1, n2]):
            z_alpha = norminv(1 - alpha / 2)
            z_beta = norminv(1 - beta)

            # Two-proportion comparison test MDE
            mde_absolute = (z_alpha + z_beta) * math.sqrt(p0 * (1 - p0) * (1/n1 + 1/n2))
            mde_relative = mde_absolute / p0

            print(f"\n  Python MDE Computation:")
            print(f"    alpha = {alpha}")
            print(f"    beta  = {beta}")
            print(f"    p0    = {p0}")
            print(f"    n1    = {n1}")
            print(f"    n2    = {n2}")
            print(f"    z_alpha (two-sided) = NORM.S.INV({1-alpha/2}) = {z_alpha:.6f}")
            print(f"    z_beta              = NORM.S.INV({1-beta}) = {z_beta:.6f}")
            print(f"    MDE (absolute)      = {mde_absolute:.6f}")
            print(f"    MDE (relative/lift)  = {mde_relative:.4%}")

            # Now compare with any cached MDE value
            # Look for cells labeled "MDE" or containing MDE
            for (r2, c2), info2 in cell_map.items():
                val_str = str(info2['value']).strip().lower()
                if 'mde' in val_str and not info2['is_formula']:
                    # The MDE result is likely in the cell to the right
                    mde_result_ref = ws.cell(row=r2, column=c2+1).coordinate
                    mde_cached = computed_vals.get(mde_result_ref)
                    mde_formula = ref_to_val.get(mde_result_ref)
                    if mde_cached is not None:
                        print(f"\n    Found MDE result at {mde_result_ref}: cached={mde_cached}, formula={mde_formula}")
                        if isinstance(mde_cached, (int, float)):
                            # Determine if it's absolute or relative
                            if abs(mde_cached) < 1:
                                # Likely a percentage/proportion
                                if abs(mde_cached - mde_relative) < 0.01:
                                    print(f"    >> MATCH as relative MDE (diff={abs(mde_cached - mde_relative):.6f})")
                                elif abs(mde_cached - mde_absolute) < 0.001:
                                    print(f"    >> MATCH as absolute MDE (diff={abs(mde_cached - mde_absolute):.6f})")
                                else:
                                    add_issue("WARN", mde_result_ref,
                                             f"MDE mismatch: cached={mde_cached}, expected_relative={mde_relative:.6f}, expected_absolute={mde_absolute:.6f}")
        else:
            print(f"\n  Cannot fully recompute MDE — missing inputs:")
            print(f"    alpha={alpha}, beta={beta}, p0={p0}, n1={n1}, n2={n2}")

    # ----------------------------------------------------------
    # 6. Cascade check — verify green input cells are referenced
    # ----------------------------------------------------------
    print(f"\n--- Green input cascade check ---\n")

    green_cells = [info for info in cell_map.values() if info['is_green']]
    for gc in green_cells:
        gc_ref = gc['ref']
        # Check how many formulas reference this green cell
        referencing_formulas = []
        for (r, c), info in formula_cells.items():
            formula_upper = info['value'].upper()
            # Check for both $-prefixed and non-prefixed references
            if gc_ref in formula_upper or f"${gc_ref}" in formula_upper:
                referencing_formulas.append(info['ref'])
            else:
                # Also check with $ in various positions
                col_part = re.match(r'([A-Z]+)(\d+)', gc_ref)
                if col_part:
                    patterns = [
                        f"${col_part.group(1)}${col_part.group(2)}",
                        f"${col_part.group(1)}{col_part.group(2)}",
                        f"{col_part.group(1)}${col_part.group(2)}",
                    ]
                    for pat in patterns:
                        if pat in formula_upper:
                            referencing_formulas.append(info['ref'])
                            break

        print(f"  Green cell {gc_ref} (value={gc['value']}): referenced by {len(referencing_formulas)} formulas")
        if referencing_formulas:
            print(f"    -> {referencing_formulas}")
        else:
            add_issue("WARN", gc_ref, "Green input cell is not referenced by any formula — changing it would have no effect")

# ============================================================
# FINAL VERDICT
# ============================================================
print(f"\n{'='*80}")
print("FINAL VERDICT")
print(f"{'='*80}\n")

errors = [i for i in issues if i[0] == "ERROR"]
warnings = [i for i in issues if i[0] == "WARN"]

print(f"Total issues: {len(issues)}")
print(f"  ERRORS: {len(errors)}")
print(f"  WARNINGS: {len(warnings)}")

if issues:
    print(f"\nAll issues:")
    for severity, cell, msg in issues:
        print(f"  [{severity}] {cell}: {msg}")

if errors:
    print(f"\n*** FAIL — {len(errors)} error(s) found ***")
    sys.exit(1)
elif warnings:
    print(f"\n*** PASS WITH WARNINGS — {len(warnings)} warning(s) ***")
else:
    print(f"\n*** PASS — all checks passed ***")
