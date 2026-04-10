"""
Diagnose #VALUE! errors in pcl_mde_calculator.xlsx.

openpyxl can't evaluate formulas — it only checks syntax.
This script does deep inspection:
  1. Dumps every formula cell with its exact string
  2. Checks common Excel compatibility issues
  3. Detects quote-prefixed formulas (written as strings)
  4. Manually evaluates formulas by substituting values
  5. Reports all suspected problem cells
"""

import openpyxl
import re
import math
import sys
import io
from pathlib import Path

# Force UTF-8 output on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ── Files to check ──────────────────────────────────────────────────────────
files = [
    Path(r"C:\Users\andre\New_projects\cards\schemas\pcl_mde_calculator.xlsx"),
    Path(r"C:\Users\andre\New_projects\cards\campaigns\PCL\pcl_mde_calculator_v2.xlsx"),
]

SEPARATOR = "=" * 80


def norm_s_inv(p):
    """Approximate NORM.S.INV using inverse error function approach."""
    from math import sqrt, log, pi, copysign
    # Rational approximation (Abramowitz and Stegun 26.2.23)
    if p <= 0 or p >= 1:
        return float('inf') if p >= 1 else float('-inf')
    # Use symmetry
    if p < 0.5:
        return -norm_s_inv(1 - p)
    t = sqrt(-2 * log(1 - p))
    # Coefficients for rational approximation
    c0 = 2.515517
    c1 = 0.802853
    c2 = 0.010328
    d1 = 1.432788
    d2 = 0.189269
    d3 = 0.001308
    return t - (c0 + c1 * t + c2 * t ** 2) / (1 + d1 * t + d2 * t ** 2 + d3 * t ** 3)


def diagnose_file(filepath):
    if not filepath.exists():
        print(f"\n  FILE NOT FOUND: {filepath}")
        return

    print(f"\n{SEPARATOR}")
    print(f"  DIAGNOSING: {filepath}")
    print(SEPARATOR)

    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb.active

    problems = []  # (cell_ref, formula, issue)

    # ── 1. Collect ALL formula cells ────────────────────────────────────────
    print("\n── ALL FORMULA CELLS ──")
    formula_cells = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                coord = cell.coordinate
                formula_cells[coord] = cell.value
                print(f"  {coord:6s} | {cell.value}")

    if not formula_cells:
        print("  (no formula cells found)")

    # ── 2. Check for quote-prefixed formulas (stored as strings) ────────────
    print("\n── QUOTE-PREFIX CHECK (formulas stored as text) ──")
    quote_prefix_found = False
    for row in ws.iter_rows():
        for cell in row:
            # Check quotePrefix attribute
            if hasattr(cell, 'quotePrefix') and cell.quotePrefix:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    problems.append((cell.coordinate, cell.value,
                                     "QUOTE PREFIX — formula stored as text string, not evaluated"))
                    print(f"  *** {cell.coordinate}: quotePrefix=True, value={cell.value}")
                    quote_prefix_found = True
            # Also check if alignment has quotePrefix via style
            if hasattr(cell, 'alignment') and cell.alignment:
                pass  # alignment doesn't carry quotePrefix, but checking anyway

    if not quote_prefix_found:
        print("  No quote-prefixed formulas found (good)")

    # ── 3. Check for merged cells ───────────────────────────────────────────
    print("\n── MERGED CELLS ──")
    merged = list(ws.merged_cells.ranges)
    if merged:
        for m in merged:
            print(f"  Merged range: {m}")
            # Check if any formula references a cell inside a merged range (not top-left)
            for coord, formula in formula_cells.items():
                for cell_in_range in m.cells:
                    r, c = cell_in_range
                    ref = ws.cell(row=r, column=c).coordinate
                    # Skip the top-left cell of the merge (that one is accessible)
                    top_left = ws.cell(row=m.min_row, column=m.min_col).coordinate
                    if ref != top_left and ref in formula:
                        problems.append((coord, formula,
                                         f"References {ref} which is inside merged range {m} (not top-left)"))
                        print(f"  *** {coord} references {ref} inside merge {m}")
    else:
        print("  No merged cells found")

    # ── 4. Formula-level checks ─────────────────────────────────────────────
    print("\n── FORMULA-LEVEL CHECKS ──")

    for coord, formula in formula_cells.items():
        f = formula

        # 4a. NORM.S.INV check
        if "NORM.S.INV" in f:
            print(f"  {coord}: Uses NORM.S.INV — Excel 2010+ required (should be fine)")
            # Check if argument could be out of range [0,1]
            # NORM.S.INV(1-C7) where C7=0.05 → 0.95, OK
            # NORM.S.INV(C8) where C8=0.80, OK

        # 4b. CHAR() check
        if "CHAR(" in f:
            match = re.findall(r"CHAR\((\d+)\)", f)
            for code in match:
                if int(code) > 255:
                    problems.append((coord, f,
                                     f"CHAR({code}) — code > 255, may cause #VALUE! in some Excel versions"))
                    print(f"  *** {coord}: CHAR({code}) code > 255")

        # 4c. TEXT() format string check
        if "TEXT(" in f:
            text_matches = re.findall(r'TEXT\([^,]+,"([^"]+)"\)', f)
            for fmt_str in text_matches:
                print(f"  {coord}: TEXT() format string: \"{fmt_str}\"")
                # Check for common issues
                if "%" in fmt_str:
                    pass  # "0.00%" is valid

        # 4d. Mismatched parentheses
        open_p = f.count("(")
        close_p = f.count(")")
        if open_p != close_p:
            problems.append((coord, f,
                             f"Mismatched parentheses: {open_p} open vs {close_p} close"))
            print(f"  *** {coord}: Mismatched parens ({open_p} open, {close_p} close)")

        # 4e. Division by zero possibility
        div_matches = re.findall(r'1/([A-Z]+\d+)', f)
        for ref in div_matches:
            # Check if the referenced cell could be zero
            ref_cell = ws[ref]
            if ref_cell.value is not None:
                if isinstance(ref_cell.value, (int, float)) and ref_cell.value == 0:
                    problems.append((coord, f, f"Division by {ref} which is 0"))
                    print(f"  *** {coord}: Division by {ref} = 0")

        # 4f. ROUND() argument count
        round_matches = re.findall(r'ROUND\(([^)]+)\)', f)
        for args_str in round_matches:
            # Count top-level commas (not inside nested parens)
            depth = 0
            commas = 0
            for ch in args_str:
                if ch == '(':
                    depth += 1
                elif ch == ')':
                    depth -= 1
                elif ch == ',' and depth == 0:
                    commas += 1
            if commas != 1:
                problems.append((coord, f,
                                 f"ROUND() has {commas + 1} arguments (expected 2)"))
                print(f"  *** {coord}: ROUND() has {commas + 1} args instead of 2")

        # 4g. Semicolon separator check (locale issue)
        if ";" in f:
            problems.append((coord, f,
                             "Uses semicolon (;) — may indicate locale mismatch"))
            print(f"  *** {coord}: Contains semicolon separator")

        # 4h. Check if formula references cells that contain text
        cell_refs = re.findall(r'\$?[A-Z]+\$?\d+', f)
        for ref in cell_refs:
            clean_ref = ref.replace("$", "")
            try:
                ref_cell = ws[clean_ref]
                if ref_cell.value is not None and isinstance(ref_cell.value, str):
                    if not ref_cell.value.startswith("="):
                        problems.append((coord, f,
                                         f"References {clean_ref} which contains text: \"{ref_cell.value[:50]}\""))
            except (ValueError, KeyError):
                pass

        # 4i. Check for empty cell references in formulas
        for ref in cell_refs:
            clean_ref = ref.replace("$", "")
            try:
                ref_cell = ws[clean_ref]
                if ref_cell.value is None:
                    # Only flag if it's used in a math context
                    if any(op in f for op in ["*", "/", "+", "-", "SQRT", "ROUND"]):
                        problems.append((coord, f,
                                         f"References {clean_ref} which is EMPTY — may cause #VALUE! in math context"))
            except (ValueError, KeyError):
                pass

    # ── 5. Manual formula evaluation ────────────────────────────────────────
    print("\n── MANUAL FORMULA EVALUATION ──")

    # Build a value map from all cells
    values = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                values[cell.coordinate] = cell.value

    # Known input values
    inputs = {
        "C3": 528000,
        "C4": 0.30,
        "C5": 0.30,
        "C6": 0.40,
        "C7": 0.05,
        "C8": 0.80,
        "C9": 0.05,
        "C10": 0.13,
    }

    # Compute derived values step by step
    derived = dict(inputs)

    # C14 = NORM.S.INV(1-C7) = NORM.S.INV(0.95)
    derived["C14"] = norm_s_inv(1 - derived["C7"])
    # C15 = NORM.S.INV(C8) = NORM.S.INV(0.80)
    derived["C15"] = norm_s_inv(derived["C8"])

    print(f"  C14 (Za) = NORM.S.INV(0.95) = {derived['C14']:.4f}  (Excel: 1.6449)")
    print(f"  C15 (Zb) = NORM.S.INV(0.80) = {derived['C15']:.4f}  (Excel: 0.8416)")

    # C18 = C10 * C9
    derived["C18"] = derived["C10"] * derived["C9"]
    print(f"  C18 (Target abs lift) = {derived['C10']} * {derived['C9']} = {derived['C18']:.6f}")

    # C19, C20, C21
    derived["C19"] = derived["C3"] * derived["C4"]
    derived["C20"] = derived["C3"] * derived["C5"]
    derived["C21"] = derived["C3"] * derived["C6"]
    print(f"  C19 (Champion n) = {derived['C19']:.0f}")
    print(f"  C20 (Challenger A n) = {derived['C20']:.0f}")
    print(f"  C21 (Challenger B n) = {derived['C21']:.0f}")

    # C22 = C4+C5+C6
    derived["C22"] = derived["C4"] + derived["C5"] + derived["C6"]
    print(f"  C22 (Allocation check) = {derived['C22']:.2f}")

    # Now evaluate each formula row in the comparison table (rows 26-28)
    print("\n  ── Comparison Table Evaluation ──")
    for row_num in range(26, 29):
        print(f"\n  Row {row_num}:")
        d_ref = f"D{row_num}"
        e_ref = f"E{row_num}"
        f_ref = f"F{row_num}"
        g_ref = f"G{row_num}"
        h_ref = f"H{row_num}"
        i_ref = f"I{row_num}"
        j_ref = f"J{row_num}"

        # Get the formulas
        d_formula = formula_cells.get(d_ref, "")
        e_formula = formula_cells.get(e_ref, "")

        # Resolve D and E (n1, n2)
        d_val = None
        e_val = None
        for ref, val in [("C19", derived.get("C19")), ("C20", derived.get("C20")), ("C21", derived.get("C21"))]:
            if ref in d_formula:
                d_val = val
            if ref in e_formula:
                e_val = val

        if d_val is None or e_val is None:
            print(f"    Could not resolve n1/n2 for row {row_num}")
            continue

        derived[d_ref] = d_val
        derived[e_ref] = e_val
        print(f"    {d_ref} (n1) = {d_val:.0f}")
        print(f"    {e_ref} (n2) = {e_val:.0f}")

        # F = D+E (Sum)
        f_val = d_val + e_val
        derived[f_ref] = f_val
        print(f"    {f_ref} (Sum) = {f_val:.0f}")

        # G = D/E (Ratio)
        if e_val != 0:
            g_val = d_val / e_val
            derived[g_ref] = g_val
            print(f"    {g_ref} (Ratio) = {g_val:.4f}")
        else:
            problems.append((g_ref, formula_cells.get(g_ref, ""), "Division by zero (n2=0)"))
            print(f"    *** {g_ref}: DIVISION BY ZERO")

        # H = $C$10 (Baseline RR)
        h_val = derived["C10"]
        derived[h_ref] = h_val
        print(f"    {h_ref} (Baseline RR) = {h_val}")

        # I = $C$18 (Target Lift abs)
        i_val = derived["C18"]
        derived[i_ref] = i_val
        print(f"    {i_ref} (Target Lift abs) = {i_val:.6f}")

        # J = ROUND((Za+Zb)*SQRT(p0*(1-p0)*(1/n1+1/n2)), 4)
        try:
            za = derived["C14"]
            zb = derived["C15"]
            p0 = h_val
            n1 = d_val
            n2 = e_val
            inner = p0 * (1 - p0) * (1 / n1 + 1 / n2)
            if inner < 0:
                problems.append((j_ref, formula_cells.get(j_ref, ""),
                                 f"SQRT of negative number: {inner}"))
                print(f"    *** {j_ref}: SQRT of negative = {inner}")
            else:
                mde = round((za + zb) * math.sqrt(inner), 4)
                derived[j_ref] = mde
                print(f"    {j_ref} (MDE) = {mde:.4f} = {mde * 100:.2f}%")

                # K: Powered?
                if mde <= i_val:
                    print(f"    K{row_num} → YES (MDE {mde:.4f} <= Target {i_val:.4f})")
                else:
                    print(f"    K{row_num} → NO (MDE {mde:.4f} > Target {i_val:.4f})")

                # L: Interpretation
                interp_formula = formula_cells.get(f"L{row_num}", "")
                print(f"    L{row_num} formula: {interp_formula}")
                # Try to evaluate: ="Can detect differences >= "&TEXT(J{row},"0.00%")&" pp"
                # This concatenates a string with TEXT() result
                text_result = f"{mde:.2%}"
                interp_result = f"Can detect differences >= {text_result} pp"
                print(f"    L{row_num} evaluated: \"{interp_result}\"")

        except Exception as e:
            problems.append((j_ref, formula_cells.get(j_ref, ""),
                             f"Evaluation error: {e}"))
            print(f"    *** {j_ref}: Evaluation error: {e}")

    # Stress test rows (39-41)
    print("\n  ── Stress Test Evaluation ──")
    stress_baselines = {39: 0.105, 40: 0.13, 41: 0.154}
    comp_refs = [("C19", "C20"), ("C19", "C21"), ("C20", "C21")]

    for row_num, baseline in stress_baselines.items():
        print(f"\n  Row {row_num} (baseline={baseline}):")
        target_abs = baseline * derived["C9"]
        print(f"    D{row_num} (Target abs) = {target_abs:.6f}")

        mde_vals = []
        for c_idx, (n1_ref, n2_ref) in enumerate(comp_refs):
            col = 5 + c_idx
            col_letter = chr(ord('A') + col - 1)  # E, F, G
            n1 = derived[n1_ref]
            n2 = derived[n2_ref]
            za = derived["C14"]
            zb = derived["C15"]
            inner = baseline * (1 - baseline) * (1 / n1 + 1 / n2)
            mde = round((za + zb) * math.sqrt(inner), 4)
            mde_vals.append(mde)
            cell_ref = f"{col_letter}{row_num}"
            derived[cell_ref] = mde
            print(f"    {cell_ref} (Comp {c_idx + 1} MDE) = {mde:.4f}")

        worst = max(mde_vals)
        powered = "YES" if worst <= target_abs else "NO"
        print(f"    H{row_num} (Worst) = {worst:.4f}, Powered? = {powered}")

    # ── 6. Deep check: openpyxl cell properties ─────────────────────────────
    print("\n── CELL PROPERTY DEEP CHECK ──")
    print("  Checking data_type, style order, protection flags...")

    for coord, formula in formula_cells.items():
        cell = ws[coord]
        dt = cell.data_type
        # In openpyxl, formula cells should have data_type 'f'
        if dt != 'f':
            problems.append((coord, formula,
                             f"data_type is '{dt}' instead of 'f' — formula may be stored as text"))
            print(f"  *** {coord}: data_type='{dt}' (expected 'f')")
        else:
            pass  # Normal

        # Check if cell is protected/locked in a way that could cause issues
        if cell.protection and cell.protection.locked is False:
            pass  # Unlocked is fine for input cells

    # ── 7. Check for the ≥ character in cell values ─────────────────────────
    print("\n── UNICODE CHARACTER CHECK ──")
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                # Check for non-ASCII characters that might cause issues
                non_ascii = [(i, ch, hex(ord(ch))) for i, ch in enumerate(cell.value)
                             if ord(ch) > 127]
                if non_ascii:
                    print(f"  {cell.coordinate}: Non-ASCII chars: {non_ascii}")
                    # The ≥ (U+2265) and — (U+2014) are display-only text, not in formulas
                    # But if they're IN a formula, that's a problem
                    if cell.value.startswith("="):
                        for pos, ch, code in non_ascii:
                            problems.append((cell.coordinate, cell.value,
                                             f"Non-ASCII char '{ch}' ({code}) at position {pos} INSIDE FORMULA"))
                            print(f"  *** {cell.coordinate}: Non-ASCII in formula!")

    # ── 8. Check string concatenation formulas specifically ─────────────────
    print("\n── STRING CONCATENATION FORMULA CHECK ──")
    for coord, formula in formula_cells.items():
        if '&' in formula or 'CONCATENATE' in formula.upper():
            print(f"  {coord}: {formula}")
            # Check if TEXT() result could be #VALUE!
            if 'TEXT(' in formula:
                # Extract what TEXT() references
                text_refs = re.findall(r'TEXT\((\$?[A-Z]+\$?\d+)', formula)
                for ref in text_refs:
                    clean = ref.replace("$", "")
                    # If the referenced cell is itself a formula, we need its result to be numeric
                    ref_cell = ws[clean]
                    if ref_cell.value and isinstance(ref_cell.value, str) and ref_cell.value.startswith("="):
                        print(f"    TEXT() references {clean} which is a formula: {ref_cell.value}")
                        # Check if that formula could produce a string instead of number
                        if "IF(" in ref_cell.value.upper():
                            print(f"    *** WARNING: {clean} contains IF() — if it returns text, TEXT() will #VALUE!")

    # ── 9. Check for any formulas referencing row 34/35 merged cells ────────
    print("\n── MERGED CELL REFERENCE CHECK ──")
    for coord, formula in formula_cells.items():
        for merged_range in merged:
            for cell_in_range in merged_range.cells:
                r, c = cell_in_range
                ref = ws.cell(row=r, column=c).coordinate
                # Skip top-left of merge
                top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col).coordinate
                if ref != top_left:
                    # Check if any formula references this non-top-left merged cell
                    if ref in formula or f"${ref}" in formula:
                        problems.append((coord, formula,
                                         f"References {ref} — inside merged range {merged_range}"))

    # ── 10. Final report ────────────────────────────────────────────────────
    print(f"\n{'=' * 80}")
    print(f"  PROBLEM SUMMARY: {len(problems)} suspected issues")
    print(f"{'=' * 80}")

    if problems:
        for i, (coord, formula, issue) in enumerate(problems, 1):
            print(f"\n  [{i}] Cell {coord}")
            print(f"      Formula: {formula}")
            print(f"      Issue:   {issue}")
    else:
        print("\n  No problems detected by this script.")
        print("\n  BUT — if Excel still shows #VALUE!, the issue might be:")
        print("    1. Excel calc engine disagreement with openpyxl formula writing")
        print("    2. Regional/locale settings (comma vs semicolon)")
        print("    3. The file was saved by openpyxl without a cached value,")
        print("       and Excel's initial recalculation hits a transient error")
        print("    4. A specific Excel version bug with NORM.S.INV or TEXT()")

    # ── 11. Final: check if build script wrote formulas AFTER styles ────────
    print(f"\n── BUILD SCRIPT WRITE-ORDER CHECK ──")
    print("  Checking if value= was set in same call as style properties...")
    print("  (openpyxl bug: setting value= in cell() call + later .value= can cause issues)")
    print()

    # Check cells where value was set as keyword arg AND as .value
    # We can't detect this from the xlsx — need to inspect the build script
    # But we CAN check if any formula cell has a cached value that's wrong
    wb_data = openpyxl.load_workbook(filepath, data_only=True)
    ws_data = wb_data.active
    print("  Cached values (data_only=True) for formula cells:")
    any_cached = False
    for coord in formula_cells:
        cached = ws_data[coord].value
        if cached is not None:
            any_cached = True
            print(f"    {coord}: cached value = {cached}")
        else:
            print(f"    {coord}: NO CACHED VALUE (will recalculate on open)")

    if not any_cached:
        print("\n  *** No cached values for ANY formula cell.")
        print("  *** This is EXPECTED for openpyxl-generated files.")
        print("  *** Excel should recalculate all on first open.")
        print("  *** If #VALUE! appears, it's a formula evaluation error, not a caching issue.")

    wb.close()
    wb_data.close()


# ── Run ──────────────────────────────────────────────────────────────────────
print("PCL MDE Calculator — Formula Diagnostic Script")
print(f"Python: {sys.version}")
print(f"openpyxl: {openpyxl.__version__}")

for f in files:
    diagnose_file(f)

print(f"\n\n{'=' * 80}")
print("DONE")
print(f"{'=' * 80}")
