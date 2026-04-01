"""
Build PCL MDE Summary Excel — clean, read-only presentation for email attachment.

No formulas, no green editable cells, no "Do Not Edit" sections.
All values hardcoded. Blue headers, thin borders, professional styling.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = r"C:\Users\andre\New_projects\cards\campaigns\PCL\pcl_mde_summary.xlsx"

# ── Styles ───────────────────────────────────────────────────────────────────
BLUE_FILL    = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
GREY_FILL    = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
GREEN_FILL   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WHITE_FILL   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
LIGHT_FILL   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

BOLD_WHITE   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
BOLD_DARK    = Font(bold=True, color="1F1F1F", name="Calibri", size=11)
BOLD_SECTION = Font(bold=True, color="1F1F1F", name="Calibri", size=11)
NORMAL       = Font(name="Calibri", size=11)
ITALIC_GREY  = Font(italic=True, color="595959", name="Calibri", size=10)
NOTE_FONT    = Font(name="Calibri", size=10, color="404040")

THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED  = Side(style="medium", color="4472C4")
TOP_BORDER  = Border(top=MED)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top",    wrap_text=True)

PCT2  = "0.00%"
PCT4  = "0.0000%"
NUM   = "#,##0"
PP    = '0.00"%"'   # not used — we write pp as plain text or formatted pct


# ── Helpers ──────────────────────────────────────────────────────────────────

def s(ws, row, col, value=None, font=None, fill=None, fmt=None,
      align=None, border=None, height=None):
    """Set cell value and styles."""
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if font   is not None: cell.font      = font
    if fill   is not None: cell.fill      = fill
    if fmt    is not None: cell.number_format = fmt
    if align  is not None: cell.alignment = align
    if border is not None: cell.border    = border
    if height is not None: ws.row_dimensions[row].height = height
    return cell


def section_header(ws, row, text, ncols=8, start_col=2):
    """Grey section divider row with bold label."""
    for c in range(1, start_col + ncols):
        ws.cell(row=row, column=c).fill = GREY_FILL
    s(ws, row, start_col, text, font=BOLD_SECTION, fill=GREY_FILL,
      align=LEFT, height=18)


def col_header_row(ws, row, headers, start_col=1):
    """Blue column header row."""
    for i, h in enumerate(headers):
        s(ws, row, start_col + i, h,
          font=BOLD_WHITE, fill=BLUE_FILL, align=CENTER, border=THIN_BORDER)
    ws.row_dimensions[row].height = 30


def data_row(ws, row, values, start_col=1, fills=None, fmts=None, bold_cols=None,
             aligns=None):
    """Write a data row with thin borders. fills/fmts/aligns are dicts {col_index: value}."""
    fills     = fills     or {}
    fmts      = fmts      or {}
    bold_cols = bold_cols or set()
    aligns    = aligns    or {}
    for i, v in enumerate(values):
        col = start_col + i
        fill  = fills.get(i, WHITE_FILL)
        fmt   = fmts.get(i)
        align = aligns.get(i, CENTER)
        font  = BOLD_DARK if i in bold_cols else NORMAL
        c = s(ws, row, col, v, font=font, fill=fill, fmt=fmt,
              align=align, border=THIN_BORDER)
    ws.row_dimensions[row].height = 22


# ════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ════════════════════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Recommended Design"

# ── Column widths ────────────────────────────────────────────────────────────
col_widths = {
    "A": 3,   # left margin
    "B": 38,  # main label / scenario col
    "C": 18,  # value / arm description
    "D": 16,
    "E": 14,
    "F": 14,
    "G": 16,
    "H": 16,
    "I": 12,
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# ── Title row ────────────────────────────────────────────────────────────────
ws.merge_cells("B1:I1")
s(ws, 1, 2, "PCL (Pre-Approved Credit Limit Increase) — MDE Summary: Recommended Design",
  font=Font(bold=True, color="FFFFFF", name="Calibri", size=13),
  fill=BLUE_FILL, align=LEFT, height=28)
for c in range(3, 10):
    ws.cell(row=1, column=c).fill = BLUE_FILL


# ════════════════════════════════════════════════════════════════════════════
# SECTION 1 — TEST DESIGN
# ════════════════════════════════════════════════════════════════════════════
r = 3
section_header(ws, r, "Section 1 — Test Design")

col_header_row(ws, r + 1, ["", "Parameter", "Value"], start_col=1)

design_rows = [
    ("Total Monthly Population",          "486,821"),
    ("Baseline RR (FY 2025 mobile avg)",  "12.51%"),
    ("Target Relative Lift",              "5%"),
    ("Target Absolute Lift",              "0.63pp"),
    ("Confidence Level",                  "95%"),
    ("Power",                             "80%"),
]

for i, (param, val) in enumerate(design_rows):
    row = r + 2 + i
    fill_row = LIGHT_FILL if i % 2 == 0 else WHITE_FILL
    s(ws, row, 1, "", fill=fill_row, border=THIN_BORDER)
    s(ws, row, 2, param,  font=BOLD_DARK, fill=fill_row, align=LEFT, border=THIN_BORDER)
    s(ws, row, 3, val,    font=NORMAL,    fill=fill_row, align=CENTER, border=THIN_BORDER)
    # Span remaining columns cleanly
    for c in range(4, 10):
        s(ws, row, c, "", fill=fill_row, border=THIN_BORDER)
    ws.row_dimensions[row].height = 20

# blank row
r = r + 2 + len(design_rows) + 1   # r = 12


# ════════════════════════════════════════════════════════════════════════════
# SECTION 2 — ALLOCATION
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, r, "Section 2 — Allocation")

col_header_row(ws, r + 1,
    ["", "Arm", "Description", "Allocation", "Monthly n"],
    start_col=1)

alloc_rows = [
    ("Champion",     "Product Page + Offers Hub (all channels)",  "10%",  "48,682"),
    ("Challenger A", "Champion + Sales Model placement",          "45%", "219,069"),
    ("Challenger B", "Champion + Mobile Dashboard placement",     "45%", "219,069"),
]

for i, (arm, desc, alloc, n) in enumerate(alloc_rows):
    row = r + 2 + i
    fill_row = LIGHT_FILL if i % 2 == 0 else WHITE_FILL
    s(ws, row, 1, "",    fill=fill_row, border=THIN_BORDER)
    s(ws, row, 2, arm,   font=BOLD_DARK, fill=fill_row, align=LEFT,   border=THIN_BORDER)
    s(ws, row, 3, desc,  font=NORMAL,    fill=fill_row, align=LEFT,   border=THIN_BORDER)
    s(ws, row, 4, alloc, font=NORMAL,    fill=fill_row, align=CENTER, border=THIN_BORDER)
    s(ws, row, 5, n,     font=NORMAL,    fill=fill_row, align=CENTER, border=THIN_BORDER)
    for c in range(6, 10):
        s(ws, row, c, "", fill=fill_row, border=THIN_BORDER)
    ws.row_dimensions[row].height = 20

r = r + 2 + len(alloc_rows) + 1   # r = 17


# ════════════════════════════════════════════════════════════════════════════
# SECTION 3 — COMPARISON TESTS
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, r, "Section 3 — Comparison Tests")

col_header_row(ws, r + 1,
    ["#", "Comparison", "What it tests", "n1", "n2",
     "Baseline RR", "Target Lift", "MDE", "Powered?"],
    start_col=1)

comp_rows = [
    (1, "Champion vs Challenger A",    "Does Sales Model lift response by >=5%?",
     "48,682",  "219,069", "12.51%", "0.63%", "0.41%", "YES"),
    (2, "Champion vs Challenger B",    "Does Dashboard lift response by >=5%?",
     "48,682",  "219,069", "12.51%", "0.63%", "0.41%", "YES"),
    (3, "Challenger A vs Challenger B","Which addition performs better?",
     "219,069", "219,069", "12.51%", "0.63%", "0.25%", "YES"),
]

for i, (num, comp, test, n1, n2, base_rr, tgt, mde, powered) in enumerate(comp_rows):
    row = r + 2 + i
    fill_row = LIGHT_FILL if i % 2 == 0 else WHITE_FILL
    powered_fill = GREEN_FILL if powered == "YES" else PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    s(ws, row, 1, num,     font=NORMAL,    fill=fill_row, align=CENTER, border=THIN_BORDER)
    s(ws, row, 2, comp,    font=BOLD_DARK, fill=fill_row, align=LEFT,   border=THIN_BORDER)
    s(ws, row, 3, test,    font=NORMAL,    fill=fill_row, align=LEFT,   border=THIN_BORDER)
    s(ws, row, 4, n1,      font=NORMAL,    fill=fill_row, align=CENTER, border=THIN_BORDER)
    s(ws, row, 5, n2,      font=NORMAL,    fill=fill_row, align=CENTER, border=THIN_BORDER)
    s(ws, row, 6, base_rr, font=NORMAL,    fill=fill_row, align=CENTER, border=THIN_BORDER)
    s(ws, row, 7, tgt,     font=NORMAL,    fill=fill_row, align=CENTER, border=THIN_BORDER)
    s(ws, row, 8, mde,     font=NORMAL,    fill=fill_row, align=CENTER, border=THIN_BORDER)
    s(ws, row, 9, powered,
      font=Font(bold=True, name="Calibri", size=11, color="375623"),
      fill=powered_fill, align=CENTER, border=THIN_BORDER)
    ws.row_dimensions[row].height = 22

r = r + 2 + len(comp_rows) + 1   # r = 24


# ════════════════════════════════════════════════════════════════════════════
# SECTION 4 — STRESS TESTS
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, r, "Section 4 — Stress Tests")

col_header_row(ws, r + 1,
    ["", "Scenario", "Baseline RR", "Target Lift (abs)",
     "Comp 1 MDE", "Comp 2 MDE", "Comp 3 MDE", "Worst Case", "Powered?"],
    start_col=1)

stress_rows = [
    ("Pessimistic (Jan 2026: 55,560/528,002)", "10.50%", "0.53%",
     "0.38%", "0.38%", "0.23%", "0.38%", "YES"),
    ("Current (FY 2025 month avg mobile)",     "12.51%", "0.63%",
     "0.41%", "0.41%", "0.25%", "0.41%", "YES"),
    ("Optimistic (Nov 2025: 99,704/648,547)",  "15.40%", "0.77%",
     "0.45%", "0.45%", "0.27%", "0.45%", "YES"),
]

for i, (scenario, base_rr, tgt, c1, c2, c3, worst, powered) in enumerate(stress_rows):
    row = r + 2 + i
    fill_row = LIGHT_FILL if i % 2 == 0 else WHITE_FILL
    powered_fill = GREEN_FILL if powered == "YES" else PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    s(ws, row, 1, "",        fill=fill_row,  border=THIN_BORDER)
    s(ws, row, 2, scenario,  font=BOLD_DARK, fill=fill_row, align=LEFT,   border=THIN_BORDER)
    s(ws, row, 3, base_rr,   font=NORMAL,    fill=fill_row, align=CENTER, border=THIN_BORDER)
    s(ws, row, 4, tgt,       font=NORMAL,    fill=fill_row, align=CENTER, border=THIN_BORDER)
    s(ws, row, 5, c1,        font=NORMAL,    fill=fill_row, align=CENTER, border=THIN_BORDER)
    s(ws, row, 6, c2,        font=NORMAL,    fill=fill_row, align=CENTER, border=THIN_BORDER)
    s(ws, row, 7, c3,        font=NORMAL,    fill=fill_row, align=CENTER, border=THIN_BORDER)
    s(ws, row, 8, worst,     font=NORMAL,    fill=fill_row, align=CENTER, border=THIN_BORDER)
    s(ws, row, 9, powered,
      font=Font(bold=True, name="Calibri", size=11, color="375623"),
      fill=powered_fill, align=CENTER, border=THIN_BORDER)
    ws.row_dimensions[row].height = 22

r = r + 2 + len(stress_rows) + 1   # r = 31


# ════════════════════════════════════════════════════════════════════════════
# SECTION 5 — NOTES
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, r, "Section 5 — Notes")

notes = [
    "1.  Champion = all existing channels (IM, EM, DM, etc). "
        "Challengers = champion + additional placement.",
    "2.  Target lift is RELATIVE (5% = 5% of baseline RR). "
        "Absolute target depends on baseline.",
    "3.  Baseline RR = mobile population-level rate "
        "(mobile responders / total population).",
    "4.  Pre-register comparisons 1 & 2 as primary; "
        "comparison 3 as exploratory.",
]

for i, note in enumerate(notes):
    row = r + 1 + i
    ws.merge_cells(
        start_row=row, start_column=2,
        end_row=row,   end_column=9
    )
    s(ws, row, 2, note, font=NOTE_FONT, fill=WHITE_FILL,
      align=LEFT_TOP, border=THIN_BORDER, height=18)
    s(ws, row, 1, "", fill=WHITE_FILL, border=THIN_BORDER)


# ── Freeze top row, zoom ──────────────────────────────────────────────────
ws.freeze_panes = "B2"
ws.sheet_view.zoomScale = 95

# ── Save ──────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"Saved: {OUTPUT_PATH}")


# ── Force recalculation via win32com (no formulas, but opens cleanly) ─────
def recalc_with_excel(filepath):
    try:
        import win32com.client
        abs_path = os.path.abspath(filepath)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb_com = excel.Workbooks.Open(abs_path)
        excel.Calculate()
        wb_com.Save()
        wb_com.Close()
        excel.Quit()
        print(f"Opened and saved via Excel: {abs_path}")
        return True
    except ImportError:
        print("win32com not available — skipping Excel open/save")
        return False
    except Exception as e:
        print(f"Excel open/save failed: {e}")
        try:
            excel.Quit()
        except Exception:
            pass
        return False

recalc_with_excel(OUTPUT_PATH)
