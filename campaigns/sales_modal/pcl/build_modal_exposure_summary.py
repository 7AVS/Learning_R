"""Build the modal-exposure-by-arm summary table (May cohort) as a clean Excel file."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule

# arm, clients, saw_modal  (exposure % is a formula = saw/clients)
rows = [
    ("Challenger (served the modal)", 138649, 963),
    ("Champion (held out)",           58867,  505),
    ("Other strategies",              375119, 2613),
]
total = ("Total (May)", sum(r[1] for r in rows), sum(r[2] for r in rows))

wb = Workbook()
ws = wb.active
ws.title = "Modal exposure"

navy  = "1F3864"
blue  = "005DAA"
light = "D9E1F2"
white = "FFFFFF"
thin  = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title
ws.merge_cells("A1:D1")
t = ws["A1"]
t.value = "PLI sales modal — who actually saw it (May)"
t.font = Font(bold=True, size=13, color=navy)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 24

# Header
headers = ["Arm", "Clients", "Saw the modal", "Exposure %"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=c, value=h)
    cell.font = Font(bold=True, color=white)
    cell.fill = PatternFill("solid", fgColor=blue)
    cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")
    cell.border = border
ws.row_dimensions[3].height = 20

# Data rows
r = 4
for arm, clients, saw in rows:
    ws.cell(row=r, column=1, value=arm).border = border
    ws.cell(row=r, column=2, value=clients).border = border
    ws.cell(row=r, column=3, value=saw).border = border
    e = ws.cell(row=r, column=4, value=f"=C{r}/B{r}")
    e.number_format = "0.00%"
    e.font = Font(bold=True)
    e.border = border
    r += 1

# Total row
ws.cell(row=r, column=1, value=total[0]).font = Font(bold=True)
ws.cell(row=r, column=2, value=total[1]).font = Font(bold=True)
ws.cell(row=r, column=3, value=total[2]).font = Font(bold=True)
tot_e = ws.cell(row=r, column=4, value=f"=C{r}/B{r}")
tot_e.number_format = "0.00%"
tot_e.font = Font(bold=True)
for c in range(1, 5):
    cell = ws.cell(row=r, column=c)
    cell.fill = PatternFill("solid", fgColor=light)
    cell.border = border
    if c > 1:
        cell.number_format = "#,##0" if c < 4 else "0.00%"

# Data bar on the three arm exposure cells — shows the pattern is flat (champion highest)
ws.conditional_formatting.add(
    "D4:D6",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=0.012,
                color=blue, showValue=True, minLength=None, maxLength=None),
)

# Thousands format on count columns
for row in range(4, r + 1):
    ws.cell(row=row, column=2).number_format = "#,##0"
    ws.cell(row=row, column=3).number_format = "#,##0"

# Takeaway line
ws.merge_cells(start_row=r + 2, start_column=1, end_row=r + 2, end_column=4)
tk = ws.cell(row=r + 2, column=1)
tk.value = ("The modal reached ~0.7% of clients — and the held-out arm saw it as much as "
            "the served arm. Exposure isn't being driven by the test setup.")
tk.font = Font(italic=True, size=11, color=navy)
tk.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r + 2].height = 40

ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 13
ws.sheet_view.showGridLines = False

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "modal_exposure_summary.xlsx")
wb.save(out)
print(f"Saved: {out}")
