# Builds campaigns/ASYNC/o2p_holdout_mde_calculator.xlsx
# O2P async BAU transition: shrink non-MB arm 47.5% -> 4.75% of cohort, keep TG7 5% untouched.
# All stats cells are LIVE formulas. Baseline and proven lift are drawn BY FORMULA from the
# "Tracker data (source)" sheet (mature cohorts), so the calculator follows the data.
# Scenario-grid holdout shares are editable inputs; every scenario carries a PASS/FAIL verdict
# (PASS = detectable lift <= proven lift). Legacy NORMSINV/NORMSDIST only (no _xlfn/@ breakage).
# Arm labels are mechanical (MB / NON_MB): the source tracker's champion/challenger tags are inverted.

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from scipy.stats import norm
import math

YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREY = PatternFill("solid", fgColor="EDEDED")
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=13)
THIN = Border(bottom=Side(style="thin", color="BBBBBB"))
GOODF = Font(color="006300", bold=True)
BADF = Font(color="C00000", bold=True)
TRK = "'Tracker data (source)'"

wb = Workbook()

# ---------------------------------------------------------------- sheet 1: calculator
ws = wb.active
ws.title = "MDE calculator"
ws.column_dimensions["A"].width = 46
for c in "BCDEFG":
    ws.column_dimensions[c].width = 16

ws["A1"] = "O2P async banner - 5% holdout MDE calculator"
ws["A1"].font = TITLE
ws["A2"] = ("Design: within TG4 (95% Action), shrink the no-banner arm from 47.5% to 4.75% of the "
            "cohort. TG7 (5% campaign control) untouched. Yellow cells are editable; everything "
            "else recalculates. Baseline and proven lift are formulas fed by the Tracker data "
            "sheet (mature cohorts Jun 18 + Jul 7) - type over them only to run what-ifs.")
ws["A2"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[2].height = 44

rows = [
    ("INPUTS", None, None),
    ("Cohort size, both TG4 arms (clients)", 505_000, "in"),                          # B5
    ("NON_MB holdout share of cohort", 0.0475, "in"),                                 # B6
    ("Baseline conversion, NON_MB arm (from tracker)",
     f"=AVERAGE({TRK}!E6,{TRK}!E8)", "in"),                                           # B7
    ("Proven lift from 50/50 read (from tracker)",
     f"=AVERAGE({TRK}!F5,{TRK}!F7)/100", "in"),                                       # B8
    ("Alpha (two-sided)", 0.05, "in"),                                                # B9
    ("Power target", 0.80, "in"),                                                     # B10
    ("DERIVED", None, None),
    ("NON_MB holdout n", "=ROUND(B5*B6,0)", None),                                    # B12
    ("MB (banner) n", "=ROUND(B5*(1-B6),0)", None),                                   # B13
    ("z alpha/2", "=NORMSINV(1-B9/2)", None),                                         # B14
    ("z power", "=NORMSINV(B10)", None),                                              # B15
    ("Standard error at baseline", "=SQRT(B7*(1-B7)*(1/B12+1/B13))", None),           # B16
    ("Minimum detectable lift (MDE)", "=(B14+B15)*B16", None),                        # B17
    ("MDE relative to baseline", "=B17/B7", None),                                    # B18
    ("Power to re-detect the proven lift", "=NORMSDIST(B8/B16-B14)", None),           # B19
    ("VERDICT at these inputs",
     '=IF(B17<=B8,"PASS - proven lift >= detectable minimum",'
     '"FAIL - holdout too small to confirm in one cycle")', None),                    # B20
    ("BUSINESS IMPACT OF THE CHANGE", None, None),
    ("Clients moved from no-banner to banner", "=ROUND(B5*(0.475-B6),0)", None),      # B22
    ("Extra target responders per cohort (at proven lift)", "=ROUND(B22*B8,0)", None),# B23
]
r = 3  # first row lands at 4: INPUTS r4, inputs B5-B10, DERIVED r11, values B12-B20
for label, val, kind in rows:
    r += 1
    ws.cell(r, 1, label)
    if val is None:
        ws.cell(r, 1).font = BOLD
        ws.cell(r, 1).fill = GREY
        ws.cell(r, 2).fill = GREY
        continue
    cell = ws.cell(r, 2, val)
    if kind == "in":
        cell.fill = YELLOW
    ws.cell(r, 1).border = THIN
    cell.border = THIN
for rr, fmt in [(6, "0.00%"), (7, "0.000%"), (8, "0.000%"), (9, "0.00"), (10, "0%"),
                (16, "0.0000%"), (17, "0.000%"), (18, "0.0%"), (19, "0.0%")]:
    ws.cell(rr, 2).number_format = fmt
for rr in (5, 12, 13, 22, 23):
    ws.cell(rr, 2).number_format = "#,##0"

# scenario grid: EDITABLE holdout shares x cohorts pooled -> MDE + PASS/FAIL
ws["A26"] = "SCENARIOS - type any holdout share in the yellow cells; MDE and verdicts recalculate"
ws["A26"].font = BOLD
hdr = ["Holdout share (editable)", "MDE, 1 cohort (pp)", "MDE, 2 cohorts (pp)", "MDE, 3 cohorts (pp)",
       "Verdict 1 cohort", "Verdict 2 cohorts", "Verdict 3 cohorts"]
for j, h in enumerate(hdr, 1):
    c = ws.cell(27, j, h)
    c.font = BOLD
    c.alignment = Alignment(wrap_text=True)
ws.row_dimensions[27].height = 30
for i, share in enumerate([0.025, 0.0475, 0.10, 0.20, 0.475]):
    rr = 28 + i
    a = ws.cell(rr, 1, share)
    a.fill = YELLOW
    a.number_format = "0.00%"
    for j in range(1, 4):  # B..D = MDE per pooled-cohort count
        ws.cell(rr, 1 + j,
                f"=($B$14+$B$15)*SQRT($B$7*(1-$B$7)*"
                f"(1/({j}*$B$5*$A{rr})+1/({j}*$B$5*(1-$A{rr}))))*100"
                ).number_format = "0.000"
    for j, col in enumerate("BCD"):  # E..G = verdicts
        ws.cell(rr, 5 + j, f'=IF({col}{rr}<=$B$8*100,"PASS","FAIL")')
ws["A34"] = ("PASS = detectable minimum <= the proven lift (B8, from tracker), i.e. that holdout "
             "still confirms the known effect in that window. pp = percentage points.")
ws["A34"].alignment = Alignment(wrap_text=True)

# conditional formatting: PASS green / FAIL red (verdict cells + main verdict)
for rng, anchor in [("E28:G32", "E28"), ("B20", "B20")]:
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("PASS",{anchor}))'], font=GOODF))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("FAIL",{anchor}))'], font=BADF))

# ---------------------------------------------------------------- sheet 2: tracker data
ws2 = wb.create_sheet("Tracker data (source)")
ws2.column_dimensions["A"].width = 20
for c in "BCDEF":
    ws2.column_dimensions[c].width = 18
ws2["A1"] = "Async_tracker_Aug pivot, transcribed 2026-09-04. Feeds the calculator's baseline and lift."
ws2["A1"].font = BOLD
ws2["A2"] = ("Source workbook tags the arms champion/challenger INVERTED - ignore those words. "
             "MB = banner served, NON_MB = no banner. TG7 not in this pivot. INTERIM: Jun 18 ran "
             "only an 18-day window (short); Jul 7 is the only full cohort; Aug 11 completes "
             "~Sep 11. Final numbers: rerun tracker after Sep 11, paste new counts in C/D - all "
             "formulas refresh.")
ws2["A2"].alignment = Alignment(wrap_text=True)
ws2.row_dimensions[2].height = 40
hdr2 = ["cohort", "arm", "total_population", "responders_target_cum", "conversion", "lift vs NON_MB (pp)"]
for j, h in enumerate(hdr2, 1):
    ws2.cell(4, j, h).font = BOLD
data = [
    ("6/18/2026 (18-day window)", "MB", 271731, 2269),
    ("6/18/2026 (18-day window)", "NON_MB", 271737, 1720),
    ("7/7/2026", "MB", 234863, 1918),
    ("7/7/2026", "NON_MB", 234866, 1464),
    ("8/11/2026 (immature)", "MB", 253657, 1232),
    ("8/11/2026 (immature)", "NON_MB", 253661, 856),
]
for i, (coh, arm, pop, resp) in enumerate(data):
    rr = 5 + i
    ws2.cell(rr, 1, coh)
    ws2.cell(rr, 2, arm)
    ws2.cell(rr, 3, pop).number_format = "#,##0"
    ws2.cell(rr, 4, resp).number_format = "#,##0"
    ws2.cell(rr, 5, f"=D{rr}/C{rr}").number_format = "0.000%"
    if arm == "MB":
        ws2.cell(rr, 6, f"=(D{rr}/C{rr}-D{rr+1}/C{rr+1})*100").number_format = "0.000"

out = r"C:\Users\andre\New_projects\cards\campaigns\ASYNC\o2p_holdout_mde_calculator.xlsx"
wb.save(out)

# ---- proof block: recompute what the live formulas should show
za, zp = norm.ppf(0.975), norm.ppf(0.80)
p = (1720 / 271737 + 1464 / 234866) / 2
lift = ((2269 / 271731 - 1720 / 271737) + (1918 / 234863 - 1464 / 234866)) / 2
N, s = 505_000, 0.0475
se = math.sqrt(p * (1 - p) * (1 / (N * s) + 1 / (N * (1 - s))))
mde = (za + zp) * se
print(f"baseline (tracker-fed) = {p:.4%}   proven lift = {lift*100:.3f}pp")
print(f"MDE @4.75% = {mde*100:.3f}pp   verdict: {'PASS' if mde <= lift else 'FAIL'}   "
      f"power = {norm.cdf(lift/se - za):.1%}")
print("grid (MDE pp / verdict @1 cohort):")
for sh in (0.025, 0.0475, 0.10, 0.20, 0.475):
    for k in (1, 2, 3):
        se_k = math.sqrt(p * (1 - p) * (1 / (k * N * sh) + 1 / (k * N * (1 - sh))))
        m = (za + zp) * se_k * 100
        print(f"  share {sh:.2%} x{k}: {m:.3f} {'PASS' if m <= lift*100 else 'FAIL'}", end="")
    print()
print("saved:", out)
