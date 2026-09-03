"""Build vbu_split_explainer.xlsx — plain-language ammunition for the 70/30 vs 50/50
conversation with marketing stakeholders. No stats jargon. Blocks sized to copy-paste
straight into PowerPoint. Numbers from b3/b4 + workstation pivot (Jun/Jul mature waves).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BOLD = Font(bold=True)
BIG = Font(bold=True, size=14)
HDR = PatternFill("solid", fgColor="D9E1F2")
HILITE = PatternFill("solid", fgColor="E2EFDA")
WRAP = Alignment(wrap_text=True, vertical="top")
BOX = Border(*[Side(style="thin")] * 4)

wb = Workbook()
ws = wb.active
ws.title = "explainer"
ws.sheet_view.showGridLines = False

r = 1
def title(text):
    global r
    ws.cell(row=r, column=1, value=text).font = BIG
    r += 1

def block_table(headers, rows, widths=None, highlight_col=None):
    global r
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=j, value=h)
        c.font = BOLD
        c.fill = HDR
        c.border = BOX
        c.alignment = WRAP
    r += 1
    for row in rows:
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BOX
            c.alignment = WRAP
            if highlight_col and j == highlight_col:
                c.fill = HILITE
        r += 1
    r += 1

def note(text):
    global r
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(italic=True)
    c.alignment = WRAP
    r += 2

# ---------------------------------------------------------------- Block 1
title("1. What we have already seen (June + July campaigns)")
block_table(
    ["What happened", "The numbers", "In plain words"],
    [
        ["Clients who RECEIVED the offer", "about 2 in every 100 upgraded "
         "(1.8 per 100 on the no-rebate offer, 2.4 per 100 on the $55-rebate offer)",
         "The campaign gets upgrades."],
        ["Clients who did NOT receive the offer", "0 upgrades out of about 3,000 clients",
         "Nobody upgrades on their own. No offer = no upgrade."],
        ["Where the upgrades come from", "9 out of 10 upgrades come from the model's top 4 "
         "score groups (group 1 alone: 3 to 8 per 100)",
         "The model's ranking works — the top of the list responds, the bottom barely does."],
    ],
)

# ---------------------------------------------------------------- Block 2
title("2. What the test does")
note("We let a coin flip decide which clients get the offer. The held-out clients tell us "
     "what would have happened WITHOUT the campaign. Because we saw 0 upgrades without the "
     "offer, we expect the held-out group to sit at ~0 — and then every upgrade in the "
     "mailed group is proven to be caused by the campaign. The only decision is HOW MANY "
     "clients to hold out.")

# ---------------------------------------------------------------- Block 3
title("3. The two scenarios, per monthly wave (~17,000–27,000 model-selected clients)")
block_table(
    ["", "Hold out half (50/50)", "Hold out 3 in 10 (70/30) — RECOMMENDED"],
    [
        ["Clients still receiving the offer", "~8,500–13,500", "~12,000–18,900"],
        ["Clients held out", "~8,500–13,500", "~5,100–8,100"],
        ["Upgrades we give up, per wave", "~170–270", "~100–160"],
        ["Answers 'does the campaign cause the upgrades?'", "Yes — after wave 1",
         "Yes — after wave 1"],
        ["Reads each of the top 4 score groups (9 of 10 upgrades)", "Yes", "Yes"],
        ["Reads each of the bottom score groups (5–9)",
         "No — too few clients there, even after 2 waves",
         "No — same (see block 4)"],
        ["Waves needed", "2 (second wave confirms it wasn't luck)", "2 (same)"],
    ],
    highlight_col=3,
)

# ---------------------------------------------------------------- Block 4
title("4. Why 70/30 is enough — the one-paragraph version")
note("The only thing a bigger holdout could buy is a verdict on the bottom score groups "
     "(5 to 9). But the model sends very few people there (a few hundred per group) and "
     "almost none of them upgrade (less than 1 in 100). Groups that small and that quiet "
     "don't produce enough upgrades to measure — not at 70/30, and not at 50/50 either, "
     "even after two waves. So going 50/50 costs about 110 extra lost upgrades every wave "
     "and buys nothing we can use. We will still report the bottom groups — as one combined "
     "bucket with a ceiling ('at most X per 100'), which is all any split can honestly say.")

# ---------------------------------------------------------------- Block 5
title("5. Why two waves")
note("Wave 1 gives the answer. Wave 2 repeats the test on the next month's clients: if the "
     "same result shows up twice, nobody can call it a one-month fluke — and it doubles the "
     "data behind the score-group read at no extra design effort.")

# ---------------------------------------------------------------- Block 6
title("6. The full dial (backup — if someone asks 'why not smaller / bigger')")
block_table(
    ["Holdout size", "Upgrades given up per wave", "What we can read",
     "Verdict"],
    [
        ["1 in 10 (90/10)", "~34–53", "Campaign yes/no + top 3 score groups",
         "Cheapest, but loses score group 4"],
        ["2 in 10 (80/20)", "~68–107", "Campaign yes/no + top 4 score groups",
         "Workable minimum"],
        ["3 in 10 (70/30) — RECOMMENDED", "~102–160",
         "Campaign yes/no + top 4 score groups, solidly, twice",
         "Best answer-per-upgrade-lost"],
        ["Half (50/50)", "~170–270",
         "Same as 70/30 — bottom groups still unreadable",
         "Pays ~70% more for the same answers"],
    ],
    highlight_col=1,
)

note("Sources: June/July 2026 waves, model-based offers (AIB_25K_NR, AIB_25K_R_55). "
     "Communicated: 1.79% / 2.40% conversion. Not communicated: 0 of 2,956. "
     "Full design: vbu_propensity_doe_report.md.")

for col, w in (("A", 46), ("B", 44), ("C", 46), ("D", 30)):
    ws.column_dimensions[col].width = w

path = (r"C:\Users\andre\New_projects\cards\campaigns\VBA_VBU\propensity_experiment"
        r"\vbu_split_explainer.xlsx")
wb.save(path)
print(f"saved {path}")
