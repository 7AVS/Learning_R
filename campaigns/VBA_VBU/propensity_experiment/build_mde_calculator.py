"""Build vbu_mde_calculator.xlsx — MDE for both VBU propensity-experiment routes.

Route A: champion/challenger (offer-level randomization, proven machinery).
Route B: score-band audit (band-level randomization, CIDM gate unconfirmed).
Baselines: b3 mature waves (NR 1.79%, R_55 2.40%); deciles: b4 v2 (certified).
All parameters live Excel formulas — edit yellow cells.
"""
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREY = PatternFill("solid", fgColor="E7E6E6")
BOLD = Font(bold=True)
PCT2 = "0.00%"
NUM = "#,##0"
MONEY = "$#,##0"

wb = Workbook()

# ---------------------------------------------------------------- README
rd = wb.active
rd.title = "README"
readme = [
    ("VBU propensity-model experiment — MDE calculator", BOLD),
    ("Built 2026-09-02. Yellow cells are editable; everything downstream recomputes.", None),
    ("", None),
    ("What the experiment tests: the propensity model's TARGETING value — does model-based", None),
    ("selection beat contacting clients the model wouldn't pick. Offer-vs-offer (rebate value)", None),
    ("is explicitly out of scope.", None),
    ("", None),
    ("Route A — champion/challenger (sheet 2): one randomization at the top. Champion arm =", None),
    ("model-selected clients, contacted. Challenger arm = random slice of contacted volume drawn", None),
    ("from below-cutoff / non-selected pool. Two-proportion two-sided test. Needs only", None),
    ("offer-level split machinery, which CIDM already runs (MCB_35K_NR 50/50 precedent).", None),
    ("", None),
    ("Route B — score-band audit (sheet 3): random below-cutoff slice sized per band to bound", None),
    ("each band's rate (rule of three when 0 conversions: upper 95% bound = 3/n). Gives", None),
    ("calibration-by-band. GATE: needs band-level (within-decile) randomization in the", None),
    ("decisioning tree — UNCONFIRMED, top CIDM question. Do not commit to Route B before that", None),
    ("answer.", None),
    ("", None),
    ("Sources:", BOLD),
    ("  Baselines: b3_baseline_conversion.sql — mature waves Jun-13/Jul-10 (Aug-14 immature,", None),
    ("  excluded). NR 1.79% both waves; R_55 2.40-2.41% both waves. COMM arms only.", None),
    ("  Deciles: b4_discrimination_read.sql v2 (2-char decile fix, certified). Decile 1 = best.", None),
    ("  Wave sizes: d1_deployment_cycle.sql — monthly deploys, ~2.5-month response windows.", None),
    ("", None),
    ("Caveats that shape the design:", BOLD),
    ("  1. Controls convert ~0 (0 target-product in 2,956+, ~0 any-product) — communication", None),
    ("     drives effectively all upgrades. Per-band NC holdouts therefore add little; the", None),
    ("     informative contrast is contacted-above-cutoff vs contacted-below-cutoff.", None),
    ("  2. Response windows overlap across monthly waves; re-entry is rare (none 3+ waves).", None),
    ("     Pooling waves is acceptable for planning; measure per-cohort.", None),
    ("  3. MDE formula is the standard normal-approximation planning formula using the champion", None),
    ("     baseline for variance. Fine at these n; exact power run before final lock.", None),
    ("  4. Cost per contact is UNKNOWN — audit fee formula is live but needs that number.", None),
]
for i, (txt, f) in enumerate(readme, start=1):
    c = rd.cell(row=i, column=1, value=txt)
    if f:
        c.font = f
rd.column_dimensions["A"].width = 100

# ------------------------------------------------- Route A: champion/challenger
ra = wb.create_sheet("Route A champion-challenger")
ra["A1"] = "Route A — champion/challenger MDE (two-proportion, two-sided)"
ra["A1"].font = BOLD
ra["A2"] = ("Cell = minimum detectable difference in conversion rate (percentage points) between "
            "champion (model-selected) and challenger (below-cutoff) arms.")

ra["A4"], ra["B4"] = "alpha (two-sided)", 0.05
ra["A5"], ra["B5"] = "power", 0.80
ra["A6"], ra["B6"] = "z alpha/2", "=NORM.S.INV(1-B4/2)"
ra["A7"], ra["B7"] = "z power", "=NORM.S.INV(B5)"
for r in (4, 5):
    ra.cell(row=r, column=2).fill = YELLOW

def route_a_block(ws, top, offer, p1, scenarios):
    ws.cell(row=top, column=1, value=f"{offer} — baseline COMM conversion (b3 mature waves)").font = BOLD
    b = ws.cell(row=top, column=2, value=p1)
    b.fill = YELLOW
    b.number_format = PCT2
    hdr = top + 1
    ws.cell(row=hdr, column=1, value="wave scenario").font = BOLD
    ws.cell(row=hdr, column=2, value="N total").font = BOLD
    for j, share in enumerate((0.10, 0.20, 0.30, 0.50)):
        c = ws.cell(row=hdr, column=3 + j, value=share)
        c.fill = YELLOW
        c.number_format = "0%"
        c.font = BOLD
    for i, (label, n) in enumerate(scenarios):
        r = hdr + 1 + i
        ws.cell(row=r, column=1, value=label)
        nc = ws.cell(row=r, column=2, value=n)
        nc.number_format = NUM
        for j in range(4):
            col = chr(ord("C") + j)
            f = (f"=($B$6+$B$7)*SQRT($B${top}*(1-$B${top})"
                 f"*(1/($B{r}*(1-{col}${hdr}))+1/($B{r}*{col}${hdr})))")
            fc = ws.cell(row=r, column=3 + j, value=f)
            fc.number_format = PCT2
    return hdr + 1 + len(scenarios)

nr_scen = [("1 wave (Jul-10)", 8540), ("1 wave (Jun-13)", 13408),
           ("2 waves (Jun+Jul)", 21948), ("3 waves (Jun+Jul+Aug)", 29012)]
r55_scen = [("1 wave (Jul-10)", 7752), ("1 wave (Jun-13)", 12162),
            ("2 waves (Jun+Jul)", 19914), ("3 waves (Jun+Jul+Aug)", 28219)]
nxt = route_a_block(ra, 9, "AIB_25K_NR", 0.0179, nr_scen)
nxt = route_a_block(ra, nxt + 1, "AIB_25K_R_55", 0.0240, r55_scen)

ra.cell(row=nxt + 1, column=1,
        value=("Read: header row = challenger share of the wave. If below-cutoff truly converts at "
               "low-decile rates (b4: ~0.1-0.8%), the true gap is ~1.0-2.2pp — a 10-20% challenger "
               "slice in a single wave already detects it. Bigger slices buy precision on the gap, "
               "not detection.")).font = Font(italic=True)
ra.column_dimensions["A"].width = 42
for col in "BCDEF":
    ra.column_dimensions[col].width = 13

# ------------------------------------------------------ Route B: score-band
rb = wb.create_sheet("Route B score-band")
rb["A1"] = "Route B — score-band audit: below-cutoff slice sizing + audit fee"
rb["A1"].font = BOLD
rb["A2"] = ("GATE: needs band-level randomization in CIDM's decisioning tree — UNCONFIRMED "
            "(top CIDM question). Offer-level precedent (MCB 50/50) does NOT prove this.")
rb["A2"].font = Font(bold=True, color="9C0006")

rb["A4"], rb["B4"] = "certify-dead threshold (upper 95% bound per band)", 0.005
rb["B4"].fill = YELLOW
rb["B4"].number_format = PCT2
rb["A5"], rb["B5"] = "rule-of-three n per band (0 conversions)", "=ROUNDUP(3/B4,0)"
rb["B5"].number_format = NUM
rb["A6"], rb["B6"] = "cost per contact ($) — UNKNOWN, get the number", 0
rb["B6"].fill = YELLOW
rb["A7"], rb["B7"] = "rebate per conversion — R_55 ($)", 55
rb["B7"].fill = YELLOW
rb["A8"], rb["B8"] = "rebate per conversion — NR ($)", 0
rb["B8"].fill = YELLOW

hdr = 10
heads = ["decile (1=best)", "NR n", "NR conv", "NR rate", "R_55 n", "R_55 conv", "R_55 rate",
         "planned slice n (per offer)", "exp conv NR", "exp conv R_55", "upper 95% if 0 conv"]
for j, h in enumerate(heads, start=1):
    c = rb.cell(row=hdr, column=j, value=h)
    c.font = BOLD
    c.fill = GREY
    c.alignment = Alignment(wrap_text=True)

# b4 v2 certified decile table (COMM arms). d10 ~0 conversions recorded as 0.
deciles = [
    (1, 4936, 136, 4710, 311), (2, 5740, 117, 4381, 107), (3, 5889, 81, 5162, 82),
    (4, 6094, 62, 5995, 53), (5, 908, 5, 1148, 10), (6, 1379, 11, 1786, 13),
    (7, 1626, 9, 2132, 13), (8, 2213, 3, 2618, 5), (9, 173, 0, 198, 3), (10, 54, 0, 89, 0),
]
for i, (d, n_nr, c_nr, n_r, c_r) in enumerate(deciles):
    r = hdr + 1 + i
    rb.cell(row=r, column=1, value=d)
    rb.cell(row=r, column=2, value=n_nr).number_format = NUM
    rb.cell(row=r, column=3, value=c_nr)
    rb.cell(row=r, column=4, value=f"=C{r}/B{r}").number_format = PCT2
    rb.cell(row=r, column=5, value=n_r).number_format = NUM
    rb.cell(row=r, column=6, value=c_r)
    rb.cell(row=r, column=7, value=f"=F{r}/E{r}").number_format = PCT2
    slice_default = 0 if d <= 4 else "=$B$5"
    sc = rb.cell(row=r, column=8, value=slice_default)
    sc.fill = YELLOW
    rb.cell(row=r, column=9, value=f"=H{r}*D{r}").number_format = "0.0"
    rb.cell(row=r, column=10, value=f"=H{r}*G{r}").number_format = "0.0"
    rb.cell(row=r, column=11, value=f'=IF(H{r}>0,3/H{r},"")').number_format = PCT2

tot = hdr + 11
rb.cell(row=tot, column=1, value="TOTAL").font = BOLD
rb.cell(row=tot, column=8, value=f"=SUM(H{hdr+1}:H{hdr+10})").font = BOLD
rb.cell(row=tot, column=8).number_format = NUM
rb.cell(row=tot, column=9, value=f"=SUM(I{hdr+1}:I{hdr+10})").number_format = "0.0"
rb.cell(row=tot, column=10, value=f"=SUM(J{hdr+1}:J{hdr+10})").number_format = "0.0"

fee = tot + 2
rb.cell(row=fee, column=1, value="AUDIT FEE (the number for the design doc)").font = BOLD
rb.cell(row=fee + 1, column=1, value="NR: slice contacts x cost + exp conv x rebate")
rb.cell(row=fee + 1, column=2, value=f"=H{tot}*B6+I{tot}*B8").number_format = MONEY
rb.cell(row=fee + 2, column=1, value="R_55: slice contacts x cost + exp conv x rebate")
rb.cell(row=fee + 2, column=2, value=f"=H{tot}*B6+J{tot}*B7").number_format = MONEY
rb.cell(row=fee + 3, column=1, value="TOTAL (both offers)").font = BOLD
rb.cell(row=fee + 3, column=2, value=f"=B{fee+1}+B{fee+2}").number_format = MONEY
rb.cell(row=fee + 3, column=2).font = BOLD
rb.cell(row=fee + 5, column=1,
        value=("Defaults: deciles 1-4 slice = 0 (already contacted at scale — calibration is free "
               "from BAU volume); deciles 5-10 slice = rule-of-three n. Note d5-10 already get "
               "thin BAU contact (~1-2.6K each) — if CIDM confirms those inclusions are "
               "quasi-random, part of the audit fee is pre-paid.")).font = Font(italic=True)
rb.column_dimensions["A"].width = 46
for col in "BCDEFGHIJK":
    rb.column_dimensions[col].width = 12

path = r"C:\Users\andre\New_projects\cards\campaigns\VBA_VBU\propensity_experiment\vbu_mde_calculator.xlsx"
wb.save(path)
print(f"saved {path}")

# ---- verification: recompute Route A grid in python (mirrors the Excel formulas)
za, zp = 1.959963985, 0.8416212336
print("\nRoute A MDE verification (pp), shares 10/20/30/50%:")
for offer, p1, scen in (("NR", 0.0179, nr_scen), ("R_55", 0.0240, r55_scen)):
    for label, n in scen:
        row = []
        for s in (0.10, 0.20, 0.30, 0.50):
            mde = (za + zp) * math.sqrt(p1 * (1 - p1) * (1 / (n * (1 - s)) + 1 / (n * s)))
            row.append(f"{mde*100:.2f}")
        print(f"  {offer:5s} {label:22s} N={n:6,d}  " + "  ".join(row))
print("\nRoute B rule-of-three at 0.5% threshold: n =", math.ceil(3 / 0.005), "per band")
