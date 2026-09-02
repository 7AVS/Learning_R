"""Build vbu_deployment_results_by_offer.xlsx — previous VBU deployment results:
volumes + baseline conversion by wave x offer x arm. MDE input workbook for Andre's
workstation. All numbers from probes run 2026-09-02 (b3 / d1), banked in
experiment_config_2026-09.md. Rates are live formulas (conv/clients).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BOLD = Font(bold=True)
GREY = PatternFill("solid", fgColor="E7E6E6")
PCT2 = "0.00%"
NUM = "#,##0"

wb = Workbook()

# ---------------- Sheet 1: baseline by wave x offer x arm (the MDE input) ----
s1 = wb.active
s1.title = "baseline by wave x offer"
s1["A1"] = ("VBU deployments — clients and target-product conversion by wave x offer x arm. "
            "Source: b3_baseline_conversion.sql run 2026-09-02. Response window ~2.5 months "
            "ending month-end; conversions computed for AIB_25K_NR and AIB_25K_R_55 only.")
s1["A1"].font = BOLD
s1["A1"].alignment = Alignment(wrap_text=True)
s1.row_dimensions[1].height = 45

heads = ["deploy date (wave)", "offer", "arm", "clients", "target conversions",
         "conversion rate", "maturity at read (2026-09-02)"]
for j, h in enumerate(heads, start=1):
    c = s1.cell(row=3, column=j, value=h)
    c.font = BOLD
    c.fill = GREY

rows = [
    ("2026-06-13", "AIB_25K_NR",   "COMM",     13408, 240, "mature (~80d window)"),
    ("2026-06-13", "AIB_25K_NR",   "NOT_COMM",   798,   0, "mature (~80d window)"),
    ("2026-06-13", "AIB_25K_R_55", "COMM",     12162, 293, "mature (~80d window)"),
    ("2026-06-13", "AIB_25K_R_55", "NOT_COMM",   619,   0, "mature (~80d window)"),
    ("2026-07-10", "AIB_25K_NR",   "COMM",      8540, 153, "mature (~50d)"),
    ("2026-07-10", "AIB_25K_NR",   "NOT_COMM",   394,   0, "mature (~50d)"),
    ("2026-07-10", "AIB_25K_R_55", "COMM",      7752, 186, "mature (~50d)"),
    ("2026-07-10", "AIB_25K_R_55", "NOT_COMM",   335,   0, "mature (~50d); 1 any-product conv"),
    ("2026-08-14", "AIB_25K_NR",   "COMM",      7064,  31, "IMMATURE (~2.5 wk) — not a baseline"),
    ("2026-08-14", "AIB_25K_NR",   "NOT_COMM",   380,   0, "IMMATURE"),
    ("2026-08-14", "AIB_25K_R_55", "COMM",      8305, 119, "IMMATURE (~2.5 wk) — not a baseline"),
    ("2026-08-14", "AIB_25K_R_55", "NOT_COMM",   430,   0, "IMMATURE"),
]
for i, (dt, offer, arm, n, conv, note) in enumerate(rows):
    r = 4 + i
    s1.cell(row=r, column=1, value=dt)
    s1.cell(row=r, column=2, value=offer)
    s1.cell(row=r, column=3, value=arm)
    s1.cell(row=r, column=4, value=n).number_format = NUM
    s1.cell(row=r, column=5, value=conv)
    s1.cell(row=r, column=6, value=f"=E{r}/D{r}").number_format = PCT2
    s1.cell(row=r, column=7, value=note)

s1.cell(row=17, column=1, value=("Baselines for planning (mature waves, stable across Jun/Jul): "
                                 "AIB_25K_NR 1.79% | AIB_25K_R_55 2.40-2.41%. NOT_COMM converts "
                                 "~0 (0 target-product across all waves) — communication drives "
                                 "effectively all upgrades.")).font = Font(italic=True)
s1.column_dimensions["A"].width = 18
s1.column_dimensions["B"].width = 16
s1.column_dimensions["C"].width = 12
s1.column_dimensions["D"].width = 10
s1.column_dimensions["E"].width = 16
s1.column_dimensions["F"].width = 14
s1.column_dimensions["G"].width = 38

# ---------------- Sheet 2: Jul-10 wave, all offers (volumes only) ------------
s2 = wb.create_sheet("Jul-10 all offers (volumes)")
s2["A1"] = ("2026-07-10 wave — ALL offers, volumes by arm. Source: d1_deployment_cycle.sql "
            "STMT 3 run 2026-09-02. Conversions not yet computed for offers outside "
            "AIB_25K_NR / AIB_25K_R_55. Totals reconcile to dashboard: 28,878 COMM ~= 28,870; "
            "11,007 NOT_COMM exact.")
s2["A1"].font = BOLD
s2["A1"].alignment = Alignment(wrap_text=True)
s2.row_dimensions[1].height = 45

for j, h in enumerate(["offer", "COMM clients", "NOT_COMM clients", "NOT_COMM share"], start=1):
    c = s2.cell(row=3, column=j, value=h)
    c.font = BOLD
    c.fill = GREY

jul_offers = [
    ("AIB_25K_NR", 8540, 394), ("AIB_25K_R_55", 7752, 335), ("AIB_25K_R_175_X", 1657, 83),
    ("AIB_35K_R_55", 706, 31), ("AIB_35K_R_175_X", 51, 2), ("MCB_35K_NR", 10164, 10162),
    ("(blank)", 8, 0),
]
for i, (offer, comm, nc) in enumerate(jul_offers):
    r = 4 + i
    s2.cell(row=r, column=1, value=offer)
    s2.cell(row=r, column=2, value=comm).number_format = NUM
    s2.cell(row=r, column=3, value=nc).number_format = NUM
    s2.cell(row=r, column=4, value=f"=C{r}/(B{r}+C{r})").number_format = PCT2
tr = 4 + len(jul_offers)
s2.cell(row=tr, column=1, value="TOTAL").font = BOLD
s2.cell(row=tr, column=2, value=f"=SUM(B4:B{tr-1})").number_format = NUM
s2.cell(row=tr, column=3, value=f"=SUM(C4:C{tr-1})").number_format = NUM
s2.cell(row=tr + 2, column=1,
        value=("MCB_35K_NR runs a deliberate ~50/50 holdout — this is what inflates the dashboard "
               "pooled control to ~28%. AIB offers hold ~4-5% NOT_COMM.")).font = Font(italic=True)
s2.column_dimensions["A"].width = 18
for col in "BCD":
    s2.column_dimensions[col].width = 16

# ---------------- Sheet 3: deployment calendar (all 5 waves) -----------------
s3 = wb.create_sheet("deployment calendar")
s3["A1"] = ("All VBU deploys 2026 — monthly cadence, tactics consolidating 5 to 1. Source: "
            "d1_deployment_cycle.sql STMT 1. Per-offer conversion results exist only for "
            "Jun/Jul/Aug at AIB_25K_NR / AIB_25K_R_55 grain (sheet 1); Apr/May per-offer "
            "results = not yet queried.")
s3["A1"].font = BOLD
s3["A1"].alignment = Alignment(wrap_text=True)
s3.row_dimensions[1].height = 45

for j, h in enumerate(["deploy date", "tactics", "total clients", "per-offer results available?"],
                      start=1):
    c = s3.cell(row=3, column=j, value=h)
    c.font = BOLD
    c.fill = GREY
calendar = [
    ("2026-04-13", 5, 39954, "no — not queried"),
    ("2026-05-13", 4, 34547, "no — not queried"),
    ("2026-06-13", 3, 52092, "yes (sheet 1, NR/R_55)"),
    ("2026-07-10", 2, 39877, "yes (sheets 1-2)"),
    ("2026-08-14", 1, 39256, "immature (sheet 1)"),
]
for i, (dt, tac, n, avail) in enumerate(calendar):
    r = 4 + i
    s3.cell(row=r, column=1, value=dt)
    s3.cell(row=r, column=2, value=tac)
    s3.cell(row=r, column=3, value=n).number_format = NUM
    s3.cell(row=r, column=4, value=avail)
s3.cell(row=10, column=1, value=("Re-entry across waves is rare: 9,270/196,456 in exactly 2 "
                                 "waves, none in 3+ — waves are clean monthly cohorts.")).font = Font(italic=True)
s3.column_dimensions["A"].width = 14
s3.column_dimensions["B"].width = 10
s3.column_dimensions["C"].width = 14
s3.column_dimensions["D"].width = 28

path = (r"C:\Users\andre\New_projects\cards\campaigns\VBA_VBU\propensity_experiment"
        r"\vbu_deployment_results_by_offer.xlsx")
wb.save(path)
print(f"saved {path}")
