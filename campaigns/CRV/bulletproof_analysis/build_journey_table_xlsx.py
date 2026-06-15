try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl")
    raise

import os

# --- Data dict (edit here; percentages computed below) ---
DATA = {
    "ANDROID": {
        "arm_total_viewers": 175158,
        "rows": [
            {
                "path": "A  CRV-action -> Both",
                "viewers": 102296,
                "crv_clicks": 21430,
                "pcl_clicks": 45199,
                "show_view_mix": True,
            },
            {
                "path": "B  CRV-control -> PCL only",
                "viewers": 9639,
                "crv_clicks": None,
                "pcl_clicks": 3401,
                "show_view_mix": False,
            },
            {
                "path": "C  No-overlap -> PCL only",
                "viewers": 61998,
                "crv_clicks": None,
                "pcl_clicks": 28114,
                "show_view_mix": False,
            },
        ],
    },
    "iOS": {
        "arm_total_viewers": 468183,
        "rows": [
            {
                "path": "A  CRV-action -> Both",
                "viewers": 390972,
                "crv_clicks": 67177,
                "pcl_clicks": 130326,
                "show_view_mix": True,
            },
            {
                "path": "B  CRV-control -> PCL only",
                "viewers": 23962,
                "crv_clicks": None,
                "pcl_clicks": 9764,
                "show_view_mix": False,
            },
            {
                "path": "C  No-overlap -> PCL only",
                "viewers": 152640,
                "crv_clicks": None,
                "pcl_clicks": 72289,
                "show_view_mix": False,
            },
        ],
    },
}

# --- % compute ---
def pct(num, denom):
    if num is None or denom == 0:
        return None
    return round(num / denom)   # whole number; stored as fraction for 0% format

def pct_val(num, denom):
    if num is None or denom == 0:
        return None
    return num / denom           # real fraction for number_format '0%'


# --- Styles ---
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
ACTION_FILL  = PatternFill("solid", fgColor="D9E1F2")   # subtle blue tint for Path A
PCL_FILL     = PatternFill("solid", fgColor="FFF2CC")   # amber tint for PCL CTR col
PCL_HEAD_FILL= PatternFill("solid", fgColor="F4B942")

WHITE_FONT   = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT    = Font(bold=True, size=10)
STD_FONT     = Font(size=10)
ITALIC_FONT  = Font(italic=True, size=9, color="595959")

THIN = Side(style="thin", color="AAAAAA")
THIN_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCL_BORDER   = Border(left=Side(style="medium", color="C9A000"),
                      right=Side(style="medium", color="C9A000"),
                      top=THIN, bottom=THIN)


def apply_border(cell, border):
    cell.border = border

def write_cell(ws, row, col, value, font=None, fill=None, alignment=None,
               number_format=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:        c.font = font
    if fill:        c.fill = fill
    if alignment:   c.alignment = alignment
    if number_format: c.number_format = number_format
    if border:      c.border = border
    return c


COL_NAMES = [
    "Path",
    "View bucket",
    "Viewers (n)",
    "View mix %",
    "Clicked CRV (n)",
    "CRV CTR %",
    "Clicked PCL (n)",
    "PCL CTR %",       # col 8 — headline
]
N_COLS = len(COL_NAMES)
PCL_CTR_COL = 8


def build(wb, platforms):
    ws = wb.active
    ws.title = "Journey Table"
    ws.freeze_panes = "A3"  # freeze after row 2 (column headers)

    row = 1

    # Row 1: title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
    c = ws.cell(row=row, column=1,
                value="CRV × PCL Banner — View→Click Journey (Android vs iOS)")
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal="center")
    row += 1

    # Column headers (row 2)
    for ci, name in enumerate(COL_NAMES, 1):
        fill = PCL_HEAD_FILL if ci == PCL_CTR_COL else HEADER_FILL
        write_cell(ws, row, ci, name,
                   font=WHITE_FONT,
                   fill=fill,
                   alignment=Alignment(horizontal="center", wrap_text=True),
                   border=THIN_BORDER)
    row += 1

    for platform_name, pdata in platforms.items():
        arm_total = pdata["arm_total_viewers"]

        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
        c = ws.cell(row=row, column=1, value=platform_name)
        c.font = WHITE_FONT
        c.fill = SECTION_FILL
        c.alignment = Alignment(horizontal="left")
        row += 1

        for r in pdata["rows"]:
            v   = r["viewers"]
            crv = r["crv_clicks"]
            pcl = r["pcl_clicks"]
            vm  = pct_val(v, arm_total) if r["show_view_mix"] else None
            crv_ctr = pct_val(crv, v)
            pcl_ctr = pct_val(pcl, v)

            is_action = r["show_view_mix"]  # Path A
            row_fill = ACTION_FILL if is_action else None

            def cell_font(bold=False):
                return Font(bold=bold, size=10)

            def wc(col, value, nf=None, extra_fill=None, extra_border=None):
                f = row_fill if extra_fill is None else extra_fill
                b = THIN_BORDER if extra_border is None else extra_border
                write_cell(ws, row, col, value,
                           font=cell_font(bold=is_action),
                           fill=f,
                           alignment=Alignment(horizontal="center"),
                           number_format=nf,
                           border=b)

            # Path (left-aligned)
            c = write_cell(ws, row, 1, r["path"],
                           font=cell_font(bold=is_action),
                           fill=row_fill,
                           alignment=Alignment(horizontal="left"),
                           border=THIN_BORDER)

            # View bucket (same as path description, abbreviated)
            bucket_label = (
                "Action — CRV + PCL" if is_action
                else ("Control — PCL" if "control" in r["path"].lower() else "No-overlap — PCL")
            )
            write_cell(ws, row, 2, bucket_label,
                       font=cell_font(), fill=row_fill,
                       alignment=Alignment(horizontal="left"),
                       border=THIN_BORDER)

            wc(3, v,       nf="#,##0")
            wc(4, vm,      nf="0%")
            wc(5, crv,     nf="#,##0")
            wc(6, crv_ctr, nf="0%")
            wc(7, pcl,     nf="#,##0")
            # PCL CTR — headline col
            wc(8, pcl_ctr, nf="0%",
               extra_fill=PCL_FILL,
               extra_border=PCL_BORDER)

            # Validation print
            crv_str = f"{round(crv_ctr*100)}%" if crv_ctr is not None else "n/a"
            pcl_str = f"{round(pcl_ctr*100)}%" if pcl_ctr is not None else "n/a"
            print(f"  [{platform_name}] {r['path'].strip()}: "
                  f"n={v:,}, CRV CTR={crv_str}, PCL CTR={pcl_str}")

            row += 1

        # Blank row between sections
        row += 1

    # --- Footnotes ---
    footnotes = [
        "View mix % = share of arm total viewers (Action arm only). "
        "CTR % = clicked / that bucket's viewers.",

        "Descriptive, not causal: Action·Both conditions on viewing both (post-exposure); "
        "no-overlap is a different population (context only). "
        "Causal cannibalization = arm-level action vs control.",

        "Figures approximate from query OCR — replace with exact query values. "
        "Flagged uncertain: (none critical in this view).",
    ]
    for fn in footnotes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
        c = ws.cell(row=row, column=1, value=fn)
        c.font = ITALIC_FONT
        c.alignment = Alignment(wrap_text=True, horizontal="left")
        ws.row_dimensions[row].height = 28
        row += 1

    # Column widths
    widths = [32, 24, 14, 12, 16, 12, 16, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Header row height
    ws.row_dimensions[2].height = 32


def main():
    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "journey_table.xlsx")

    wb = openpyxl.Workbook()
    print("Building journey table…")
    build(wb, DATA)
    wb.save(out_path)
    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()
