"""
Build CRV Suppression Coverage Curve Excel.

Input: cells_v3_pooled.csv — 100 policy-matrix cells (elig_txn_bin x
mobile_login_bin x prior_contact_bin), pooled Sep-2025-Mar-2026 cohorts.
Action = served mobile banner, Control = 5% holdout.

Output: coverage_curve.xlsx, 3 sheets:
  Cells   — one row per cell, sorted by lift ascending, live formulas.
  Curve   — same order, cumulative surface/converter columns + chart.
  Summary — pooled totals + a 10/20/30/40/50% surface-cut scenario table,
            interpolated off the Curve sheet.

All rate/lift/cumulative math is Excel formulas referencing raw counts, so
editing a leads/conv count recalculates everything downstream. NOTE: the
row ORDER (sorted by lift) and the scenario table's bracket-row references
are fixed at build time in Python — if edited counts are large enough to
reorder cells by lift, rerun this script to re-sort and re-bracket.
"""

import csv
import os

import openpyxl
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CSV_PATH = os.path.join(SCRIPT_DIR, "cells_v3_pooled.csv")
OUT_PATH = os.path.join(SCRIPT_DIR, "coverage_curve.xlsx")

N_CELLS = 100
FIRST_ROW = 2
LAST_ROW = FIRST_ROW + N_CELLS - 1  # 101

# ── Styles ────────────────────────────────────────────────────────────────
NAVY_FILL = PatternFill("solid", fgColor="1F4E79")
WHITE_BOLD = Font(bold=True, color="FFFFFF", size=10)
BOLD = Font(bold=True, size=10)
ITALIC_GREY = Font(italic=True, size=9, color="595959")
BOLD_RED = Font(italic=True, bold=True, size=9, color="C00000")

THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PCT_FMT = "0.00%"
PCT_FMT_0 = "0%"
NUM_FMT = "#,##0"

HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center")
LEFT_ALIGN = Alignment(horizontal="left")


def write_header_row(ws, row, headers, widths=None):
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = WHITE_BOLD
        c.fill = NAVY_FILL
        c.alignment = HEADER_ALIGN
        c.border = BORDER
    ws.row_dimensions[row].height = 30
    if widths:
        for ci, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w


def section_header(ws, row, text, n_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = WHITE_BOLD
    c.fill = NAVY_FILL
    c.alignment = LEFT_ALIGN


# ═════════════════════════════════════════════════════════════════════════
# 1. Read + sort data (Python side — needed for row order & scenario brackets)
# ═════════════════════════════════════════════════════════════════════════
rows = []
with open(CSV_PATH, newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for r in reader:
        la, lc = int(r["leads_action"]), int(r["leads_control"])
        ca, cc = int(r["conv_action"]), int(r["conv_control"])
        rra, rrc = ca / la, cc / lc
        rows.append({
            "elig": r["elig_txn_bin"], "mobile": r["mobile_login_bin"],
            "contact": r["prior_contact_bin"],
            "la": la, "lc": lc, "ca": ca, "cc": cc,
            "rra": rra, "rrc": rrc, "lift": rra - rrc,
        })

assert len(rows) == N_CELLS, f"expected {N_CELLS} cells, got {len(rows)}"
rows.sort(key=lambda x: x["lift"])  # ascending lift = weakest cells first

total_la = sum(r["la"] for r in rows)
total_lc = sum(r["lc"] for r in rows)
total_ca = sum(r["ca"] for r in rows)
total_cc = sum(r["cc"] for r in rows)
overall_rra = total_ca / total_la
overall_rrc = total_cc / total_lc
overall_lift = overall_rra - overall_rrc

# Cumulative pass (mirrors the Curve-sheet formulas exactly, integer-for-integer)
cum_la, cum_ca, cum_incr = 0, 0, 0
cum_kept_num = 0.0  # running SUMPRODUCT(leads_action, rr_control) for cut cells
curve_rows = []
for i, r in enumerate(rows, start=1):
    incr = round(r["lift"] * r["la"])
    cum_la += r["la"]
    cum_ca += r["ca"]
    cum_incr += incr
    cum_kept_num += r["la"] * r["rrc"]
    surf_pct = cum_la / total_la
    kept_abs = total_ca - cum_ca + cum_kept_num
    curve_rows.append({
        "rank": i, "excel_row": i + 1,  # Curve/Cells data starts at row 2
        "la": r["la"], "ca": r["ca"], "incr": incr,
        "cum_la": cum_la, "surf_pct": surf_pct,
        "cum_incr": cum_incr, "kept_abs": kept_abs,
    })

bottomup_incr_total = cum_incr  # == curve_rows[-1]["cum_incr"]


def find_bracket(target_pct):
    """Largest curve row with surf_pct <= target, and the next row (for interp)."""
    lo = None
    for cr in curve_rows:
        if cr["surf_pct"] >= target_pct:
            hi = cr
            break
        lo = cr
    else:
        raise ValueError(f"target {target_pct} exceeds max surface pct")
    if lo is None:
        raise ValueError(
            f"target {target_pct} falls before the first cell's cumulative "
            "surface share — origin-interpolation not implemented (not needed "
            "for the current data)."
        )
    return lo, hi


SCENARIO_CUTS = [0.10, 0.20, 0.30, 0.40, 0.50]
scenario_brackets = [(cut, *find_bracket(cut)) for cut in SCENARIO_CUTS]

print("=" * 78)
print("PYTHON-SIDE INDEPENDENT COMPUTATION (cross-check vs. Excel formulas)")
print("=" * 78)
print(f"Total leads_action={total_la:,}  leads_control={total_lc:,}")
print(f"Total conv_action={total_ca:,}  conv_control={total_cc:,}")
print(f"Overall RR action={overall_rra:.4%}  control={overall_rrc:.4%}  "
      f"lift={overall_lift:.4%}pp")
print(f"Bottom-up sum of per-cell incr_converters (ties to Curve total) = "
      f"{bottomup_incr_total:,}")
print(f"(Pooled overall_lift * total_la = {overall_lift * total_la:,.0f} — "
      f"differs from the bottom-up sum because control-rate weighting varies "
      f"cell to cell; this is expected, not an error.)")

# ═════════════════════════════════════════════════════════════════════════
# 2. Build workbook
# ═════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ── Sheet 1: Cells ──────────────────────────────────────────────────────
ws_c = wb.active
ws_c.title = "Cells"
ws_c.sheet_view.showGridLines = False
ws_c.freeze_panes = "A2"

cells_headers = [
    "elig_txn_bin", "mobile_login_bin", "prior_contact_bin",
    "leads_action", "leads_control", "conv_action", "conv_control",
    "rr_action", "rr_control", "lift_pp", "incr_converters",
    "surface_share", "converter_share",
]
cells_widths = [14, 16, 16, 13, 13, 12, 13, 11, 11, 10, 15, 13, 15]
write_header_row(ws_c, 1, cells_headers, cells_widths)

for i, r in enumerate(rows, start=FIRST_ROW):
    ws_c.cell(row=i, column=1, value=r["elig"])
    ws_c.cell(row=i, column=2, value=r["mobile"])
    ws_c.cell(row=i, column=3, value=r["contact"])
    ws_c.cell(row=i, column=4, value=r["la"])
    ws_c.cell(row=i, column=5, value=r["lc"])
    ws_c.cell(row=i, column=6, value=r["ca"])
    ws_c.cell(row=i, column=7, value=r["cc"])
    ws_c.cell(row=i, column=8, value=f"=F{i}/D{i}")
    ws_c.cell(row=i, column=9, value=f"=G{i}/E{i}")
    ws_c.cell(row=i, column=10, value=f"=H{i}-I{i}")
    ws_c.cell(row=i, column=11, value=f"=ROUND(J{i}*D{i},0)")
    ws_c.cell(row=i, column=12, value=f"=D{i}/SUM($D${FIRST_ROW}:$D${LAST_ROW})")
    ws_c.cell(row=i, column=13, value=f"=F{i}/SUM($F${FIRST_ROW}:$F${LAST_ROW})")

    for ci in (4, 5, 6, 7, 11):
        ws_c.cell(row=i, column=ci).number_format = NUM_FMT
    for ci in (8, 9, 10, 12, 13):
        ws_c.cell(row=i, column=ci).number_format = PCT_FMT
    for ci in range(1, 14):
        cell = ws_c.cell(row=i, column=ci)
        cell.border = BORDER
        if ci <= 3:
            cell.alignment = LEFT_ALIGN
        else:
            cell.alignment = CENTER_ALIGN

# ── Sheet 2: Curve ──────────────────────────────────────────────────────
ws_v = wb.create_sheet("Curve")
ws_v.sheet_view.showGridLines = False
ws_v.freeze_panes = "A2"

curve_headers = [
    "rank", "elig_txn_bin", "mobile_login_bin", "prior_contact_bin",
    "leads_action", "conv_action", "rr_control", "incr_converters",
    "cum_leads_action", "cum_surface_pct", "cum_incr_converters_lost",
    "cum_incr_pct_lost", "converters_kept_abs", "converters_kept_pct",
]
curve_widths = [7, 14, 16, 16, 13, 12, 11, 15, 16, 15, 20, 17, 19, 19]
write_header_row(ws_v, 1, curve_headers, curve_widths)

for i in range(FIRST_ROW, LAST_ROW + 1):
    rank = i - 1
    ws_v.cell(row=i, column=1, value=rank)
    ws_v.cell(row=i, column=2, value=f"=Cells!A{i}")
    ws_v.cell(row=i, column=3, value=f"=Cells!B{i}")
    ws_v.cell(row=i, column=4, value=f"=Cells!C{i}")
    ws_v.cell(row=i, column=5, value=f"=Cells!D{i}")
    ws_v.cell(row=i, column=6, value=f"=Cells!F{i}")
    ws_v.cell(row=i, column=7, value=f"=Cells!I{i}")
    ws_v.cell(row=i, column=8, value=f"=Cells!K{i}")
    ws_v.cell(row=i, column=9, value=f"=SUM($E${FIRST_ROW}:E{i})")
    ws_v.cell(row=i, column=10, value=f"=I{i}/SUM($E${FIRST_ROW}:$E${LAST_ROW})")
    ws_v.cell(row=i, column=11, value=f"=SUM($H${FIRST_ROW}:H{i})")
    ws_v.cell(row=i, column=12, value=f"=K{i}/SUM($H${FIRST_ROW}:$H${LAST_ROW})")
    ws_v.cell(
        row=i, column=13,
        value=(f"=SUM($F${FIRST_ROW}:$F${LAST_ROW})-SUM($F${FIRST_ROW}:F{i})"
               f"+SUMPRODUCT($E${FIRST_ROW}:E{i},$G${FIRST_ROW}:G{i})"),
    )
    ws_v.cell(row=i, column=14, value=f"=M{i}/SUM($F${FIRST_ROW}:$F${LAST_ROW})")

    for ci in (1, 5, 6, 8, 9, 11, 13):
        ws_v.cell(row=i, column=ci).number_format = NUM_FMT
    for ci in (7, 10, 12, 14):
        ws_v.cell(row=i, column=ci).number_format = PCT_FMT
    for ci in range(1, 15):
        cell = ws_v.cell(row=i, column=ci)
        cell.border = BORDER
        cell.alignment = CENTER_ALIGN if ci != 2 and ci != 3 and ci != 4 else LEFT_ALIGN

# Coverage-curve chart: X = cum_surface_pct (J), Y = converters_kept_pct (N).
# Scatter (not category-line) so the X axis is a true numeric axis — cells vary
# hugely in size, so evenly-spaced categories would distort the curve.
chart = ScatterChart()
chart.title = "Coverage Curve — Converters Kept vs. Surface Cut"
chart.style = 2
chart.x_axis.title = "Cumulative surface cut (% of Action leads)"
chart.y_axis.title = "Converters kept (%)"
chart.x_axis.numFmt = "0%"
chart.y_axis.numFmt = "0%"
chart.x_axis.majorGridlines = None
chart.height = 12
chart.width = 24

xvalues = Reference(ws_v, min_col=10, min_row=FIRST_ROW, max_row=LAST_ROW)
yvalues = Reference(ws_v, min_col=14, min_row=1, max_row=LAST_ROW)
series = Series(yvalues, xvalues, title_from_data=True)
series.marker = Marker(symbol="none")
series.graphicalProperties.line.width = 20000  # ~1.6pt
series.graphicalProperties.line.solidFill = "1F4E79"
series.smooth = False
chart.series.append(series)
ws_v.add_chart(chart, "P2")

# ── Sheet 3: Summary ────────────────────────────────────────────────────
ws_s = wb.create_sheet("Summary")
ws_s.sheet_view.showGridLines = False
ws_s.column_dimensions["A"].width = 3
ws_s.column_dimensions["B"].width = 32
for col in "CDEF":
    ws_s.column_dimensions[col].width = 20

ws_s.merge_cells("B1:F1")
title_cell = ws_s.cell(row=1, column=2, value="CRV Suppression — Coverage Curve Summary")
title_cell.font = Font(bold=True, size=13, color="1F4E79")

# -- Totals block --
section_header(ws_s, 3, "Totals (Pooled)", 6)
totals = [
    ("Total leads — Action", "=SUM(Cells!D2:D101)", NUM_FMT),
    ("Total leads — Control", "=SUM(Cells!E2:E101)", NUM_FMT),
    ("Total converters — Action", "=SUM(Cells!F2:F101)", NUM_FMT),
    ("Total converters — Control", "=SUM(Cells!G2:G101)", NUM_FMT),
]
row = 4
for label, formula, fmt in totals:
    ws_s.cell(row=row, column=2, value=label).font = BOLD
    v = ws_s.cell(row=row, column=3, value=formula)
    v.number_format = fmt
    v.border = BORDER
    row += 1

leads_a_row, leads_c_row, conv_a_row, conv_c_row = 4, 5, 6, 7

ws_s.cell(row=row, column=2, value="Overall RR — Action (pooled)").font = BOLD
overall_rra_cell = f"C{conv_a_row}/C{leads_a_row}"
c = ws_s.cell(row=row, column=3, value=f"={overall_rra_cell}")
c.number_format = PCT_FMT
c.border = BORDER
rra_row = row
row += 1

ws_s.cell(row=row, column=2, value="Overall RR — Control (pooled)").font = BOLD
overall_rrc_cell = f"C{conv_c_row}/C{leads_c_row}"
c = ws_s.cell(row=row, column=3, value=f"={overall_rrc_cell}")
c.number_format = PCT_FMT
c.border = BORDER
rrc_row = row
row += 1

ws_s.cell(row=row, column=2, value="Overall lift (pp, pooled counts)").font = Font(bold=True, color="1F4E79")
c = ws_s.cell(row=row, column=3, value=f"=C{rra_row}-C{rrc_row}")
c.number_format = PCT_FMT
c.border = BORDER
c.font = Font(bold=True, color="1F4E79")
row += 1

ws_s.cell(row=row, column=2, value="Total incremental converters").font = BOLD
c = ws_s.cell(row=row, column=3, value="=SUM(Cells!K2:K101)")
c.number_format = NUM_FMT
c.border = BORDER
row += 2

# -- Scenario table --
scen_header_row = row
section_header(ws_s, scen_header_row, "Contact-Suppression Scenarios (cut weakest-lift cells first)", 6)
row += 1
scen_col_headers = [
    "Surface cut %", "Leads cut", "Incremental converters lost",
    "Absolute converters kept", "% converters kept",
]
for ci, h in enumerate(scen_col_headers, start=2):
    c = ws_s.cell(row=row, column=ci, value=h)
    c.font = WHITE_BOLD
    c.fill = NAVY_FILL
    c.alignment = HEADER_ALIGN
    c.border = BORDER
ws_s.row_dimensions[row].height = 30
row += 1

scenario_first_data_row = row
for cut, lo, hi in scenario_brackets:
    lo_r, hi_r = lo["excel_row"], hi["excel_row"]
    cut_cell = ws_s.cell(row=row, column=2, value=cut)
    cut_cell.number_format = PCT_FMT_0
    cut_cell.border = BORDER
    cut_cell.alignment = CENTER_ALIGN
    cut_cell.font = BOLD

    frac = f"((B{row}-Curve!J{lo_r})/(Curve!J{hi_r}-Curve!J{lo_r}))"

    leads_cut = ws_s.cell(
        row=row, column=3,
        value=f"=ROUND(Curve!I{lo_r}+{frac}*(Curve!I{hi_r}-Curve!I{lo_r}),0)",
    )
    leads_cut.number_format = NUM_FMT

    incr_lost = ws_s.cell(
        row=row, column=4,
        value=f"=ROUND(Curve!K{lo_r}+{frac}*(Curve!K{hi_r}-Curve!K{lo_r}),0)",
    )
    incr_lost.number_format = NUM_FMT

    kept_abs = ws_s.cell(
        row=row, column=5,
        value=f"=ROUND(Curve!M{lo_r}+{frac}*(Curve!M{hi_r}-Curve!M{lo_r}),0)",
    )
    kept_abs.number_format = NUM_FMT

    kept_pct = ws_s.cell(
        row=row, column=6,
        value=f"=Curve!N{lo_r}+{frac}*(Curve!N{hi_r}-Curve!N{lo_r})",
    )
    kept_pct.number_format = PCT_FMT

    for ci in range(3, 7):
        ws_s.cell(row=row, column=ci).border = BORDER
        ws_s.cell(row=row, column=ci).alignment = CENTER_ALIGN

    row += 1

row += 1

# -- Footer notes --
footer_notes = [
    "Control cells are small — single-cell lift is noisy; read rollups and "
    "the pooled curve, not individual cells.",
    "2 control-converter values (e.7+/c.10-29/d.5-9 and e.10+) inferred from "
    "displayed rates due to screenshot tooltip occlusion.",
    "Cells sheet is sorted by lift ascending; Curve and this scenario table "
    "reference that fixed order. If edited counts reorder cells by lift, "
    "rerun build_coverage_curve.py to re-sort and re-bracket the scenarios.",
]
for note in footer_notes:
    ws_s.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws_s.cell(row=row, column=2, value=note)
    c.font = BOLD_RED if "Control cells" in note or "inferred" in note else ITALIC_GREY
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws_s.row_dimensions[row].height = 28
    row += 1

# ── Recalc + save ─────────────────────────────────────────────────────────
wb.calculation = CalcProperties(fullCalcOnLoad=True, calcMode="auto")

try:
    wb.save(OUT_PATH)
    print(f"\nSaved: {OUT_PATH}")
except PermissionError:
    alt = OUT_PATH.replace(".xlsx", "_v2.xlsx")
    wb.save(alt)
    OUT_PATH = alt
    print(f"\nOriginal locked — saved as: {OUT_PATH}")


def recalc_with_excel(filepath):
    """Open in Excel, force recalculation, save+close, so cached formula
    values are populated (openpyxl can't evaluate formulas itself)."""
    try:
        import win32com.client
        abs_path = os.path.abspath(filepath)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb_com = excel.Workbooks.Open(abs_path)
        excel.CalculateFullRebuild()
        wb_com.Save()
        wb_com.Close()
        excel.Quit()
        print(f"Recalculated via Excel: {abs_path}")
        return True
    except ImportError:
        print("win32com not available — skipping Excel recalculation")
        return False
    except Exception as e:
        print(f"Excel recalculation failed: {e}")
        try:
            excel.Quit()
        except Exception:
            pass
        return False


recalc_with_excel(OUT_PATH)

# ═════════════════════════════════════════════════════════════════════════
# 3. Verify: re-open, print key cells, cross-check vs. Python computation
# ═════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 78)
print("VERIFICATION — reopening coverage_curve.xlsx with openpyxl (data_only)")
print("=" * 78)

wb2 = openpyxl.load_workbook(OUT_PATH, data_only=True)
print("Sheets:", wb2.sheetnames)

ws_s2 = wb2["Summary"]
excel_leads_a = ws_s2.cell(row=leads_a_row, column=3).value
excel_leads_c = ws_s2.cell(row=leads_c_row, column=3).value
excel_conv_a = ws_s2.cell(row=conv_a_row, column=3).value
excel_conv_c = ws_s2.cell(row=conv_c_row, column=3).value
excel_rra = ws_s2.cell(row=rra_row, column=3).value
excel_rrc = ws_s2.cell(row=rrc_row, column=3).value
excel_lift = ws_s2.cell(row=rrc_row + 1, column=3).value
excel_total_incr = ws_s2.cell(row=rrc_row + 2, column=3).value

print("\n--- Summary sheet totals (Excel-recalculated) ---")
print(f"leads_action={excel_leads_a:,}  leads_control={excel_leads_c:,}")
print(f"conv_action={excel_conv_a:,}  conv_control={excel_conv_c:,}")
print(f"RR action={excel_rra:.4%}  RR control={excel_rrc:.4%}  "
      f"lift={excel_lift:.4%}")
print(f"total incremental converters={excel_total_incr:,}")

# Cross-check vs. independent Python computation
checks = [
    ("total leads_action", excel_leads_a, total_la),
    ("total leads_control", excel_leads_c, total_lc),
    ("total conv_action", excel_conv_a, total_ca),
    ("total conv_control", excel_conv_c, total_cc),
    ("total incremental converters", excel_total_incr, bottomup_incr_total),
]
all_pass = True
for label, excel_val, py_val in checks:
    ok = excel_val == py_val
    all_pass &= ok
    print(f"  CHECK {label}: excel={excel_val:,} python={py_val:,} "
          f"-> {'PASS' if ok else 'FAIL'}")
if abs(excel_rra - overall_rra) > 1e-9 or abs(excel_rrc - overall_rrc) > 1e-9:
    all_pass = False
    print("  CHECK overall rates: FAIL (excel vs python mismatch)")
else:
    print("  CHECK overall rates: PASS")
print(f"\nOVERALL CROSS-CHECK: {'PASS' if all_pass else 'FAIL — investigate before using'}")

print("\n--- 10 lowest-lift cells (Cells sheet, rows 2-11) ---")
ws_c2 = wb2["Cells"]
for r in range(2, 12):
    vals = [ws_c2.cell(row=r, column=c).value for c in range(1, 14)]
    (elig, mob, con, la, lc, ca, cc, rra, rrc, lift, incr, surf, convshare) = vals
    print(f"  row{r:>3}: {elig:>7s}/{mob:>8s}/{con:>7s}  "
          f"la={la:>9,} ca={ca:>6,}  rr_a={rra:7.3%} rr_c={rrc:7.3%} "
          f"lift={lift:+7.3%}  incr={incr:>6,}  surf_share={surf:6.2%} "
          f"conv_share={convshare:6.2%}")

print("\n--- Scenario table (Summary sheet) ---")
for r in range(scenario_first_data_row, scenario_first_data_row + len(SCENARIO_CUTS)):
    cut = ws_s2.cell(row=r, column=2).value
    leads_cut = ws_s2.cell(row=r, column=3).value
    incr_lost = ws_s2.cell(row=r, column=4).value
    kept_abs = ws_s2.cell(row=r, column=5).value
    kept_pct = ws_s2.cell(row=r, column=6).value
    print(f"  cut={cut:>5.0%}  leads_cut={leads_cut:>10,}  "
          f"incr_converters_lost={incr_lost:>7,}  "
          f"converters_kept_abs={kept_abs:>8,}  converters_kept_pct={kept_pct:7.2%}")

print("\nDone.")
