"""
Build PCL MDE Calculator Excel — INCREMENTAL test framing.

Champion = all existing channels (current experience).
Challenger A = Champion + Sales Model placement.
Challenger B = Champion + Mobile Dashboard placement.

Question: does ADDING a new placement improve response by >=5% relative lift?

All calculated cells use Excel formulas so the user can change green input cells
and see everything recalculate.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties
import shutil
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Calculate Minimum Lift"

# ── Styles ──────────────────────────────────────────────────────────────────
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
bold = Font(bold=True)
bold_white = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

pct_fmt = "0.00%"
pct_fmt_4 = "0.0000%"
num_fmt = "#,##0"
ratio_fmt = "0.00"

# ── Column widths ───────────────────────────────────────────────────────────
col_widths = {
    "A": 5, "B": 50, "C": 20, "D": 16, "E": 16,
    "F": 16, "G": 16, "H": 16, "I": 18, "J": 16,
    "K": 16, "L": 50,
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w


# ── Helper: apply style to a cell ──────────────────────────────────────────
def style(cell, font=None, fill=None, fmt=None, align=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    return cell


# ════════════════════════════════════════════════════════════════════════════
# SECTION 1: STATISTICAL INPUTS  (Rows 2-10)
# ════════════════════════════════════════════════════════════════════════════

# Row 2 — section header
for c in range(1, 13):
    style(ws.cell(row=2, column=c), fill=grey_fill)
style(ws.cell(row=2, column=2, value="Statistical Inputs"), font=bold, fill=grey_fill)

# Row 3 — Total sample size
style(ws.cell(row=3, column=2, value="Total sample size"), font=bold)
style(ws.cell(row=3, column=3, value=528000), fill=green_fill, fmt=num_fmt)

# Row 4 — Champion allocation
style(ws.cell(row=4, column=2, value="Champion allocation"), font=bold)
style(ws.cell(row=4, column=3, value=0.30), fill=green_fill, fmt=pct_fmt)

# Row 5 — Challenger A (Sales Model) allocation
style(ws.cell(row=5, column=2, value="Challenger A (Sales Model) allocation"), font=bold)
style(ws.cell(row=5, column=3, value=0.30), fill=green_fill, fmt=pct_fmt)

# Row 6 — Challenger B (Mobile Dashboard) allocation
style(ws.cell(row=6, column=2, value="Challenger B (Mobile Dashboard) allocation"), font=bold)
style(ws.cell(row=6, column=3, value=0.40), fill=green_fill, fmt=pct_fmt)

# Row 7 — Significance level (alpha)
style(ws.cell(row=7, column=2, value="Significance level (alpha)"), font=bold)
style(ws.cell(row=7, column=3, value=0.05), fill=green_fill, fmt=pct_fmt)

# Row 8 — Power level
style(ws.cell(row=8, column=2, value="Power level"), font=bold)
style(ws.cell(row=8, column=3, value=0.80), fill=green_fill, fmt=pct_fmt)

# Row 9 — Target relative lift
style(ws.cell(row=9, column=2, value="Target relative lift"), font=bold)
style(ws.cell(row=9, column=3, value=0.05), fill=green_fill, fmt=pct_fmt)
ws.cell(row=9, column=4, value="The lift we want to detect (5% = 5% of baseline)")
ws.cell(row=9, column=4).font = Font(italic=True, color="404040")

# Row 10 — Baseline RR (mobile, population-level)
style(ws.cell(row=10, column=2, value="Baseline RR (mobile, population-level)"), font=bold)
style(ws.cell(row=10, column=3, value=0.13), fill=green_fill, fmt=pct_fmt)
ws.cell(row=10, column=4, value="3-month avg mobile RR")
ws.cell(row=10, column=4).font = Font(italic=True, color="404040")


# ════════════════════════════════════════════════════════════════════════════
# SECTION 2: DO NOT EDIT  (Rows 12-15)
# ════════════════════════════════════════════════════════════════════════════

for c in range(1, 13):
    style(ws.cell(row=12, column=c), fill=grey_fill)
style(ws.cell(row=12, column=2, value="Do Not Edit \u2014 Statistical Constants"), font=bold, fill=grey_fill)

# Row 14 — Za = NORM.S.INV(1 - alpha) — hardcoded for alpha=0.05
style(ws.cell(row=14, column=2, value="Critical Value for Significance (Za)"), font=bold)
ws.cell(row=14, column=3).value = 1.6449
ws.cell(row=14, column=3).number_format = "0.0000"

# Row 15 — Zb = NORM.S.INV(power) — hardcoded for power=0.80
style(ws.cell(row=15, column=2, value="Critical Value for Power (Zb)"), font=bold)
ws.cell(row=15, column=3).value = 0.8416
ws.cell(row=15, column=3).number_format = "0.0000"

# Row 16 — note about Za/Zb
ws.cell(row=16, column=2).value = (
    "Za and Zb are computed for alpha=5% and power=80%. "
    "If you change alpha/power, update these manually or use "
    "=NORM.S.INV(1-alpha) and =NORM.S.INV(power)."
)
ws.cell(row=16, column=2).font = Font(italic=True, color="808080", size=9)


# ════════════════════════════════════════════════════════════════════════════
# SECTION 3: DERIVED VALUES  (Rows 17-22)
# ════════════════════════════════════════════════════════════════════════════

for c in range(1, 13):
    style(ws.cell(row=17, column=c), fill=grey_fill)
style(ws.cell(row=17, column=2, value="Derived Values"), font=bold, fill=grey_fill)

# Row 18 — Target absolute lift (pp)
style(ws.cell(row=18, column=2, value="Target absolute lift (pp)"), font=bold)
ws.cell(row=18, column=3).value = "=C10*C9"
ws.cell(row=18, column=3).number_format = pct_fmt
ws.cell(row=18, column=4, value="Converts relative lift to absolute pp")
ws.cell(row=18, column=4).font = Font(italic=True, color="404040")

# Row 19 — Champion n
style(ws.cell(row=19, column=2, value="Champion n"), font=bold)
ws.cell(row=19, column=3).value = "=C3*C4"
ws.cell(row=19, column=3).number_format = num_fmt

# Row 20 — Challenger A n
style(ws.cell(row=20, column=2, value="Challenger A (Sales Model) n"), font=bold)
ws.cell(row=20, column=3).value = "=C3*C5"
ws.cell(row=20, column=3).number_format = num_fmt

# Row 21 — Challenger B n
style(ws.cell(row=21, column=2, value="Challenger B (Mobile Dashboard) n"), font=bold)
ws.cell(row=21, column=3).value = "=C3*C6"
ws.cell(row=21, column=3).number_format = num_fmt

# Row 22 — Allocation check
style(ws.cell(row=22, column=2, value="Allocation check"), font=bold)
ws.cell(row=22, column=3).value = "=C4+C5+C6"
ws.cell(row=22, column=3).number_format = pct_fmt
ws.cell(row=22, column=4, value="Should show 100%")
ws.cell(row=22, column=4).font = Font(italic=True, color="404040")


# ════════════════════════════════════════════════════════════════════════════
# SECTION 4: COMPARISON TESTS  (Rows 24-33)
# ════════════════════════════════════════════════════════════════════════════

for c in range(1, 13):
    style(ws.cell(row=24, column=c), fill=grey_fill)
style(ws.cell(row=24, column=2, value="Comparison Tests"), font=bold, fill=grey_fill)

# Row 25 — Column headers
headers_s4 = [
    "#", "Comparison", "What it tests",
    "n1", "n2", "Sum", "Ratio",
    "Baseline RR", "Target Lift (abs)", "MDE",
    "Powered?", "Interpretation",
]
for i, h in enumerate(headers_s4, start=1):
    cell = ws.cell(row=25, column=i, value=h)
    style(cell, font=bold_white, fill=header_fill, border=thin_border,
          align=Alignment(horizontal="center", wrap_text=True))

# Define the 7 comparisons
comparisons = [
    # Primary comparisons
    {
        "num": 1,
        "label": "Champion vs Challenger A (Sales Model)",
        "test": "Does adding Sales Model lift response by >=5%?",
        "n1": "C19",
        "n2": "C20",
    },
    {
        "num": 2,
        "label": "Champion vs Challenger B (Dashboard)",
        "test": "Does adding Dashboard lift response by >=5%?",
        "n1": "C19",
        "n2": "C21",
    },
    {
        "num": 3,
        "label": "Challenger A vs Challenger B",
        "test": "Which addition performs better?",
        "n1": "C20",
        "n2": "C21",
    },
]

for idx, comp in enumerate(comparisons):
    row = 26 + idx  # rows 26-32

    ws.cell(row=row, column=1, value=comp["num"])
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=comp["label"])
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=3, value=comp["test"])
    ws.cell(row=row, column=3).border = thin_border

    # n1 — either a direct cell ref or a formula
    if "n1_formula" in comp:
        ws.cell(row=row, column=4).value = comp["n1_formula"]
    else:
        ws.cell(row=row, column=4).value = f"={comp['n1']}"
    ws.cell(row=row, column=4).number_format = num_fmt
    ws.cell(row=row, column=4).border = thin_border

    # n2 — either a direct cell ref or a formula
    if "n2_formula" in comp:
        ws.cell(row=row, column=5).value = comp["n2_formula"]
    else:
        ws.cell(row=row, column=5).value = f"={comp['n2']}"
    ws.cell(row=row, column=5).number_format = num_fmt
    ws.cell(row=row, column=5).border = thin_border

    # Sum = n1 + n2
    ws.cell(row=row, column=6).value = f"=D{row}+E{row}"
    ws.cell(row=row, column=6).number_format = num_fmt
    ws.cell(row=row, column=6).border = thin_border

    # Ratio = n1/n2
    ws.cell(row=row, column=7).value = f"=D{row}/E{row}"
    ws.cell(row=row, column=7).number_format = ratio_fmt
    ws.cell(row=row, column=7).border = thin_border

    # Baseline RR — references the input cell C10
    ws.cell(row=row, column=8).value = "=$C$10"
    ws.cell(row=row, column=8).number_format = pct_fmt
    ws.cell(row=row, column=8).border = thin_border

    # Target Lift (abs) — references derived C18
    ws.cell(row=row, column=9).value = "=$C$18"
    ws.cell(row=row, column=9).number_format = pct_fmt
    ws.cell(row=row, column=9).border = thin_border

    # MDE = ROUND((Za+Zb)*SQRT(p0*(1-p0)*(1/n1+1/n2)), 4)
    mde_formula = (
        f"=ROUND(($C$14+$C$15)*SQRT(H{row}*(1-H{row})*(1/D{row}+1/E{row})),4)"
    )
    ws.cell(row=row, column=10).value = mde_formula
    ws.cell(row=row, column=10).number_format = pct_fmt_4
    ws.cell(row=row, column=10).border = thin_border

    # Powered? — compares MDE to Target Lift (abs)
    powered_formula = (
        f'=IF(J{row}<=I{row},"YES - MDE <= Target","NO - need more n or lower target")'
    )
    ws.cell(row=row, column=11).value = powered_formula
    ws.cell(row=row, column=11).border = thin_border
    ws.cell(row=row, column=11).alignment = Alignment(wrap_text=True)

    # Interpretation
    interp_formula = (
        f'="Can detect differences >= "&TEXT(J{row},"0.00%")&" pp"'
    )
    # Note: ">=" is plain text, no CHAR(8805)
    ws.cell(row=row, column=12).value = interp_formula
    ws.cell(row=row, column=12).border = thin_border
    ws.cell(row=row, column=12).alignment = Alignment(wrap_text=True)


# ── Section 5: Bonferroni note (row 34) ──────────────────────────────────────
bonferroni_row = 34
ws.merge_cells(start_row=bonferroni_row, start_column=2, end_row=bonferroni_row, end_column=12)
bonferroni_cell = ws.cell(
    row=bonferroni_row, column=2,
    value=(
        "With 3 comparisons, Bonferroni-adjusted alpha = 5%/3 = 1.67% per comparison."
    ),
)
style(bonferroni_cell, font=Font(italic=True, bold=True, color="C00000"),
      align=Alignment(wrap_text=True, vertical="top"))
ws.row_dimensions[bonferroni_row].height = 35

# ── Section 6: MDE summary note (row 35) ────────────────────────────────────
summary_row = 35
ws.merge_cells(start_row=summary_row, start_column=2, end_row=summary_row, end_column=12)
summary_cell = ws.cell(
    row=summary_row, column=2,
    value=(
        "The MDE is the minimum lift (in percentage points) detectable with 80% power "
        "at the specified significance level."
    ),
)
style(summary_cell, font=Font(italic=True, color="404040"),
      align=Alignment(wrap_text=True, vertical="top"))
ws.row_dimensions[summary_row].height = 35


# ════════════════════════════════════════════════════════════════════════════
# SECTION 7: STRESS TEST SCENARIOS  (Rows 37-42)
# ════════════════════════════════════════════════════════════════════════════

stress_start = 37
for c in range(1, 13):
    style(ws.cell(row=stress_start, column=c), fill=grey_fill)
style(ws.cell(row=stress_start, column=2, value="Stress Test Scenarios"), font=bold, fill=grey_fill)

# Row 38 — headers
stress_headers = [
    "", "Scenario", "Baseline RR", "Target Lift (abs)",
    "Comp 1 MDE", "Comp 2 MDE", "Comp 3 MDE",
    "Worst Case", "Powered?",
]
for i, h in enumerate(stress_headers, start=1):
    cell = ws.cell(row=stress_start + 1, column=i, value=h)
    style(cell, font=bold_white, fill=header_fill, border=thin_border,
          align=Alignment(horizontal="center", wrap_text=True))

# Comp n1/n2 references for the 3 comparisons
comp_n_refs = [
    ("C19", "C20"),                          # Comp 1: Champ vs Chall A
    ("C19", "C21"),                          # Comp 2: Champ vs Chall B
    ("C20", "C21"),                          # Comp 3: Chall A vs Chall B
]

# Scenarios: (label, baseline_rr, note)
scenarios = [
    ("Pessimistic", 0.105, "Jan 2026 mobile: 55,560/528,002"),
    ("Current", 0.13, "3-month avg mobile"),
    ("Optimistic", 0.154, "Nov 2025 mobile: 99,704/648,547"),
]

for s_idx, (scenario_label, baseline, note) in enumerate(scenarios):
    row = stress_start + 2 + s_idx  # rows 39, 40, 41

    ws.cell(row=row, column=2, value=f"{scenario_label} ({note})")
    ws.cell(row=row, column=2).border = thin_border

    # Baseline RR — editable (green)
    style(ws.cell(row=row, column=3, value=baseline), fill=green_fill, fmt=pct_fmt)
    ws.cell(row=row, column=3).border = thin_border

    # Target Lift (abs) = scenario baseline * C9 (the relative lift %)
    ws.cell(row=row, column=4).value = f"=C{row}*$C$9"
    ws.cell(row=row, column=4).number_format = pct_fmt
    ws.cell(row=row, column=4).border = thin_border

    # Comp 1-3 MDE formulas (columns E through G)
    for c_idx, (n1_ref, n2_ref) in enumerate(comp_n_refs):
        col = 5 + c_idx  # columns E=5, F=6, G=7
        p0_ref = f"C{row}"  # this scenario's baseline RR
        mde_f = (
            f"=ROUND(($C$14+$C$15)*SQRT({p0_ref}*(1-{p0_ref})*(1/{n1_ref}+1/{n2_ref})),4)"
        )
        ws.cell(row=row, column=col).value = mde_f
        ws.cell(row=row, column=col).number_format = pct_fmt_4
        ws.cell(row=row, column=col).border = thin_border

    # Worst Case MDE = MAX of the 3 (columns E through G)
    ws.cell(row=row, column=8).value = f"=MAX(E{row}:G{row})"
    ws.cell(row=row, column=8).number_format = pct_fmt_4
    ws.cell(row=row, column=8).border = thin_border

    # Powered? = IF(worst case MDE <= target lift abs, "YES", "NO")
    ws.cell(row=row, column=9).value = f'=IF(H{row}<=D{row},"YES","NO")'
    ws.cell(row=row, column=9).border = thin_border


# ════════════════════════════════════════════════════════════════════════════
# SECTION 8: NOTES  (Rows 44+)
# ════════════════════════════════════════════════════════════════════════════

notes_start = 44
for c in range(1, 13):
    style(ws.cell(row=notes_start, column=c), fill=grey_fill)
style(ws.cell(row=notes_start, column=2, value="Notes"), font=bold, fill=grey_fill)

notes = [
    "1. Champion = all existing channels. Challengers = champion + additional placement.",
    "2. Target lift is RELATIVE (5% = 5% of baseline RR). Absolute target depends on baseline.",
    "3. Baseline RR = mobile population-level rate (mobile responders / total population).",
    "4. PCL mnemonic confirmed by Daniel Chin (campaign called PLI internally).",
    "5. Green cells are editable \u2014 change them to recalculate everything.",
    "6. Pre-register comparisons 1 & 2 as primary; comparison 3 as exploratory.",
]
for i, note in enumerate(notes):
    ws.cell(row=notes_start + 1 + i, column=2, value=note)


# ── Zoom ──────────────────────────────────────────────────────────────────
ws.sheet_view.zoomScale = 90

# ── Force full recalculation on load ─────────────────────────────────────
wb.calculation = CalcProperties(fullCalcOnLoad=True, calcMode='auto')

# ── Save ──────────────────────────────────────────────────────────────────
output_path = r"C:\Users\andre\New_projects\cards\campaigns\PCL\pcl_mde_calculator.xlsx"
# Try primary path, fall back to _v2 if locked
try:
    wb.save(output_path)
    print(f"Saved: {output_path}")
except PermissionError:
    output_path = r"C:\Users\andre\New_projects\cards\campaigns\PCL\pcl_mde_calculator_v2.xlsx"
    wb.save(output_path)
    print(f"Original locked — saved as: {output_path}")

# Copy to schemas folder
copy_dest = r"C:\Users\andre\New_projects\cards\schemas\pcl_mde_calculator.xlsx"
try:
    shutil.copy2(output_path, copy_dest)
    print(f"Copied to: {copy_dest}")
except PermissionError:
    copy_dest = r"C:\Users\andre\New_projects\cards\schemas\pcl_mde_calculator_v2.xlsx"
    shutil.copy2(output_path, copy_dest)
    print(f"Schemas copy locked — saved as: {copy_dest}")

# ── Force recalculation via win32com (Windows only) ──────────────────────
def recalc_with_excel(filepath):
    """Open in Excel, force recalculate all formulas, save and close."""
    try:
        import win32com.client
        abs_path = os.path.abspath(filepath)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb_com = excel.Workbooks.Open(abs_path)
        excel.Calculate()
        wb_com.Save()
        wb_com.Close()
        excel.Quit()
        print(f"Recalculated via Excel: {abs_path}")
        return True
    except ImportError:
        print("win32com not available — skipping Excel recalculation")
        return False
    except Exception as e:
        print(f"Excel recalculation failed: {e}")
        try:
            excel.Quit()
        except:
            pass
        return False

recalc_with_excel(output_path)
recalc_with_excel(copy_dest)
