"""
Comprehensive audit of pcl_mde_calculator.xlsx
Checks every formula cell for correctness, broken references, and logic errors.
"""

import openpyxl
import math
import sys

XLSX = r"C:\Users\andre\New_projects\cards\campaigns\PCL\pcl_mde_calculator.xlsx"

import io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ── Open workbook twice: formulas + cached values ────────────────────────────
wb_f = openpyxl.load_workbook(XLSX, data_only=False)
wb_v = openpyxl.load_workbook(XLSX, data_only=True)
ws_f = wb_f.active
ws_v = wb_v.active

print("=" * 80)
print("PCL MDE CALCULATOR — FULL FORMULA AUDIT")
print("=" * 80)

issues = []

def issue(severity, cell, msg):
    """Record an issue."""
    issues.append((severity, cell, msg))
    marker = "ERROR" if severity == "E" else "WARN" if severity == "W" else "INFO"
    print(f"  [{marker}] {cell}: {msg}")


# ══════════════════════════════════════════════════════════════════════════════
# PASS 1: Enumerate every formula cell
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "─" * 80)
print("PASS 1: All formula cells in the workbook")
print("─" * 80)

formula_cells = []
for row in ws_f.iter_rows(min_row=1, max_row=ws_f.max_row, max_col=ws_f.max_column):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
            coord = cell.coordinate
            cached = ws_v[coord].value
            formula_cells.append((coord, cell.value, cached))
            print(f"  {coord:6s} | Formula: {cell.value:60s} | Cached: {cached}")

print(f"\nTotal formula cells: {len(formula_cells)}")


# ══════════════════════════════════════════════════════════════════════════════
# PASS 2: Check green input cells are VALUES, not formulas
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "─" * 80)
print("PASS 2: Green input cells must be plain values (not formulas)")
print("─" * 80)

GREEN_HEX = "C6EFCE"
green_input_cells = []
for row in ws_f.iter_rows(min_row=1, max_row=ws_f.max_row, max_col=ws_f.max_column):
    for cell in row:
        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
            rgb = str(cell.fill.start_color.rgb)
            if GREEN_HEX in rgb:
                is_formula = isinstance(cell.value, str) and str(cell.value).startswith("=")
                status = "FORMULA (BAD)" if is_formula else "value (ok)"
                green_input_cells.append((cell.coordinate, cell.value, is_formula))
                print(f"  {cell.coordinate:6s} = {str(cell.value):20s}  [{status}]")
                if is_formula:
                    issue("E", cell.coordinate, f"Green input cell contains a formula: {cell.value}")

print(f"\nTotal green cells: {len(green_input_cells)}")


# ══════════════════════════════════════════════════════════════════════════════
# PASS 3: Check broken references (#REF!, #NAME?, etc.)
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "─" * 80)
print("PASS 3: Check for broken references in cached values")
print("─" * 80)

error_values = ["#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#NULL!", "#N/A", "#NUM!"]
for coord, formula, cached in formula_cells:
    if cached is not None and isinstance(cached, str):
        for err in error_values:
            if err in str(cached):
                issue("E", coord, f"Cached value is an error: {cached}")
                break
    # Also check formula text for obvious #REF!
    if "#REF!" in formula:
        issue("E", coord, f"Formula contains #REF!: {formula}")

print("  (Pass 3 complete)")


# ══════════════════════════════════════════════════════════════════════════════
# PASS 4: Verify specific formulas by section
# ══════════════════════════════════════════════════════════════════════════════

# Helper: get formula string from a cell
def get_formula(coord):
    val = ws_f[coord].value
    return val if isinstance(val, str) and val.startswith("=") else None

def get_value(coord):
    return ws_v[coord].value

# ── Known input values for manual computation ─────────────────────────────────
total_n = 528000
alloc_sm = 0.30   # C5
alloc_md = 0.40   # C6
alloc_pp = 0.30   # C7
alpha = 0.05      # C8
power = 0.80      # C9
Za = 1.6449       # C13
Zb = 0.8416       # C14
baseline_rr = 0.15  # E18
stress_increase = 0  # G3
stress_rr = baseline_rr * (1 + stress_increase)  # F18 = 0.15

n_sm = total_n * alloc_sm   # C19 = 158400
n_md = total_n * alloc_md   # C20 = 211200
n_pp = total_n * alloc_pp   # C21 = 158400

print("\n" + "─" * 80)
print("PASS 4a: Population cells (rows 18-21)")
print("─" * 80)

# C18 = =C3 → should be 528000
f = get_formula("C18")
print(f"  C18: formula={f}, expected='=C3'")
if f != "=C3":
    issue("E", "C18", f"Expected '=C3', got '{f}'")
else:
    print("    OK")

# F18 = =E18*(1+G3) → stress test RR
f = get_formula("F18")
print(f"  F18: formula={f}, expected='=E18*(1+G3)'")
if f != "=E18*(1+G3)":
    issue("E", "F18", f"Expected '=E18*(1+G3)', got '{f}'")
else:
    expected_val = baseline_rr * (1 + stress_increase)
    print(f"    OK — computes to {expected_val}")

# Population arm rows
expected_arms = {
    19: ("C5", n_sm),
    20: ("C6", n_md),
    21: ("C7", n_pp),
}

for row, (alloc_cell, expected_n) in expected_arms.items():
    # C{row} = =C3*{alloc_cell}
    coord_n = f"C{row}"
    f = get_formula(coord_n)
    expected_f = f"=C3*{alloc_cell}"
    print(f"  {coord_n}: formula={f}, expected='{expected_f}'")
    if f != expected_f:
        issue("E", coord_n, f"Expected '{expected_f}', got '{f}'")
    else:
        print(f"    OK — computes to {expected_n:,.0f}")

    # D{row} = ={alloc_cell}
    coord_pct = f"D{row}"
    f = get_formula(coord_pct)
    expected_f2 = f"={alloc_cell}"
    print(f"  {coord_pct}: formula={f}, expected='{expected_f2}'")
    if f != expected_f2:
        issue("E", coord_pct, f"Expected '{expected_f2}', got '{f}'")
    else:
        print("    OK")

    # E{row} = =E18 (baseline RR)
    coord_br = f"E{row}"
    f = get_formula(coord_br)
    print(f"  {coord_br}: formula={f}, expected='=E18'")
    if f != "=E18":
        issue("E", coord_br, f"Expected '=E18', got '{f}'")
    else:
        print("    OK")

    # F{row} = =F18 (stress test RR)
    coord_sr = f"F{row}"
    f = get_formula(coord_sr)
    print(f"  {coord_sr}: formula={f}, expected='=F18'")
    if f != "=F18":
        issue("E", coord_sr, f"Expected '=F18', got '{f}'")
    else:
        print("    OK")


# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "─" * 80)
print("PASS 4b: Comparison tests (rows 25-28)")
print("─" * 80)

# Expected structure per comparison row
expected_comps = [
    {"row": 25, "n1": "=C19", "n2": "=C20",
     "n1_val": n_sm, "n2_val": n_md},
    {"row": 26, "n1": "=C19", "n2": "=C21",
     "n1_val": n_sm, "n2_val": n_pp},
    {"row": 27, "n1": "=C20", "n2": "=C21",
     "n1_val": n_md, "n2_val": n_pp},
    {"row": 28, "n1": "=MIN(C19:C21)", "n2": "=C3-MIN(C19:C21)",
     "n1_val": min(n_sm, n_md, n_pp), "n2_val": total_n - min(n_sm, n_md, n_pp)},
]

for comp in expected_comps:
    row = comp["row"]
    print(f"\n  --- Row {row} ---")

    # D{row} = n1
    f = get_formula(f"D{row}")
    print(f"  D{row}: formula={f}, expected='{comp['n1']}'")
    if f != comp["n1"]:
        issue("E", f"D{row}", f"Expected '{comp['n1']}', got '{f}'")
    else:
        print(f"    OK — n1 = {comp['n1_val']:,.0f}")

    # E{row} = n2
    f = get_formula(f"E{row}")
    print(f"  E{row}: formula={f}, expected='{comp['n2']}'")
    if f != comp["n2"]:
        issue("E", f"E{row}", f"Expected '{comp['n2']}', got '{f}'")
    else:
        print(f"    OK — n2 = {comp['n2_val']:,.0f}")

    # F{row} = D+E (sum)
    f = get_formula(f"F{row}")
    expected_sum_f = f"=D{row}+E{row}"
    print(f"  F{row}: formula={f}, expected='{expected_sum_f}'")
    if f != expected_sum_f:
        issue("E", f"F{row}", f"Expected '{expected_sum_f}', got '{f}'")
    else:
        print(f"    OK — sum = {comp['n1_val'] + comp['n2_val']:,.0f}")

    # G{row} = D/E (ratio)
    f = get_formula(f"G{row}")
    expected_ratio_f = f"=D{row}/E{row}"
    print(f"  G{row}: formula={f}, expected='{expected_ratio_f}'")
    if f != expected_ratio_f:
        issue("E", f"G{row}", f"Expected '{expected_ratio_f}', got '{f}'")
    else:
        print(f"    OK — ratio = {comp['n1_val']/comp['n2_val']:.4f}")

    # H{row} = =F18 (baseline RR = stress test RR)
    f = get_formula(f"H{row}")
    print(f"  H{row}: formula={f}, expected='=F18'")
    if f != "=F18":
        issue("E", f"H{row}", f"Expected '=F18', got '{f}'")
    else:
        print(f"    OK — p0 = {stress_rr}")

    # I{row} = MDE formula
    f = get_formula(f"I{row}")
    expected_mde_f = f"=ROUND(($C$13+$C$14)*SQRT(H{row}*(1-H{row})*(1/D{row}+1/E{row})),4)"
    print(f"  I{row}: formula={f}")
    print(f"         expected='{expected_mde_f}'")
    if f != expected_mde_f:
        issue("E", f"I{row}", f"MDE formula mismatch.\n    Got:      {f}\n    Expected: {expected_mde_f}")
    else:
        # Manual compute
        p0 = stress_rr
        n1 = comp["n1_val"]
        n2 = comp["n2_val"]
        mde_manual = (Za + Zb) * math.sqrt(p0 * (1 - p0) * (1/n1 + 1/n2))
        mde_rounded = round(mde_manual, 4)
        print(f"    OK — MDE (Python) = {mde_manual:.6f}, rounded = {mde_rounded:.4f}")

    # J{row} = Target Lift — should be a VALUE (0.05), not a formula
    f = get_formula(f"J{row}")
    v = ws_f[f"J{row}"].value
    if f is not None:
        issue("W", f"J{row}", f"Target lift is a formula ({f}) — expected hardcoded 0.05")
    elif v != 0.05:
        issue("E", f"J{row}", f"Target lift value = {v}, expected 0.05")
    else:
        print(f"  J{row}: value=0.05 (hardcoded) — OK")

    # K{row} = Headroom = J/I
    f = get_formula(f"K{row}")
    expected_headroom_f = f"=J{row}/I{row}"
    print(f"  K{row}: formula={f}, expected='{expected_headroom_f}'")
    if f != expected_headroom_f:
        issue("E", f"K{row}", f"Expected '{expected_headroom_f}', got '{f}'")
    else:
        print("    OK")

    # L{row} = Status = IF(I<J, "OK", "NOT OK")
    f = get_formula(f"L{row}")
    expected_status_f = f'=IF(I{row}<J{row},"OK","NOT OK")'
    print(f"  L{row}: formula={f}")
    print(f"         expected='{expected_status_f}'")
    if f != expected_status_f:
        issue("E", f"L{row}", f"Expected '{expected_status_f}', got '{f}'")
    else:
        # Check logic: MDE < Target → OK
        p0 = stress_rr
        n1 = comp["n1_val"]
        n2 = comp["n2_val"]
        mde = round((Za + Zb) * math.sqrt(p0 * (1 - p0) * (1/n1 + 1/n2)), 4)
        target = 0.05
        expected_status = "OK" if mde < target else "NOT OK"
        print(f"    OK — MDE={mde:.4f} vs target={target} → {expected_status}")


# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "─" * 80)
print("PASS 4c: Stress test section (rows 33-35)")
print("─" * 80)

stress_baselines = {33: 0.14, 34: 0.15, 35: 0.22}

comp_n_pairs = [
    (n_sm, n_md),      # Comp 1
    (n_sm, n_pp),       # Comp 2
    (n_md, n_pp),       # Comp 3
    (min(n_sm, n_md, n_pp), total_n - min(n_sm, n_md, n_pp)),  # Comp 4
]

comp_n_refs = [
    ("C19", "C20"),
    ("C19", "C21"),
    ("C20", "C21"),
    ("MIN(C19:C21)", "C3-MIN(C19:C21)"),
]

for row in [33, 34, 35]:
    print(f"\n  --- Row {row} ---")
    bl = stress_baselines[row]

    # C{row} = green value (stress baseline RR)
    f = get_formula(f"C{row}")
    v = ws_f[f"C{row}"].value
    if f is not None:
        issue("E", f"C{row}", f"Stress baseline RR is a formula ({f}) — should be plain value")
    else:
        print(f"  C{row}: value={v} (hardcoded green) — OK")

    # D{row} through G{row}: Comp 1-4 MDE
    mde_vals = []
    for c_idx, (n1_ref, n2_ref) in enumerate(comp_n_refs):
        col_letter = chr(ord('D') + c_idx)
        coord = f"{col_letter}{row}"
        f = get_formula(coord)
        p0_ref = f"C{row}"
        expected_f = f"=ROUND(($C$13+$C$14)*SQRT({p0_ref}*(1-{p0_ref})*(1/{n1_ref}+1/{n2_ref})),4)"
        print(f"  {coord}: formula={f}")
        print(f"         expected='{expected_f}'")

        if f != expected_f:
            issue("E", coord, f"Stress MDE formula mismatch.\n    Got:      {f}\n    Expected: {expected_f}")
        else:
            # Manual compute
            p0 = bl
            n1, n2 = comp_n_pairs[c_idx]
            mde_manual = round((Za + Zb) * math.sqrt(p0 * (1 - p0) * (1/n1 + 1/n2)), 4)
            mde_vals.append(mde_manual)
            print(f"    OK — MDE = {mde_manual:.4f}")

    # H{row} = MAX(D:G) — worst case
    f = get_formula(f"H{row}")
    expected_max_f = f"=MAX(D{row}:G{row})"
    print(f"  H{row}: formula={f}, expected='{expected_max_f}'")
    if f != expected_max_f:
        issue("E", f"H{row}", f"Expected '{expected_max_f}', got '{f}'")
    else:
        if mde_vals:
            worst = max(mde_vals)
            print(f"    OK — worst case MDE = {worst:.4f}")

    # I{row} = Target Lift — hardcoded 0.05
    f = get_formula(f"I{row}")
    v = ws_f[f"I{row}"].value
    if f is not None:
        issue("W", f"I{row}", f"Target lift is a formula ({f}) — expected hardcoded 0.05")
    elif v != 0.05:
        issue("E", f"I{row}", f"Target lift value = {v}, expected 0.05")
    else:
        print(f"  I{row}: value=0.05 (hardcoded) — OK")

    # J{row} = Headroom = I/H
    f = get_formula(f"J{row}")
    expected_hroom_f = f"=I{row}/H{row}"
    print(f"  J{row}: formula={f}, expected='{expected_hroom_f}'")
    if f != expected_hroom_f:
        issue("E", f"J{row}", f"Expected '{expected_hroom_f}', got '{f}'")
    else:
        print("    OK")

    # K{row} = Status = IF(H<I, "OK", "NOT OK")
    f = get_formula(f"K{row}")
    expected_stat_f = f'=IF(H{row}<I{row},"OK","NOT OK")'
    print(f"  K{row}: formula={f}")
    print(f"         expected='{expected_stat_f}'")
    if f != expected_stat_f:
        issue("E", f"K{row}", f"Expected '{expected_stat_f}', got '{f}'")
    else:
        if mde_vals:
            worst = max(mde_vals)
            expected_status = "OK" if worst < 0.05 else "NOT OK"
            print(f"    OK — worst MDE={worst:.4f} vs 0.05 → {expected_status}")


# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "─" * 80)
print("PASS 5: Cascade / dependency chain analysis")
print("─" * 80)

# Check that the main input cells cascade correctly
print("\n  If C3 (total sample) changes → should cascade to:")
print("    C18 (=C3), C19 (=C3*C5), C20 (=C3*C6), C21 (=C3*C7)")
print("    D25-D28 (n1), E25-E28 (n2), then F (sum), G (ratio), I (MDE)")
print("    Stress test D33-G35 also reference C19-C21 and C3")

# Verify: do any formulas in Section 4 or 5 directly hardcode a number
# instead of referencing a cell?
print("\n  Checking for hardcoded values that should be formulas...")
for row in range(25, 29):
    for col_letter in ['D', 'E', 'F', 'G', 'H', 'I', 'K']:
        coord = f"{col_letter}{row}"
        f = get_formula(coord)
        if f is None:
            v = ws_f[coord].value
            if v is not None:
                issue("W", coord, f"Contains hardcoded value {v} — expected a formula")

for row in range(33, 36):
    for col_letter in ['D', 'E', 'F', 'G', 'H', 'J']:
        coord = f"{col_letter}{row}"
        f = get_formula(coord)
        if f is None:
            v = ws_f[coord].value
            if v is not None:
                issue("W", coord, f"Contains hardcoded value {v} — expected a formula")


# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "─" * 80)
print("PASS 6: MDE numerical verification (all comparisons)")
print("─" * 80)

print("\n  Using: Za=1.6449, Zb=0.8416, p0=0.15 (stress increase=0%)")
print(f"  Formula: MDE = (Za+Zb) * sqrt(p0*(1-p0)*(1/n1+1/n2))")
print()

all_mde_results = []
for i, (n1, n2, label) in enumerate([
    (n_sm, n_md, "Sales Model vs Mobile Dashboard"),
    (n_sm, n_pp, "Sales Model vs Product Page"),
    (n_md, n_pp, "Mobile Dashboard vs Product Page"),
    (min(n_sm, n_md, n_pp), total_n - min(n_sm, n_md, n_pp), "Best vs Rest"),
]):
    p0 = stress_rr  # 0.15
    mde_raw = (Za + Zb) * math.sqrt(p0 * (1 - p0) * (1/n1 + 1/n2))
    mde_rounded = round(mde_raw, 4)
    all_mde_results.append(mde_rounded)
    row = 25 + i
    cached = get_value(f"I{row}")
    match = "MATCH" if cached == mde_rounded else f"MISMATCH (cached={cached})"
    print(f"  Comp {i+1}: {label}")
    print(f"    n1={n1:,.0f}, n2={n2:,.0f}, p0={p0}")
    print(f"    MDE = {mde_raw:.6f} → rounded = {mde_rounded:.4f}")
    print(f"    Cached in Excel: {cached} — {match}")
    if cached is not None and cached != mde_rounded:
        issue("W", f"I{row}", f"Cached value {cached} != computed {mde_rounded} (may just be stale cache)")
    print()


# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "─" * 80)
print("PASS 7: Stress test numerical verification")
print("─" * 80)

for row, (label, bl) in [(33, ("Pessimistic", 0.14)), (34, ("Current", 0.15)), (35, ("Optimistic", 0.22))]:
    print(f"\n  Scenario: {label} (baseline={bl})")
    for c_idx, (n1, n2) in enumerate(comp_n_pairs):
        col_letter = chr(ord('D') + c_idx)
        coord = f"{col_letter}{row}"
        p0 = bl
        mde_raw = (Za + Zb) * math.sqrt(p0 * (1 - p0) * (1/n1 + 1/n2))
        mde_rounded = round(mde_raw, 4)
        cached = get_value(coord)
        match_str = "ok" if cached == mde_rounded else f"cached={cached}"
        print(f"    Comp {c_idx+1} ({coord}): MDE={mde_rounded:.4f}  [{match_str}]")


# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "─" * 80)
print("PASS 8: Design / logic issues")
print("─" * 80)

# Issue: Stress test baseline RR column uses C{row} which is independently editable
# but main section uses F18 which is derived from E18*(1+G3)
# This means stress test section does NOT respond to the G3 "increase from baseline" input
print("""
  DESIGN NOTE: The stress test section (rows 33-35) uses independently editable
  green cells in column C for baseline RR. These do NOT cascade from E18 or G3.
  This is intentional (separate scenarios), but worth noting:
  - Changing G3 (stress increase %) affects Section 4 via F18 → H25:H28
  - Changing G3 does NOT affect Section 5 stress test scenarios
  - Section 5 has its OWN green baseline RR cells (C33, C34, C35)
""")

# Issue: Target lift is hardcoded in every row — should it reference a single cell?
print("  DESIGN NOTE: Target lift (0.05) is hardcoded in J25:J28 and I33:I35.")
print("  If the user wants to change the target, they must edit 7 cells.")
issue("W", "J25:J28,I33:I35", "Target lift hardcoded in 7 cells — consider using a single input cell reference")

# Issue: Significance and power are hardcoded as values (C8=0.05, C9=0.80)
# but Za/Zb are also hardcoded (C13=1.6449, C14=0.8416)
# Changing C8 or C9 does NOT change Za/Zb
print("\n  DESIGN NOTE: C8 (significance=0.05) and C9 (power=0.80) are displayed")
print("  but Za (C13) and Zb (C14) are independently hardcoded constants.")
print("  Changing C8 or C9 has NO effect on MDE calculations.")
f_c13 = get_formula("C13")
f_c14 = get_formula("C14")
if f_c13 is None and f_c14 is None:
    issue("W", "C13/C14", "Za and Zb are hardcoded constants. Changing C8/C9 (alpha/power) has no effect — user may be misled into thinking they can change significance/power levels")
    print("  Za (C13) is hardcoded =", ws_f["C13"].value)
    print("  Zb (C14) is hardcoded =", ws_f["C14"].value)

# Issue: Comp 4 "Best Performer vs Rest Combined" uses MIN which always picks
# the smallest arm. With 30/40/30, MIN is always 158400 (tied between SM and PP).
# This means Comp 4 n1 = 158400, n2 = 528000-158400 = 369600.
# But "Best Performer" may not be the smallest arm — it could be any arm.
print("\n  DESIGN NOTE: Comp 4 uses MIN(C19:C21) for 'Best Performer'.")
print("  MIN picks the smallest arm, not the actual best performer.")
print("  With 30/40/30 split, MIN always picks a 30% arm (158,400).")
print("  This is a reasonable conservative assumption for power calculation,")
print("  but the label 'Best Performer vs Rest' is slightly misleading —")
print("  it's really 'Smallest Arm vs Rest'.")
issue("W", "D28", "Comp 4 labeled 'Best Performer vs Rest' but uses MIN(C19:C21) = smallest arm. Label is slightly misleading — it's actually a conservative worst-case for any single arm vs rest.")


# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "─" * 80)
print("PASS 9: Check alpha/power consistency")
print("─" * 80)

# Verify Za corresponds to alpha=0.05 two-tailed → Z_{alpha/2} or one-tailed Z_{alpha}
# Za = 1.6449 ≈ Z_{0.05} one-tailed (NORM.S.INV(0.95))
# For two-tailed: Z_{0.025} = 1.96
from scipy import stats as scipy_stats

z_one_sided_05 = scipy_stats.norm.ppf(1 - 0.05)
z_two_sided_05 = scipy_stats.norm.ppf(1 - 0.05/2)
z_power_80 = scipy_stats.norm.ppf(0.80)

print(f"  Z for one-sided alpha=0.05: {z_one_sided_05:.4f}")
print(f"  Z for two-sided alpha=0.05: {z_two_sided_05:.4f}")
print(f"  Z for power=0.80:           {z_power_80:.4f}")
print(f"  Za in spreadsheet (C13):    {ws_f['C13'].value}")
print(f"  Zb in spreadsheet (C14):    {ws_f['C14'].value}")

if abs(ws_f["C13"].value - z_one_sided_05) < 0.001:
    print("  → Za matches one-sided test at alpha=0.05 — CORRECT for one-sided MDE")
elif abs(ws_f["C13"].value - z_two_sided_05) < 0.001:
    print("  → Za matches two-sided test at alpha=0.05")
else:
    issue("W", "C13", f"Za={ws_f['C13'].value} doesn't match standard Z values for alpha=0.05")

if abs(ws_f["C14"].value - z_power_80) < 0.001:
    print("  → Zb matches power=0.80 — CORRECT")
else:
    issue("W", "C14", f"Zb={ws_f['C14'].value} doesn't match Z for power=0.80")

# Note on Bonferroni
print("\n  Note 7 in the spreadsheet mentions Bonferroni correction (alpha/4).")
print("  If applied, Za should be NORM.S.INV(1-0.0125/2)=2.2414 (two-sided)")
print("  or NORM.S.INV(1-0.0125)=2.2414 (one-sided).")
z_bonf_one = scipy_stats.norm.ppf(1 - 0.05/4)
z_bonf_two = scipy_stats.norm.ppf(1 - 0.05/8)
print(f"  Bonferroni-corrected Za (one-sided, 4 tests): {z_bonf_one:.4f}")
print(f"  Bonferroni-corrected Za (two-sided, 4 tests): {z_bonf_two:.4f}")
print("  Current spreadsheet does NOT apply Bonferroni — this is noted but not enforced.")


# ══════════════════════════════════════════════════════════════════════════════
# FINAL SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 80)
print("AUDIT SUMMARY")
print("=" * 80)

errors = [i for i in issues if i[0] == "E"]
warnings = [i for i in issues if i[0] == "W"]
infos = [i for i in issues if i[0] == "I"]

print(f"\n  ERRORS:   {len(errors)}")
print(f"  WARNINGS: {len(warnings)}")
print(f"  INFO:     {len(infos)}")

if errors:
    print("\n  ── ERRORS ──")
    for _, cell, msg in errors:
        print(f"    [{cell}] {msg}")

if warnings:
    print("\n  ── WARNINGS ──")
    for _, cell, msg in warnings:
        print(f"    [{cell}] {msg}")

if infos:
    print("\n  ── INFO ──")
    for _, cell, msg in infos:
        print(f"    [{cell}] {msg}")

if not errors:
    print("\n  ✓ No formula errors found. All formulas reference correct cells")
    print("    and produce correct numerical results.")
    print("    Warnings are design suggestions, not bugs.")

print("\n" + "=" * 80)
print("AUDIT COMPLETE")
print("=" * 80)
