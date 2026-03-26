"""
build_campaign_tracker.py
Generates campaign_tracker.xlsx from the Cards Pod campaign tracker data.
"""

import os
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Border, Side, Alignment
)
from openpyxl.utils import get_column_letter

# ── Output path ──────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
out_path = os.path.join(SCRIPT_DIR, "campaign_tracker.xlsx")

# ── Style constants ───────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="4472C4")
HEADER_FONT   = Font(bold=True, color="FFFFFF")
ALT_FILL      = PatternFill("solid", fgColor="F2F2F2")
NO_FILL       = PatternFill("none")
TBD_FILL      = PatternFill("solid", fgColor="C6EFCE")

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Column headers ────────────────────────────────────────────────────────────
COLUMNS = [
    "LOB",
    "Mnemonic",
    "Description",
    "Rationale (Success Metric)",
    "Schedule",
    "Important Set (Control Structure)",
    "Typical Monthly Volume",
    "Response %",
    "Measurement Window (Days)",
    "Targeting",
    "Channels",
    "Results / Status",
    "Dashboard Link",
    "Source Table",
]

# ── Campaign rows ─────────────────────────────────────────────────────────────
ROWS = [
    # PCQ
    [
        "Cards",
        "PCQ",
        "Credit Card Acquisition — Third Party Acquisition (TPA/ITA) for existing RBC clients without a credit card",
        "Credit card application approved (app_approved); application completed with period-specific ASC (associates offer rate, terms, points, fee waiver)",
        "Monthly (major deployments ~biweekly; ~500K–626K clients per wave)",
        "No formal control group (BAU vs. challenger segments: BAU/NextGen-A; Challenger vs. Champion; NextGen vs. Model)",
        "~530K–626K per major wave; ~2.5M YTD (Jan–Mar 2026 across all tactics)",
        "1.5% (app_completed); ~1.0% approval rate (EDA: app_approved_pct); historical DM lift: 0.8%",
        "60–90 days (dependent on TPA offer validity)",
        "PACS TPA list + all 10 deciles; model-scored; DM test expanding to 7th decile Apr 2026",
        "DM, DO (Display Offer — near 100% coverage), EM (~65%), IM/OLB (~60%), MB (~60%), IV (IVR ~24%), CC",
        "TBD",
        "TBD",
        "dw00_im.dl_mr_prod.cards_tpa_pcq_decision_resp",
    ],
    # PCL / PLI
    [
        "Cards",
        "PCL / PLI",
        "Proactive Credit Limit Increase — pre-approved credit limit increase for existing cardholders",
        "Credit limit increase accepted (responder_cli); decisioned account",
        "Twice a month",
        "No explicit holdout in Confluence tracker; FY2025 test structure: Challenger / Sole-only Mobile eligible / Champion (30% / 40% / 30% sub-splits)",
        "~480K/month (FY2025 avg); PLI table: 20.4M rows (historical)",
        "16–18% (Confluence: 16%; EDA: 17.3%; FY2025 testing approach: 18% RR)",
        "60 days",
        "Top deciles; propensity model; mobile-eligible segment flagged separately; channels: Sales Model (30%), Mobile Dashboard (40%), Product Page (30%)",
        "CC (100%), DO (100%), IN (100%), OLB; RD not used",
        "TBD",
        "TBD",
        "dw00_im.dl_mr_prod.cards_pli_decision_resp (alias DL_MR_PROD.CARDS_PLI_DECISION_RESP)",
    ],
    # PCD
    [
        "Cards",
        "PCD",
        "Product Card Upgrade — NBO offer for existing cardholders to upgrade to a higher-tier card (e.g., RBC ION+ Visa)",
        "Product upgrade completed; card upgraded (resp_target_pct); responder flag",
        "Active, always-on (periodicity TBD — ongoing decision/response table suggests continuous)",
        "No Control noted in Confluence; Act/Control seg: Action = 77.69%; invitation_to_upgrade = N for 95.65% (non-upgrade segment dominates)",
        "~17M rows in PCD Ongoing table (historical); typical monthly deployment volume TBD",
        "2.0% (resp_any_pct per EDA); 1.6% resp_target_pct",
        "TBD (EDA shows response_start: 2022–2026)",
        "All segments; credit phase Emerging Prospects (34.5%) and Builders (30.9%) most common; product_at_decision top = RVC (24.3%); OLB and mobile eligible",
        "CC (~77%), EM (~56%), OLB (45%), Mobile (~79%), async banner (4 promo names confirmed: SalesModel_JVP, SalesModel_JAV, PPCN, Offer_Hub_Banner)",
        "Async banner pending (confirmed Mar 2026)",
        "TBD",
        "dw00_jm.dl_mr_prod.cards_pcd_ongoing_decis_resp",
    ],
    # AUH
    [
        "Cards",
        "AUH",
        "Authorized User Acquisition — targeting existing cardholders to add an authorized user (Phase 2: Rewards Cards)",
        "Authorized user card opened; relationship_cd = 'Z', card_sts in ('A', '') in D3CV12A.ACCT_CRD_OWN_DLY_DELTA",
        "Phase 1: complete. Phase 2 launch: 2026-04-30; monitored bi-weekly for 30-day window",
        "70% test / 30% control (final design); two-arm: Arm 1 (Comm+Offer) 50% vs Arm 2 (Comm Only) 50%; each arm: 70T/30C; total pop 539,620",
        "539,620 (Phase 2 total); sub-segments: Web 64.7K, Model 63.2K, Random 141.9K (per arm)",
        "0.045–0.120% (control RR by segment); lift baseline ~0.06% overall; treatment ~0.104%",
        "TBD (Phase 1 used similar window; Phase 2 30-day bi-weekly monitoring)",
        "Web segment (GA4 activity, 24%); Model segment (propensity, 23%); Random (53%); Rewards cards: ADP, GPR, GCP, MC2",
        "Email + OLB banner; mobile banner CANNOT be fulfilled",
        "Phase 1 complete; Phase 2 DOE finalized 2026-03-12; results pending launch",
        "TBD",
        "DG6V01.CLNT_DERIV_DTA_HIST; D3CV12A.ACCT_CRD_OWN_DLY_DELTA (response); DTZV01.VENDOR_FEEDBACK_MASTER (email)",
    ],
    # CLI
    [
        "Cards",
        "CLI",
        "Credit Limit Increase — always-on campaign; Priority #1 for Cards pod",
        "Credit limit increase actioned",
        "Active, always-on",
        "TBD — tactic EDA shows MGA_C and MGA_A segments (likely control/action groups)",
        "TBD (tactic data spans Feb–Mar 2026 in EDA; lead_resp counts in range 37K–202K visible)",
        "TBD",
        "TBD",
        "TBD — CLI tactic codes observed: CLI, CLO, NCT, MCP, VPN product filters in EDA queries",
        "TBD",
        "TBD — EDA in progress",
        "TBD",
        "DTZV01.TACTIC_EVNT_IP_AR_H60M (tactic events); response table TBD",
    ],
    # IPC
    [
        "Payments",
        "IPC",
        "International Money Transfer — Proactive Campaign; outbound targeting of clients for IMT offers",
        "IMT event completed (activity 031); conversion to international money transfer",
        "~2x/year (observed: 2025-05 and 2025-12 cohorts)",
        "~95% test (TG4) / ~5% control (TG7); 12 report groups (PIPCAG01–12) across segments",
        "~1.5M total clients across all segments (2025-12 TG4 sum); ranges from 479 (smallest segment) to 641K (largest)",
        "0.18%–1.69% (varies by segment and window); 61–73 day windows",
        "61–73 days (61 days for PIPCAG04–12; 73 days for PIPCAG01–03)",
        "12 segments (PIPCAG01–12); some segments email-enabled, some OLB/mobile only",
        "Email (select segments, ~60–67% open rate), OLB, Mobile; control receives no email",
        "Results available via imt_pipeline.py output; see ipc_iri_results.md",
        "TBD",
        "DTZTA_T_TACTIC_EVNT_HIST (Teradata); response via Teradata activity tables",
    ],
    # IRI
    [
        "Payments",
        "IRI",
        "International Money Transfer — Reactive/Trigger Campaign; triggered by client behavior",
        "IMT event completed (activity 031)",
        "Monthly",
        "~90% test (TG4) / ~10% control (TG7); single report group PIRIAG01",
        "76K–95K/month (alternates; 76K in shorter months, 95K in others)",
        "0.29%–0.42% (TG4 test); 0.14%–0.39% (TG7 control); minimal observed lift",
        "90 days",
        "Single segment; behavioral trigger-based targeting",
        "Email (~70K/month, ~60–68% open rate, declining trend), OLB, Mobile; control receives no email",
        "Monthly results available Jan 2025–Feb 2026; see ipc_iri_results.md; observed lift minimal",
        "TBD",
        "DTZTA_T_TACTIC_EVNT_HIST (Teradata)",
    ],
    # FHA
    [
        "HEF",
        "FHA",
        "First-time Home Buyer Acquisition",
        "Mortgage Funded",
        "Once every week",
        "5% test / 5% control",
        "20–5,000",
        "0.8%",
        "180",
        "All 10 Deciles",
        "HE, BU",
        "Without a monthly report. Refer to dashboard for details.",
        "Home Equity Financing — Campaign Results; Tableau Server (filter by MNE)",
        "N/A",
    ],
    # MSW
    [
        "HEF",
        "MSW",
        "Mortgage Switch",
        "Mortgage Funded",
        "Once every month",
        "50% test / 50% control",
        "Prior to expansion: 275,000–30,000; After expansion: 275,000–530,000",
        "2.4%",
        "365",
        "Top 5 Deciles",
        "HE, BU, WG, FG",
        "Without a monthly report. Refer to dashboard for details.",
        "Home Equity Financing — Campaign Results; Tableau Server (filter by MNE)",
        "N/A",
    ],
    # HMO
    [
        "HEF",
        "HMO",
        "New Mortgage Opportunity",
        "Mortgage Funded",
        "Once every month",
        "50% test / 50% control",
        "Prior to expansion: 10,000–16,000; After expansion: 220,000–370,000",
        "1.3%",
        "180",
        "Top 5 Percentiles",
        "HE, BU, WG",
        "Without a monthly report. Refer to dashboard for details.",
        "Home Equity Financing — Campaign Results; Tableau Server (filter by MNE)",
        "N/A",
    ],
    # MCE
    [
        "FSI",
        "MCE",
        "Investment Advice",
        "Funds converted to long-term investment",
        "Once every month",
        "50% test / 5% control",
        "60–120 days (ETI needs more time; extended window required)",
        "10%",
        "60",
        "TBD",
        "TBD",
        "Refer to dashboard",
        "Dashboard link (Confluence: 'Click here for dashboard...')",
        "dl_mr_test.area_report_el_monthly",
    ],
]

# ── Gaps data (sheet 2) ───────────────────────────────────────────────────────
GAPS = [
    ("PCQ", [
        "Dashboard Link: No dashboard URL captured; listed as TBD in original tracker",
        "Results: Confluence says 'Click here to expand' — no actual result summary captured",
        "Channels: Confluence notes 'TBD [TBD, TBD, TBD]' — from artifacts we know DO, DM, EM, IM, MB, IV, CC, RD are deployed; full channel priority order not formalized",
        "Control structure: No holdout control; DM test (7th decile expansion) control % options are 20% or 50% — not yet finalized as of Mar 24, 2026",
        "Response % by channel: DM lift known at 0.8% historically; overall response by channel breakdown not in files",
        "Schedule precision: Major waves appear biweekly but schedule is not explicitly stated as such in any source",
    ]),
    ("PCL / PLI", [
        "Dashboard Link: Not present in any file",
        "Results: No results summary in files",
        "Important Set: FY2025 testing approach doc shows Challenger/Champion/Mobile-eligible split but this is experiment-specific; standing BAU control % not confirmed",
        "Measurement Window: Confluence says 60 days; FY2025 testing approach shows 18% RR over unspecified window — verify alignment",
        "Targeting: Decile breakdown not specified; MDE calculator uses Work/Model/Random segments; top-decile filtering not confirmed",
        "Channels: CC, DO, IN confirmed always-on; email and mobile present but coverage % not stated",
        "Monthly Volume: 480K cited in PLI testing approach as FY2025 avg; current (2026) volume not confirmed",
    ]),
    ("PCD", [
        "Schedule: Exact deployment frequency not found; 'ongoing' decision/response table suggests continuous",
        "Dashboard Link: Not found",
        "Results: Async banner integration confirmed Mar 2026 but no outcome data",
        "Important Set / Control %: No control holdout documented; PCD-NBO is single mnemonic (100%); test group structure not found",
        "Measurement Window: Not found in any file; EDA shows response dates spanning 2022–2026 but no stated window",
        "Typical Monthly Volume: Historical table has 17M rows but current monthly deployment counts not found",
    ]),
    ("AUH", [
        "Measurement Window: Phase 2 has bi-weekly monitoring at 30 days; final declared measurement window not stated",
        "Dashboard Link: Not found",
        "Results: Phase 1 results not captured in these files; Phase 2 pending (launch Apr 30, 2026)",
        "Typical Monthly Volume: Phase 2 total pop is 539,620 (one-time experiment, not recurring monthly); recurring schedule post-experiment not defined",
        "Schedule (post-experiment): Phase 1 complete; Phase 2 is a structured experiment — standing deployment frequency not defined",
        "Channels (full list): Email + OLB banner confirmed; CC/branch confirmed not in scope; mobile banner explicitly excluded",
    ]),
    ("CLI", [
        "ALL columns are substantially incomplete. CLI is listed as Priority #1 but artifacts are limited to a low-resolution tactic EDA spreadsheet.",
        "Missing: Monthly volume, response %, measurement window, control structure, targeting criteria, channels, dashboard, results",
        "Partial data: Tactic events sourced from DTZV01.TACTIC_EVNT_IP_AR_H60M; MGA_C and MGA_A are likely test group codes; date range Feb–Mar 2026 visible; product filters include CLI, CLO, NCT, MCP, VPN",
    ]),
    ("IPC", [
        "Schedule: Only 2 cohorts visible (2025-05 and 2025-12) — appears semi-annual but not confirmed",
        "Dashboard Link: Not found",
        "Description detail: 'Proactive' campaign — full targeting criteria and eligibility rules not in files",
        "Targeting: 12 segment definitions (PIPCAG01–12) — the criteria distinguishing each segment not documented in these files",
        "Source table name: EDW Teradata table name for response not confirmed beyond DTZTA_T_TACTIC_EVNT_HIST",
    ]),
    ("IRI", [
        "Dashboard Link: Not found",
        "Description detail: Trigger/reactive mechanism not fully defined — what behavioral trigger fires the campaign?",
        "Targeting: Trigger criteria not documented",
        "Lift significance: Observed TG4 vs TG7 rates are very close (e.g., Jan 2025: 0.40% vs 0.39%) — no statistical test result on file confirming significance or lack thereof",
    ]),
    ("FHA", [
        "Channels: 'TBD [TBD, TBD, TBD]' — channel list not resolved in original Confluence tracker; HE and BU are noted in the original tracker",
        "Dashboard Link: Tableau Server reference exists but no URL",
        "Source Table: Listed as N/A in original tracker — actual underlying table not known",
    ]),
    ("MSW", [
        "Channels: HE, BU, WG, FG noted; no further detail",
        "Dashboard Link: Tableau Server reference exists but no URL",
        "Volume note: Two ranges (pre/post expansion) — expansion date not documented",
        "Source Table: N/A in original tracker",
    ]),
    ("HMO", [
        "Targeting: 'Top 5 Percentiles' — unusual phrasing (percentiles vs. deciles); not clarified in files",
        "Channels: HE, BU, WG noted; no further detail",
        "Dashboard Link: Tableau Server reference exists but no URL",
        "Source Table: N/A in original tracker",
    ]),
    ("MCE", [
        "Channels: Not found in any file; original tracker left blank",
        "Targeting: Not documented in any file; original tracker left blank",
        "Important Set: 50% test / 5% control is unusual asymmetry — confirm this is intentional",
        "Volume: '60–120 days ETI' appears in the Volume column of the original tracker, which seems to be a measurement note rather than a volume figure — actual monthly volume unknown",
        "Dashboard Link: URL not captured (Confluence has a placeholder link)",
    ]),
]

TBD_TOKENS = {"tbd", "unknown", "n/a"}


def is_tbd(value):
    """Return True if the cell value is TBD, Unknown, or N/A (case-insensitive)."""
    if value is None:
        return False
    return str(value).strip().lower() in TBD_TOKENS


def apply_border(cell):
    cell.border = BORDER


def set_column_widths(ws, max_width=40):
    """Auto-size columns, capped at max_width; enable wrap_text for capped cols."""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                # measure each line in the cell (wrap already applied)
                cell_len = max(len(str(line)) for line in str(cell.value).splitlines())
                max_len = max(max_len, cell_len)
        width = min(max_len + 2, max_width)
        ws.column_dimensions[col_letter].width = width


# ─────────────────────────────────────────────────────────────────────────────
# Build workbook
# ─────────────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Sheet 1: Campaign Tracker ─────────────────────────────────────────────────
ws = wb.active
ws.title = "Campaign Tracker"

# Header row
for col_idx, col_name in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
    cell.alignment = HEADER_ALIGN

# Auto-filter on header
ws.auto_filter.ref = ws.dimensions  # will be updated after data is written

# Freeze top row
ws.freeze_panes = "A2"

# Data rows
for row_idx, row_data in enumerate(ROWS, start=2):
    use_alt = (row_idx % 2 == 0)
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = BORDER
        cell.alignment = WRAP
        # TBD highlight takes priority; otherwise alternating shading
        if is_tbd(value):
            cell.fill = TBD_FILL
        elif use_alt:
            cell.fill = ALT_FILL
        else:
            cell.fill = NO_FILL

# Now set auto_filter to the full data range
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

# Column widths
set_column_widths(ws, max_width=40)

# Row heights — give data rows a bit of breathing room
for row_idx in range(2, len(ROWS) + 2):
    ws.row_dimensions[row_idx].height = 60  # points; enough for 3–4 wrapped lines

# ── Sheet 2: Gaps ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Gaps")

gap_headers = ["Campaign", "Gap Description"]
for col_idx, h in enumerate(gap_headers, start=1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
    cell.alignment = HEADER_ALIGN

ws2.auto_filter.ref = "A1:B1"
ws2.freeze_panes = "A2"

gap_row = 2
for campaign, gaps in GAPS:
    use_alt_block = (gap_row % 2 == 0)
    for gap_text in gaps:
        cell_c = ws2.cell(row=gap_row, column=1, value=campaign)
        cell_g = ws2.cell(row=gap_row, column=2, value=gap_text)
        for cell in (cell_c, cell_g):
            cell.border = BORDER
            cell.alignment = WRAP
            if use_alt_block:
                cell.fill = ALT_FILL
            else:
                cell.fill = NO_FILL
        gap_row += 1
        use_alt_block = not use_alt_block  # alternate within each campaign block

# Column widths for gaps sheet
ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 80
for row_idx in range(2, gap_row):
    ws2.row_dimensions[row_idx].height = 45

# ─────────────────────────────────────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────────────────────────────────────
wb.save(out_path)
print(f"Saved: {out_path}")

# ── Force recalculation via win32com (optional) ───────────────────────────────
try:
    import win32com.client
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    wb_com = xl.Workbooks.Open(os.path.abspath(out_path))
    wb_com.Save()
    wb_com.Close()
    xl.Quit()
    print("win32com recalculation complete.")
except Exception as e:
    print(f"win32com not available or failed ({e}); skipping recalculation.")
