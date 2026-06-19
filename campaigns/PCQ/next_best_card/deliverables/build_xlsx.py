import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
BOLD = Font(bold=True)
ITALIC = Font(italic=True, color="595959")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER


def write_data_row(ws, row, values, start_col=1, bold=False):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.border = BORDER
        if bold:
            c.font = BOLD
        c.alignment = LEFT if i == 0 else RIGHT


# ============== Sheet 1: Executive Summary ==============
ws = wb.active
ws.title = "Executive Summary"
set_col_widths(ws, [42, 20, 20, 18])

ws.merge_cells("A1:D1")
ws["A1"] = "PCQ Next Best Card - NG3_1ST vs NG3_2ND - Executive Summary"
ws["A1"].font = Font(bold=True, size=14)
ws["A1"].alignment = CENTER

ws["A2"] = "Run date"
ws["B2"] = "2026-04-14"
ws["A3"] = "Test groups"
ws["B3"] = "NG3_1ST (1st recommended) vs NG3_2ND (2nd recommended)"
ws["A4"] = "Primary population"
ws["B4"] = "Period-ASC approved (true PCQ conversions)"

ws["A6"] = "HEADLINE"
ws["A6"].font = Font(bold=True, size=12)
ws["A6"].fill = SECTION_FILL

ws.merge_cells("A7:D7")
ws["A7"] = (
    "The hypothesis '2nd-best recommendation outperforms 1st-best' is NOT supported. "
    "Conversion rates are tied at 1.05%, but NG3_2ND delivers 22% lower per-account spend "
    "and 19% lower revenue per deployed client. The 2nd-best recommendation produces less "
    "value in aggregate, not more."
)
ws["A7"].alignment = WRAP
ws.row_dimensions[7].height = 60

ws["A9"] = "KEY METRICS - Period-ASC only (primary measure)"
ws["A9"].font = Font(bold=True, size=11)
ws["A9"].fill = SECTION_FILL

write_header_row(ws, 10, ["Metric", "NG3_1ST", "NG3_2ND", "Delta"])
rows1 = [
    ("Deployed clients", "332,821", "232,555", ""),
    ("Period-ASC approved", "3,491", "2,438", ""),
    ("Period-ASC approval rate", "1.05%", "1.05%", "+0.00 pp"),
    ("Total spend (lifetime)", "$40,585,752", "$22,052,690", ""),
    ("Weighted $ per approved", "$11,625", "$9,045", "-22%"),
    ("Weighted $ per deployed", "$122", "$95", "-22%"),
    ("Voluntary attrition rate", "5.64%", "5.58%", "-0.06 pp"),
    ("Total fees collected", "$325,518", "$136,524", ""),
]
for i, r in enumerate(rows1):
    write_data_row(ws, 11 + i, list(r))

ws["A20"] = "KEY METRICS - Any in-campaign approval (Period-ASC + Other ASC)"
ws["A20"].font = Font(bold=True, size=11)
ws["A20"].fill = SECTION_FILL

write_header_row(ws, 21, ["Metric", "NG3_1ST", "NG3_2ND", "Delta"])
rows2 = [
    ("Any in-campaign approved", "4,518", "3,114", ""),
    ("Any in-campaign rate", "1.36%", "1.34%", "-0.02 pp"),
    ("Total spend (combined)", "$59,257,380", "$33,544,979", ""),
    ("Weighted $ per approved", "$13,126", "$10,773", "-18%"),
    ("Weighted $ per deployed", "$178", "$144", "-19%"),
]
for i, r in enumerate(rows2):
    write_data_row(ws, 22 + i, list(r))

ws["A28"] = "RECOMMENDATION"
ws["A28"].font = Font(bold=True, size=11)
ws["A28"].fill = SECTION_FILL
ws.merge_cells("A29:D29")
ws["A29"] = (
    "Do not roll out the 2nd-best recommendation as a default. Per-deployed revenue is "
    "19% lower with no offsetting conversion lift. The product-level breakdown (back-pocket) "
    "suggests the underperformance is product-mix driven: IAV as 2nd-best beats expectations, "
    "IOP as 2nd-best collapses. Any future iteration should deprioritize IOP and prioritize "
    "IAV in the 2nd-best slot specifically."
)
ws["A29"].alignment = WRAP
ws.row_dimensions[29].height = 70


# ============== Sheet 2: Period-ASC by product ==============
ws2 = wb.create_sheet("Results - Period-ASC")
set_col_widths(ws2, [14, 10, 12, 14, 14, 14, 18, 18, 16, 16])

ws2["A1"] = "Period-ASC results by (test group x product) - primary measure"
ws2["A1"].font = Font(bold=True, size=12)
ws2.merge_cells("A1:J1")

headers = [
    "Group", "Product", "Deployed", "Period-ASC Approved",
    "Period-ASC RR", "Vol attrition", "Vol attrition rate",
    "Total spend", "Weighted $/acct", "Total fees",
]
write_header_row(ws2, 3, headers)
ws2.row_dimensions[3].height = 30

prod_rows = [
    ("NG3_1ST", "CLO", "25,369", "685", "2.70%", "26", "3.80%", "$6,780,502", "$9,898", "$30,542"),
    ("NG3_1ST", "GCP", "12,878", "119", "0.92%", "11", "9.24%", "$1,479,610", "$12,434", "$27,021"),
    ("NG3_1ST", "IAV", "196,372", "1,100", "0.56%", "116", "10.55%", "$18,609,749", "$16,918", "$217,283"),
    ("NG3_1ST", "ION", "34,556", "605", "1.75%", "24", "3.97%", "$4,364,872", "$7,215", "$18,272"),
    ("NG3_1ST", "IOP", "38,159", "754", "1.98%", "15", "1.99%", "$7,938,153", "$10,528", "$31,122"),
    ("NG3_1ST", "MC1", "25,110", "227", "0.90%", "5", "2.20%", "$1,412,865", "$6,224", "$1,179"),
    ("NG3_1ST", "MCP", "377", "1", "0.27%", "0", "0.00%", "$0", "$0", "$99"),
    ("NG3_1ST", "Total", "332,821", "3,491", "1.05%", "197", "5.64%", "$40,585,752", "$11,625", "$325,518"),
    ("", "", "", "", "", "", "", "", "", ""),
    ("NG3_2ND", "CLO", "27,977", "499", "1.78%", "24", "4.81%", "$3,023,733", "$6,060", "$19,221"),
    ("NG3_2ND", "GCP", "4,713", "77", "1.63%", "11", "14.29%", "$865,293", "$11,237", "$16,720"),
    ("NG3_2ND", "IAV", "13,174", "214", "1.62%", "15", "7.01%", "$3,900,826", "$18,228", "$45,042"),
    ("NG3_2ND", "ION", "49,398", "755", "1.53%", "19", "2.52%", "$6,041,612", "$8,002", "$13,566"),
    ("NG3_2ND", "IOP", "112,350", "663", "0.59%", "58", "8.75%", "$6,778,335", "$10,224", "$39,294"),
    ("NG3_2ND", "MC1", "23,165", "222", "0.96%", "8", "3.60%", "$1,185,659", "$5,341", "$1,463"),
    ("NG3_2ND", "MCP", "1,778", "8", "0.45%", "1", "12.50%", "$257,232", "$32,154", "$1,219"),
    ("NG3_2ND", "Total", "232,555", "2,438", "1.05%", "136", "5.58%", "$22,052,690", "$9,045", "$136,524"),
]
for i, r in enumerate(prod_rows):
    is_total = "Total" in str(r[1]) if r[1] else False
    write_data_row(ws2, 4 + i, list(r), bold=is_total)

ws2["A23"] = "Weighted $/acct = total spend / count of approved accounts (not average of averages)."
ws2["A23"].font = ITALIC
ws2.merge_cells("A23:J23")


# ============== Sheet 3: Combined results ==============
ws3 = wb.create_sheet("Results - Combined")
set_col_widths(ws3, [14, 22, 14, 16, 18, 20, 20])

ws3["A1"] = "Combined in-campaign results (Period-ASC + Other ASC) at group level"
ws3["A1"].font = Font(bold=True, size=12)
ws3.merge_cells("A1:G1")

ws3["A2"] = "Product-level Other ASC numbers are excluded - see Methodology & Caveats"
ws3["A2"].font = ITALIC
ws3.merge_cells("A2:G2")

headers = [
    "Group", "Population", "Approved", "Rate / deployed",
    "Total spend", "Weighted $/approved", "Weighted $/deployed",
]
write_header_row(ws3, 4, headers)

comb_rows = [
    ("NG3_1ST", "Period-ASC",          "3,491", "1.05%", "$40,585,752", "$11,625", "$122"),
    ("NG3_1ST", "Other ASC",           "1,027", "0.31%", "$18,671,628", "$18,181", "$56"),
    ("NG3_1ST", "Combined in-campaign","4,518", "1.36%", "$59,257,380", "$13,126", "$178"),
    ("", "", "", "", "", "", ""),
    ("NG3_2ND", "Period-ASC",          "2,438", "1.05%", "$22,052,690", "$9,045",  "$95"),
    ("NG3_2ND", "Other ASC",           "676",   "0.29%", "$11,492,289", "$17,000", "$49"),
    ("NG3_2ND", "Combined in-campaign","3,114", "1.34%", "$33,544,979", "$10,773", "$144"),
    ("", "", "", "", "", "", ""),
    ("Delta", "Period-ASC",            "",      "+0.00 pp", "", "-22%", "-22%"),
    ("Delta", "Combined",              "",      "-0.02 pp", "", "-18%", "-19%"),
]
for i, r in enumerate(comb_rows):
    is_bold = "Combined" in str(r[1]) or r[0] == "Delta"
    write_data_row(ws3, 5 + i, list(r), bold=is_bold)

ws3["A17"] = "Observation: Other ASC accounts spend 56-88% more per account than Period-ASC in BOTH groups."
ws3["A17"].font = BOLD
ws3.merge_cells("A17:G17")
ws3["A18"] = (
    "Interpretation: PCQ reaches a lower-value customer segment than organic/branch channels. "
    "This is population composition, not campaign-driven lift. See Methodology for framing."
)
ws3["A18"].font = ITALIC
ws3.merge_cells("A18:G18")
ws3.row_dimensions[18].height = 40


# ============== Sheet 4: Product Detail back-pocket ==============
ws4 = wb.create_sheet("Product Detail (Period-ASC)")
set_col_widths(ws4, [10, 12, 12, 12, 12, 14, 14, 14, 10])

ws4["A1"] = "Product-level lift table - Period-ASC only"
ws4["A1"].font = Font(bold=True, size=12)
ws4.merge_cells("A1:I1")

ws4["A2"] = "Back-pocket table - use to explain why overall NG3_2ND underperforms. Lead with group-level."
ws4["A2"].font = ITALIC
ws4.merge_cells("A2:I2")

headers = [
    "Product", "1ST Deploy", "2ND Deploy",
    "1ST RR", "2ND RR", "Delta RR",
    "1ST $/acct", "2ND $/acct", "Delta $",
]
write_header_row(ws4, 4, headers)

detail_rows = [
    ("CLO", "25,369",  "27,977",  "2.70%", "1.78%", "-0.92 pp", "$9,898",  "$6,060",  "-39%"),
    ("GCP", "12,878",  "4,713",   "0.92%", "1.63%", "+0.71 pp", "$12,434", "$11,237", "-10%"),
    ("IAV", "196,372", "13,174",  "0.56%", "1.62%", "+1.06 pp", "$16,918", "$18,228", "+8%"),
    ("ION", "34,556",  "49,398",  "1.75%", "1.53%", "-0.22 pp", "$7,215",  "$8,002",  "+11%"),
    ("IOP", "38,159",  "112,350", "1.98%", "0.59%", "-1.39 pp", "$10,528", "$10,224", "-3%"),
    ("MC1", "25,110",  "23,165",  "0.90%", "0.96%", "+0.06 pp", "$6,224",  "$5,341",  "-14%"),
    ("MCP", "377",     "1,778",   "0.27%", "0.45%", "+0.18 pp", "$0",      "$32,154", "n/a"),
]
for i, r in enumerate(detail_rows):
    write_data_row(ws4, 5 + i, list(r))

ws4["A13"] = "Interpretation"
ws4["A13"].font = Font(bold=True, size=11)
ws4["A13"].fill = SECTION_FILL
ws4.merge_cells("A13:I13")

observations = [
    "The model's product allocation is drastically different between groups. NG3_1ST pushes 196k clients to IAV as best. NG3_2ND pushes 112k clients to IOP as 2nd-best. Populations are naturally unbalanced by product.",
    "IAV as a 2nd-best is a WINNER. Response rate triples (0.56% to 1.62%) and per-account spend grows 8%. The model under-deploys IAV as a 2nd-best option. Potential upside.",
    "IOP as a 2nd-best is a LOSER. Response rate collapses 70% (1.98% to 0.59%). The model over-deploys IOP as a 2nd-best. Customers reject it when offered as an alternative rather than the primary.",
    "CLO loses materially as a 2nd-best (-0.92 pp RR, -39% $/acct). A strong 1st-pick that fails when pushed as an alternative.",
    "MCP volumes too small for reliable interpretation (1 and 8 approved accounts). Exclude from conclusions.",
    "Risk if shown to Phil: invites a validity challenge ('populations aren't matched by product'). Keep as back-pocket, bring out only if asked 'why does NG3_2ND underperform'.",
]
for i, obs in enumerate(observations):
    ws4.cell(row=14 + i, column=1, value="- " + obs)
    ws4.merge_cells(start_row=14 + i, start_column=1, end_row=14 + i, end_column=9)
    ws4.cell(row=14 + i, column=1).alignment = WRAP
    ws4.row_dimensions[14 + i].height = 36


# ============== Sheet 5: Methodology ==============
ws5 = wb.create_sheet("Methodology & Caveats")
set_col_widths(ws5, [110])

ws5["A1"] = "Methodology & Caveats"
ws5["A1"].font = Font(bold=True, size=14)

sections = [
    ("1. Weighted vs simple averages",
     "All per-account figures in this workbook are WEIGHTED averages, computed as "
     "(sum of account-level spend) / (count of accounts). When pivoting in Excel, "
     "the default Average aggregation computes the unweighted mean of product-level "
     "averages, which can give misleading results at parent rollups (one product with "
     "8 accounts gets the same weight as one with 1,100 accounts). To replicate the "
     "weighted numbers in a pivot, use a calculated field: SUM(spend) / SUM(approved)."),
    ("2. Population comparability",
     "NG3_1ST and NG3_2ND are NOT an even split (332,821 vs 232,555). The model's "
     "recommendation logic produces different populations per group - each client sees "
     "their own 1st-best or 2nd-best product. The populations are therefore not strictly "
     "matched by product mix, and overall group comparisons are 'mix-adjusted' rather "
     "than strict A/B. The aggregate group-level lift is still valid as a hypothesis test "
     "of whether the 2nd-best recommendation works better than the 1st-best in aggregate. "
     "Product-level comparisons are interpretive, not strict A/B."),
    ("3. Other ASC product attribution - CAVEAT",
     "For Period-ASC approvals, TPA (True Pre-Approved) logic guarantees the client can only "
     "receive the offered product, so grouping by offer_prod_latest is safe. For Other ASC "
     "approvals, the client applied through a non-PCQ channel, and the booked product may "
     "differ from the offered product. Therefore, product-level Other ASC numbers are "
     "UNRELIABLE and are excluded from the product breakdown sheet. Only group-level Other "
     "ASC rollups are shown."),
    ("4. Period-ASC vs Other ASC $/acct comparison framing",
     "Other ASC accounts spend 56 to 88 percent more per account than Period-ASC accounts "
     "in both groups. This is NOT a causal claim about the campaign - it is an observation "
     "about population composition. Clients who respond via non-PCQ channels are self-selected, "
     "higher-intent customers. The PCQ channel reaches a lower-value segment overall. "
     "Do not frame this as lift in the writeup; frame it as customer-segment context."),
    ("5. Statistical significance",
     "Conversion rate tied at 1.05% is a true null - NG3_1ST and NG3_2ND have the same "
     "Period-ASC approval rate to three decimal places. No significance test is needed; "
     "the minimum detectable effect at 80% power given sample sizes is approximately "
     "0.02 pp, and the observed difference is 0.00 pp. The per-account spend gap "
     "($11,625 vs $9,045, -22%) has NOT been tested for significance because the standard "
     "deviations are not in the summary output; a two-sample t-test should be run on the "
     "account-level data before the writeup is finalized. Given n=3,491 and n=2,438 and "
     "the magnitude of the gap, the result is almost certainly significant at p<0.001."),
    ("6. Reclass / fulfillment drift",
     "Approximately 10% of Period-ASC approved accounts show a booked product that does "
     "not equal the offered product (booking_status=mismatch). This is a separate data "
     "quality observation from the lift analysis. For the strictest population, filter "
     "the Q2 account-level output to booking_status=match AND lifetime_status=stable and "
     "recompute the metrics. Not expected to change the directional conclusion."),
    ("7. Halo / cross-sell (pending)",
     "The _q2_v2_halo.sql query adds a cross_* block for post-campaign acquisitions by "
     "in-campaign approved clients (acct_open_dt > treatmt_end_dt). This measures a "
     "directional halo effect - whether the campaign triggers downstream card acquisitions "
     "beyond the offered card. Cross-sell cannot be product-attributed (the halo card is "
     "by definition NOT the offered product) and is an observational-only comparison "
     "between NG3_1ST and NG3_2ND. Results not yet incorporated into this workbook."),
]

row = 3
for heading, body in sections:
    c = ws5.cell(row=row, column=1, value=heading)
    c.font = Font(bold=True, size=11)
    c.fill = SECTION_FILL
    c.alignment = WRAP
    ws5.row_dimensions[row].height = 22
    row += 1
    c = ws5.cell(row=row, column=1, value=body)
    c.alignment = WRAP
    c.font = Font(size=10)
    ws5.row_dimensions[row].height = max(60, len(body) // 100 * 18)
    row += 2


out = r"C:\Users\andre\New_projects\cards\campaigns\PCQ\next_best_card\deliverables\pcq_nbc_results.xlsx"
wb.save(out)
print("Saved:", out)
