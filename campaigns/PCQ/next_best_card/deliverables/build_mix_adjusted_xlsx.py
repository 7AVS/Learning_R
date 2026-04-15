"""
Build pcq_nbc_mix_adjusted.xlsx with LIVE Excel formulas.

Everything in the workbook that isn't an input is a formula. You can
click any computed cell and see how it's built. Modify an input count
or $/acct and the whole workbook recomputes.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill(start_color="0051A5", end_color="0051A5", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
RESULT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BOLD = Font(bold=True)
ITALIC = Font(italic=True, color="595959", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_INT = '#,##0'
FMT_USD = '"$"#,##0'
FMT_USD2 = '"$"#,##0.00'
FMT_PCT = '0.0%;-0.0%;0.0%'


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER


def cell(ws, row, col, value, fmt=None, bold=False, fill=None, align_right=True):
    c = ws.cell(row=row, column=col, value=value)
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    if bold:
        c.font = BOLD
    if fill:
        c.fill = fill
    c.alignment = RIGHT if align_right else LEFT
    return c


def text_cell(ws, row, col, value, bold=False, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.border = BORDER
    c.alignment = LEFT
    if bold:
        c.font = BOLD
    if fill:
        c.fill = fill
    return c


ws = wb.active
ws.title = "Mix-Adjusted Spend"
set_col_widths(ws, [44, 16, 16, 18, 18, 18, 18])

# ============== TITLE ==============
ws.merge_cells("A1:G1")
ws["A1"] = "PCQ Next Best Card — Mix-Adjusted Average Spend per Approved Account"
ws["A1"].font = Font(bold=True, size=13)
ws["A1"].alignment = CENTER

ws.merge_cells("A2:G2")
ws["A2"] = (
    "All computed cells are live Excel formulas. Modify any input in the "
    "INPUT DATA block below and every downstream calculation updates."
)
ws["A2"].font = ITALIC
ws["A2"].alignment = WRAP
ws.row_dimensions[2].height = 24

# ============== INPUT DATA ==============
row_input_header = 4
text_cell(ws, row_input_header, 1, "INPUT DATA — per product (edit these cells to recompute)",
          bold=True, fill=SECTION_FILL)
ws.cell(row=row_input_header, column=1).font = Font(bold=True, size=11)
ws.merge_cells(start_row=row_input_header, start_column=1, end_row=row_input_header, end_column=7)

row_input_cols = row_input_header + 1
write_header(ws, row_input_cols, [
    "Product",
    "1st approved",
    "2nd approved",
    "1st $/acct",
    "2nd $/acct",
    "Δ $ per product",
    "Δ % per product",
])

# Per-product raw data (Period-ASC only)
products = [
    # (name, 1st_approved, 2nd_approved, 1st_dpa, 2nd_dpa)
    ("CLO", 685, 499, 9898, 6060),
    ("GCP", 119, 77, 12434, 11237),
    ("IAV", 1100, 214, 16918, 18228),
    ("ION", 605, 755, 7215, 8002),
    ("IOP", 754, 663, 10528, 10224),
    ("MC1", 227, 222, 6224, 5341),
    ("MCP", 1, 8, 0, 32154),
]

DATA_START = row_input_cols + 1
DATA_END = DATA_START + len(products) - 1

for i, (p_name, app1, app2, dpa1, dpa2) in enumerate(products):
    r = DATA_START + i
    text_cell(ws, r, 1, p_name)
    cell(ws, r, 2, app1, fmt=FMT_INT, fill=INPUT_FILL)
    cell(ws, r, 3, app2, fmt=FMT_INT, fill=INPUT_FILL)
    cell(ws, r, 4, dpa1, fmt=FMT_USD, fill=INPUT_FILL)
    cell(ws, r, 5, dpa2, fmt=FMT_USD, fill=INPUT_FILL)
    # Delta $ and Delta % — formulas
    cell(ws, r, 6, f"=E{r}-D{r}", fmt=FMT_USD)
    cell(ws, r, 7, f"=IF(D{r}=0,\"\",(E{r}-D{r})/D{r})", fmt=FMT_PCT)

# Totals row — raw weighted averages from SUMPRODUCT
TOTAL_ROW = DATA_END + 1
text_cell(ws, TOTAL_ROW, 1, "Total (raw weighted average)", bold=True, fill=SECTION_FILL)
cell(ws, TOTAL_ROW, 2, f"=SUM(B{DATA_START}:B{DATA_END})", fmt=FMT_INT, bold=True, fill=SECTION_FILL)
cell(ws, TOTAL_ROW, 3, f"=SUM(C{DATA_START}:C{DATA_END})", fmt=FMT_INT, bold=True, fill=SECTION_FILL)
cell(ws, TOTAL_ROW, 4,
     f"=SUMPRODUCT(B{DATA_START}:B{DATA_END},D{DATA_START}:D{DATA_END})/B{TOTAL_ROW}",
     fmt=FMT_USD, bold=True, fill=SECTION_FILL)
cell(ws, TOTAL_ROW, 5,
     f"=SUMPRODUCT(C{DATA_START}:C{DATA_END},E{DATA_START}:E{DATA_END})/C{TOTAL_ROW}",
     fmt=FMT_USD, bold=True, fill=SECTION_FILL)
cell(ws, TOTAL_ROW, 6, f"=E{TOTAL_ROW}-D{TOTAL_ROW}", fmt=FMT_USD, bold=True, fill=SECTION_FILL)
cell(ws, TOTAL_ROW, 7, f"=(E{TOTAL_ROW}-D{TOTAL_ROW})/D{TOTAL_ROW}", fmt=FMT_PCT, bold=True, fill=SECTION_FILL)

# ============== MIX-ADJUSTMENT #1: 1st-best mix as reference ==============
MIX1_HEADER_ROW = TOTAL_ROW + 3
text_cell(ws, MIX1_HEADER_ROW, 1,
          "MIX-ADJUSTMENT #1 — reference = 1st-best product distribution",
          bold=True, fill=SECTION_FILL)
ws.cell(row=MIX1_HEADER_ROW, column=1).font = Font(bold=True, size=11)
ws.merge_cells(start_row=MIX1_HEADER_ROW, start_column=1, end_row=MIX1_HEADER_ROW, end_column=7)

MIX1_EXPLAIN_ROW = MIX1_HEADER_ROW + 1
c = ws.cell(row=MIX1_EXPLAIN_ROW, column=1,
            value=('Question: "If 2nd-best customers had the same product distribution as '
                   '1st-best customers, what would their average spend per account be?"'))
c.font = ITALIC
c.alignment = WRAP
ws.merge_cells(start_row=MIX1_EXPLAIN_ROW, start_column=1, end_row=MIX1_EXPLAIN_ROW, end_column=7)
ws.row_dimensions[MIX1_EXPLAIN_ROW].height = 28

MIX1_COLS_ROW = MIX1_EXPLAIN_ROW + 1
write_header(ws, MIX1_COLS_ROW, [
    "Product",
    "1st-mix weight (ref)",
    "1st $/acct",
    "2nd $/acct",
    "weight × 1st $",
    "weight × 2nd $",
    "",
])

MIX1_DATA_START = MIX1_COLS_ROW + 1
MIX1_DATA_END = MIX1_DATA_START + len(products) - 1

for i in range(len(products)):
    r = MIX1_DATA_START + i
    input_row = DATA_START + i
    text_cell(ws, r, 1, f"=A{input_row}")
    cell(ws, r, 2, f"=B{input_row}", fmt=FMT_INT)  # 1st-mix weight pulled from input
    cell(ws, r, 3, f"=D{input_row}", fmt=FMT_USD)  # 1st $/acct
    cell(ws, r, 4, f"=E{input_row}", fmt=FMT_USD)  # 2nd $/acct
    cell(ws, r, 5, f"=B{r}*C{r}", fmt=FMT_USD)      # weight × 1st $
    cell(ws, r, 6, f"=B{r}*D{r}", fmt=FMT_USD)      # weight × 2nd $

# Sum row for mix-adj #1
MIX1_SUM_ROW = MIX1_DATA_END + 1
text_cell(ws, MIX1_SUM_ROW, 1, "Sum", bold=True, fill=SECTION_FILL)
cell(ws, MIX1_SUM_ROW, 2, f"=SUM(B{MIX1_DATA_START}:B{MIX1_DATA_END})",
     fmt=FMT_INT, bold=True, fill=SECTION_FILL)
cell(ws, MIX1_SUM_ROW, 3, "—", bold=True, fill=SECTION_FILL)
cell(ws, MIX1_SUM_ROW, 4, "—", bold=True, fill=SECTION_FILL)
cell(ws, MIX1_SUM_ROW, 5, f"=SUM(E{MIX1_DATA_START}:E{MIX1_DATA_END})",
     fmt=FMT_USD, bold=True, fill=SECTION_FILL)
cell(ws, MIX1_SUM_ROW, 6, f"=SUM(F{MIX1_DATA_START}:F{MIX1_DATA_END})",
     fmt=FMT_USD, bold=True, fill=SECTION_FILL)

# Result block for mix-adj #1
MIX1_RESULT_LABEL = MIX1_SUM_ROW + 2
text_cell(ws, MIX1_RESULT_LABEL, 1, "RESULT — mix-adjustment #1", bold=True, fill=RESULT_FILL)
ws.merge_cells(start_row=MIX1_RESULT_LABEL, start_column=1, end_row=MIX1_RESULT_LABEL, end_column=7)

MIX1_RESULT_HEADER = MIX1_RESULT_LABEL + 1
write_header(ws, MIX1_RESULT_HEADER, ["", "1st best", "2nd best", "Δ $", "Δ %", "", ""])

# Raw row
MIX1_RAW_ROW = MIX1_RESULT_HEADER + 1
text_cell(ws, MIX1_RAW_ROW, 1, "Raw avg spend per approved")
cell(ws, MIX1_RAW_ROW, 2, f"=D{TOTAL_ROW}", fmt=FMT_USD)
cell(ws, MIX1_RAW_ROW, 3, f"=E{TOTAL_ROW}", fmt=FMT_USD)
cell(ws, MIX1_RAW_ROW, 4, f"=C{MIX1_RAW_ROW}-B{MIX1_RAW_ROW}", fmt=FMT_USD)
cell(ws, MIX1_RAW_ROW, 5, f"=(C{MIX1_RAW_ROW}-B{MIX1_RAW_ROW})/B{MIX1_RAW_ROW}", fmt=FMT_PCT)

# Mix-adjusted row
MIX1_ADJ_ROW = MIX1_RAW_ROW + 1
text_cell(ws, MIX1_ADJ_ROW, 1,
          "Mix-adjusted (2nd-best $ × 1st-best mix)", bold=True, fill=RESULT_FILL)
cell(ws, MIX1_ADJ_ROW, 2, f"=E{MIX1_SUM_ROW}/B{MIX1_SUM_ROW}",
     fmt=FMT_USD, bold=True, fill=RESULT_FILL)
cell(ws, MIX1_ADJ_ROW, 3, f"=F{MIX1_SUM_ROW}/B{MIX1_SUM_ROW}",
     fmt=FMT_USD, bold=True, fill=RESULT_FILL)
cell(ws, MIX1_ADJ_ROW, 4, f"=C{MIX1_ADJ_ROW}-B{MIX1_ADJ_ROW}",
     fmt=FMT_USD, bold=True, fill=RESULT_FILL)
cell(ws, MIX1_ADJ_ROW, 5, f"=(C{MIX1_ADJ_ROW}-B{MIX1_ADJ_ROW})/B{MIX1_ADJ_ROW}",
     fmt=FMT_PCT, bold=True, fill=RESULT_FILL)


# ============== MIX-ADJUSTMENT #2: 2nd-best mix as reference ==============
MIX2_HEADER_ROW = MIX1_ADJ_ROW + 3
text_cell(ws, MIX2_HEADER_ROW, 1,
          "MIX-ADJUSTMENT #2 — reference = 2nd-best product distribution",
          bold=True, fill=SECTION_FILL)
ws.cell(row=MIX2_HEADER_ROW, column=1).font = Font(bold=True, size=11)
ws.merge_cells(start_row=MIX2_HEADER_ROW, start_column=1, end_row=MIX2_HEADER_ROW, end_column=7)

MIX2_EXPLAIN_ROW = MIX2_HEADER_ROW + 1
c = ws.cell(row=MIX2_EXPLAIN_ROW, column=1,
            value=('Question: "If 1st-best customers had the same product distribution as '
                   '2nd-best customers, what would their average spend per account be?"'))
c.font = ITALIC
c.alignment = WRAP
ws.merge_cells(start_row=MIX2_EXPLAIN_ROW, start_column=1, end_row=MIX2_EXPLAIN_ROW, end_column=7)
ws.row_dimensions[MIX2_EXPLAIN_ROW].height = 28

MIX2_COLS_ROW = MIX2_EXPLAIN_ROW + 1
write_header(ws, MIX2_COLS_ROW, [
    "Product",
    "2nd-mix weight (ref)",
    "1st $/acct",
    "2nd $/acct",
    "weight × 1st $",
    "weight × 2nd $",
    "",
])

MIX2_DATA_START = MIX2_COLS_ROW + 1
MIX2_DATA_END = MIX2_DATA_START + len(products) - 1

for i in range(len(products)):
    r = MIX2_DATA_START + i
    input_row = DATA_START + i
    text_cell(ws, r, 1, f"=A{input_row}")
    cell(ws, r, 2, f"=C{input_row}", fmt=FMT_INT)  # 2nd-mix weight
    cell(ws, r, 3, f"=D{input_row}", fmt=FMT_USD)  # 1st $/acct
    cell(ws, r, 4, f"=E{input_row}", fmt=FMT_USD)  # 2nd $/acct
    cell(ws, r, 5, f"=B{r}*C{r}", fmt=FMT_USD)      # weight × 1st $
    cell(ws, r, 6, f"=B{r}*D{r}", fmt=FMT_USD)      # weight × 2nd $

MIX2_SUM_ROW = MIX2_DATA_END + 1
text_cell(ws, MIX2_SUM_ROW, 1, "Sum", bold=True, fill=SECTION_FILL)
cell(ws, MIX2_SUM_ROW, 2, f"=SUM(B{MIX2_DATA_START}:B{MIX2_DATA_END})",
     fmt=FMT_INT, bold=True, fill=SECTION_FILL)
cell(ws, MIX2_SUM_ROW, 3, "—", bold=True, fill=SECTION_FILL)
cell(ws, MIX2_SUM_ROW, 4, "—", bold=True, fill=SECTION_FILL)
cell(ws, MIX2_SUM_ROW, 5, f"=SUM(E{MIX2_DATA_START}:E{MIX2_DATA_END})",
     fmt=FMT_USD, bold=True, fill=SECTION_FILL)
cell(ws, MIX2_SUM_ROW, 6, f"=SUM(F{MIX2_DATA_START}:F{MIX2_DATA_END})",
     fmt=FMT_USD, bold=True, fill=SECTION_FILL)

MIX2_RESULT_LABEL = MIX2_SUM_ROW + 2
text_cell(ws, MIX2_RESULT_LABEL, 1, "RESULT — mix-adjustment #2", bold=True, fill=RESULT_FILL)
ws.merge_cells(start_row=MIX2_RESULT_LABEL, start_column=1, end_row=MIX2_RESULT_LABEL, end_column=7)

MIX2_RESULT_HEADER = MIX2_RESULT_LABEL + 1
write_header(ws, MIX2_RESULT_HEADER, ["", "1st best", "2nd best", "Δ $", "Δ %", "", ""])

MIX2_RAW_ROW = MIX2_RESULT_HEADER + 1
text_cell(ws, MIX2_RAW_ROW, 1, "Raw avg spend per approved")
cell(ws, MIX2_RAW_ROW, 2, f"=D{TOTAL_ROW}", fmt=FMT_USD)
cell(ws, MIX2_RAW_ROW, 3, f"=E{TOTAL_ROW}", fmt=FMT_USD)
cell(ws, MIX2_RAW_ROW, 4, f"=C{MIX2_RAW_ROW}-B{MIX2_RAW_ROW}", fmt=FMT_USD)
cell(ws, MIX2_RAW_ROW, 5, f"=(C{MIX2_RAW_ROW}-B{MIX2_RAW_ROW})/B{MIX2_RAW_ROW}", fmt=FMT_PCT)

MIX2_ADJ_ROW = MIX2_RAW_ROW + 1
text_cell(ws, MIX2_ADJ_ROW, 1,
          "Mix-adjusted (1st-best $ × 2nd-best mix)", bold=True, fill=RESULT_FILL)
cell(ws, MIX2_ADJ_ROW, 2, f"=E{MIX2_SUM_ROW}/B{MIX2_SUM_ROW}",
     fmt=FMT_USD, bold=True, fill=RESULT_FILL)
cell(ws, MIX2_ADJ_ROW, 3, f"=F{MIX2_SUM_ROW}/B{MIX2_SUM_ROW}",
     fmt=FMT_USD, bold=True, fill=RESULT_FILL)
cell(ws, MIX2_ADJ_ROW, 4, f"=C{MIX2_ADJ_ROW}-B{MIX2_ADJ_ROW}",
     fmt=FMT_USD, bold=True, fill=RESULT_FILL)
cell(ws, MIX2_ADJ_ROW, 5, f"=(C{MIX2_ADJ_ROW}-B{MIX2_ADJ_ROW})/B{MIX2_ADJ_ROW}",
     fmt=FMT_PCT, bold=True, fill=RESULT_FILL)


# ============== SUMMARY ==============
SUMMARY_HEADER = MIX2_ADJ_ROW + 3
text_cell(ws, SUMMARY_HEADER, 1,
          "SUMMARY — three ways to compute the avg-spend gap", bold=True, fill=SECTION_FILL)
ws.cell(row=SUMMARY_HEADER, column=1).font = Font(bold=True, size=11)
ws.merge_cells(start_row=SUMMARY_HEADER, start_column=1, end_row=SUMMARY_HEADER, end_column=7)

SUMMARY_COLS = SUMMARY_HEADER + 1
write_header(ws, SUMMARY_COLS, ["Calculation", "1st best", "2nd best", "Δ $", "Δ %", "", ""])

# Raw
SUM_RAW = SUMMARY_COLS + 1
text_cell(ws, SUM_RAW, 1, "Raw (no mix adjustment)")
cell(ws, SUM_RAW, 2, f"=D{TOTAL_ROW}", fmt=FMT_USD)
cell(ws, SUM_RAW, 3, f"=E{TOTAL_ROW}", fmt=FMT_USD)
cell(ws, SUM_RAW, 4, f"=C{SUM_RAW}-B{SUM_RAW}", fmt=FMT_USD)
cell(ws, SUM_RAW, 5, f"=(C{SUM_RAW}-B{SUM_RAW})/B{SUM_RAW}", fmt=FMT_PCT)

# Mix #1
SUM_MIX1 = SUM_RAW + 1
text_cell(ws, SUM_MIX1, 1, "Mix-adjusted #1 (2nd $ × 1st mix)")
cell(ws, SUM_MIX1, 2, f"=B{MIX1_ADJ_ROW}", fmt=FMT_USD)
cell(ws, SUM_MIX1, 3, f"=C{MIX1_ADJ_ROW}", fmt=FMT_USD)
cell(ws, SUM_MIX1, 4, f"=C{SUM_MIX1}-B{SUM_MIX1}", fmt=FMT_USD)
cell(ws, SUM_MIX1, 5, f"=(C{SUM_MIX1}-B{SUM_MIX1})/B{SUM_MIX1}", fmt=FMT_PCT)

# Mix #2
SUM_MIX2 = SUM_MIX1 + 1
text_cell(ws, SUM_MIX2, 1, "Mix-adjusted #2 (1st $ × 2nd mix)")
cell(ws, SUM_MIX2, 2, f"=B{MIX2_ADJ_ROW}", fmt=FMT_USD)
cell(ws, SUM_MIX2, 3, f"=C{MIX2_ADJ_ROW}", fmt=FMT_USD)
cell(ws, SUM_MIX2, 4, f"=C{SUM_MIX2}-B{SUM_MIX2}", fmt=FMT_USD)
cell(ws, SUM_MIX2, 5, f"=(C{SUM_MIX2}-B{SUM_MIX2})/B{SUM_MIX2}", fmt=FMT_PCT)


# ============== FORMULA GLOSSARY ==============
GLOSS_HEADER = SUM_MIX2 + 3
text_cell(ws, GLOSS_HEADER, 1, "FORMULA GLOSSARY — how each calculation is built",
          bold=True, fill=SECTION_FILL)
ws.cell(row=GLOSS_HEADER, column=1).font = Font(bold=True, size=11)
ws.merge_cells(start_row=GLOSS_HEADER, start_column=1, end_row=GLOSS_HEADER, end_column=7)

glossary_lines = [
    ("Raw weighted average (per group)",
     "= SUMPRODUCT(approved_counts, $_per_acct) / SUM(approved_counts)"),
    ("  Intuition",
     "For each product, multiply the number of approved accounts by the $/acct, sum across "
     "products, divide by total approved. Each approved account contributes once."),
    ("", ""),
    ("Mix-adjustment #1 (2nd $ × 1st mix)",
     "= SUMPRODUCT(1st_approved_counts, 2nd_$_per_acct) / SUM(1st_approved_counts)"),
    ("  Intuition",
     "Weight the 2nd-best per-product spend by the 1st-best product mix. This tells you what "
     "the 2nd-best group WOULD have spent per account if the product distribution had matched "
     "the 1st-best group. Isolates per-customer behavior from mix."),
    ("", ""),
    ("Mix-adjustment #2 (1st $ × 2nd mix)",
     "= SUMPRODUCT(2nd_approved_counts, 1st_$_per_acct) / SUM(2nd_approved_counts)"),
    ("  Intuition",
     "Symmetric version. Weight the 1st-best per-product spend by the 2nd-best product mix. "
     "Tells you what the 1st-best group would have spent if the product distribution had "
     "matched the 2nd-best group. The two adjustments bracket the true per-customer effect."),
    ("", ""),
    ("Why both directions matter",
     "The two methods give slightly different numbers because the reference mix is different. "
     "Both are valid. Reporting the range (-3% to -5%) is more honest than picking one."),
    ("", ""),
    ("Reading the raw -22% in context",
     "The raw aggregate gap mixes two effects: (a) per-customer spend differences at the "
     "product level, and (b) product-mix shifts between groups. The mix-adjusted calculations "
     "separate them. Raw -22% = mostly (b), not (a). At the product level, per-customer spend "
     "is roughly flat (IAV and ION actually UP in the 2nd-best slot)."),
]

r = GLOSS_HEADER + 1
for label, body in glossary_lines:
    if not label and not body:
        r += 1
        continue
    c = ws.cell(row=r, column=1, value=label)
    c.font = BOLD if label and not label.startswith("  ") else Font(size=10, italic=True)
    c.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c2 = ws.cell(row=r, column=3, value=body)
    c2.font = Font(size=10)
    c2.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    ws.row_dimensions[r].height = max(18, len(body) // 80 * 18 + 18) if body else 18
    r += 1


out = r"C:\Users\andre\New_projects\cards\campaigns\PCQ\next_best_card\deliverables\pcq_nbc_mix_adjusted.xlsx"
wb.save(out)
print("Saved:", out)
print()
print("Row map (for reference):")
print(f"  Input data          rows {DATA_START}-{DATA_END}")
print(f"  Raw totals          row  {TOTAL_ROW}")
print(f"  Mix-adj #1 data     rows {MIX1_DATA_START}-{MIX1_DATA_END}")
print(f"  Mix-adj #1 sum      row  {MIX1_SUM_ROW}")
print(f"  Mix-adj #1 result   row  {MIX1_ADJ_ROW}")
print(f"  Mix-adj #2 data     rows {MIX2_DATA_START}-{MIX2_DATA_END}")
print(f"  Mix-adj #2 sum      row  {MIX2_SUM_ROW}")
print(f"  Mix-adj #2 result   row  {MIX2_ADJ_ROW}")
print(f"  Summary             rows {SUM_RAW}-{SUM_MIX2}")
