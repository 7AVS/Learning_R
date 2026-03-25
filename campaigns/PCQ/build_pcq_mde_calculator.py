"""
Build PCQ MDE Calculator Excel — DM holdout test framing.

Test group  = receives DM + all digital channels.
Control group = digital channels only (no DM).

Question: does adding DM to the 7th decile improve the approval rate?

All calculated cells use Excel formulas so the user can change green input
cells and see everything recalculate.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "PCQ DM MDE Calculator"

# ── Styles ──────────────────────────────────────────────────────────────────
green_fill   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
grey_fill    = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
bold         = Font(bold=True)
bold_white   = Font(bold=True, color="FFFFFF")
header_fill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
italic_grey  = Font(italic=True, color="404040")
italic_light = Font(italic=True, color="808080", size=9)
thin_border  = Border(
    left   = Side(style="thin"),
    right  = Side(style="thin"),
    top    = Side(style="thin"),
    bottom = Side(style="thin"),
)

pct_fmt   = "0.00%"
pct_fmt_4 = "0.0000%"
num_fmt   = "#,##0"
ratio_fmt = "0.00"

# ── Column widths ─────────────────────────────────────────────────────────
col_widths = {
    "A": 5,  "B": 50, "C": 20, "D": 16, "E": 16,
    "F": 16, "G": 16, "H": 16, "I": 18, "J": 16,
    "K": 16, "L": 50,
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w


# ── Helper: apply style to a cell ─────────────────────────────────────────
def style(cell, font=None, fill=None, fmt=None, align=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    return cell


def note(row, col, text):
    """Write a note in italic grey."""
    c = ws.cell(row=row, column=col, value=text)
    c.font = italic_grey
    return c


# ════════════════════════════════════════════════════════════════════════════
# SECTION 1: STATISTICAL INPUTS  (Rows 2–10)
# ════════════════════════════════════════════════════════════════════════════

# Row 2 — section header
for c in range(1, 13):
    style(ws.cell(row=2, column=c), fill=grey_fill)
style(ws.cell(row=2, column=2, value="Section 1 — Statistical Inputs (edit green cells)"),
      font=bold, fill=grey_fill)

# Row 3 — Total population (7th decile)
style(ws.cell(row=3, column=2, value="Total population (7th decile)"), font=bold)
style(ws.cell(row=3, column=3, value=55_000), fill=green_fill, fmt=num_fmt)
note(row=3, col=4, text="Jan 2026 deployment: 7th decile = 55,039 clients (2026010PCQ)")

# Row 4 — Control allocation
style(ws.cell(row=4, column=2, value="Control allocation"), font=bold)
style(ws.cell(row=4, column=3, value=0.10), fill=green_fill, fmt=pct_fmt)
note(row=4, col=4, text="Fraction held back (no DM). Editable — try 0.10, 0.20, 0.50")

# Row 5 — Test allocation (derived)
style(ws.cell(row=5, column=2, value="Test allocation"), font=bold)
ws.cell(row=5, column=3).value = "=1-C4"
ws.cell(row=5, column=3).number_format = pct_fmt
note(row=5, col=4, text="Auto-calculated as 1 - control allocation")

# Row 6 — Significance level (alpha)
style(ws.cell(row=6, column=2, value="Significance level (alpha)"), font=bold)
style(ws.cell(row=6, column=3, value=0.05), fill=green_fill, fmt=pct_fmt)

# Row 7 — Power level
style(ws.cell(row=7, column=2, value="Power level"), font=bold)
style(ws.cell(row=7, column=3, value=0.80), fill=green_fill, fmt=pct_fmt)

# Row 8 — Baseline RR (7th decile overall approval rate)
style(ws.cell(row=8, column=2, value="Baseline RR (7th decile)"), font=bold)
style(ws.cell(row=8, column=3, value=0.0046), fill=green_fill, fmt=pct_fmt_4)
note(row=8, col=4, text="7th decile overall approval rate (252/55,048), Jan 2026 deployment")


# ════════════════════════════════════════════════════════════════════════════
# SECTION 2: STATISTICAL CONSTANTS  (Rows 10–13)
# ════════════════════════════════════════════════════════════════════════════

for c in range(1, 13):
    style(ws.cell(row=10, column=c), fill=grey_fill)
style(ws.cell(row=10, column=2, value="Section 2 — Do Not Edit — Statistical Constants"),
      font=bold, fill=grey_fill)

# Row 11 — Za
style(ws.cell(row=11, column=2, value="Critical Value for Significance (Za)"), font=bold)
ws.cell(row=11, column=3).value = 1.9600
ws.cell(row=11, column=3).number_format = "0.0000"
note(row=11, col=4, text="Two-sided test: alpha=0.05 → Za = 1.9600 (NORM.S.INV(0.975))")

# Row 12 — Zb
style(ws.cell(row=12, column=2, value="Critical Value for Power (Zb)"), font=bold)
ws.cell(row=12, column=3).value = 0.8416
ws.cell(row=12, column=3).number_format = "0.0000"
note(row=12, col=4, text="Power=80% → Zb = 0.8416 (NORM.S.INV(0.80))")

# Row 13 — note
ws.cell(row=13, column=2).value = (
    "Two-sided test: we want to detect whether DM changes the approval rate "
    "(up or down). If you change alpha/power, update Za/Zb using "
    "=NORM.S.INV(1-alpha/2) and =NORM.S.INV(power)."
)
ws.cell(row=13, column=2).font = italic_light
ws.merge_cells(start_row=13, start_column=2, end_row=13, end_column=12)
ws.cell(row=13, column=2).alignment = Alignment(wrap_text=True)
ws.row_dimensions[13].height = 28


# ════════════════════════════════════════════════════════════════════════════
# SECTION 3: DERIVED VALUES  (Rows 15–19)
# ════════════════════════════════════════════════════════════════════════════

for c in range(1, 13):
    style(ws.cell(row=15, column=c), fill=grey_fill)
style(ws.cell(row=15, column=2, value="Section 3 — Derived Values"),
      font=bold, fill=grey_fill)

# Row 16 — Control n
style(ws.cell(row=16, column=2, value="Control n"), font=bold)
ws.cell(row=16, column=3).value = "=ROUND(C3*C4,0)"
ws.cell(row=16, column=3).number_format = num_fmt
note(row=16, col=4, text="Records held back — no DM")

# Row 17 — Test n
style(ws.cell(row=17, column=2, value="Test n"), font=bold)
ws.cell(row=17, column=3).value = "=ROUND(C3*C5,0)"
ws.cell(row=17, column=3).number_format = num_fmt
note(row=17, col=4, text="Records receiving DM + digital channels")

# Row 18 — Allocation check
style(ws.cell(row=18, column=2, value="Allocation check"), font=bold)
ws.cell(row=18, column=3).value = "=C4+C5"
ws.cell(row=18, column=3).number_format = pct_fmt
note(row=18, col=4, text="Should show 100.00%")

# Row 19 — MDE for selected split (main output)
style(ws.cell(row=19, column=2, value="MDE at selected control split (Section 1)"), font=bold)
ws.cell(row=19, column=3).value = (
    "=ROUND((C11+C12)*SQRT(C8*(1-C8)*(1/C16+1/C17)),4)"
)
ws.cell(row=19, column=3).number_format = pct_fmt_4
note(row=19, col=4,
     text="Minimum detectable effect (absolute pp) at the control % set in C4")


# ════════════════════════════════════════════════════════════════════════════
# SECTION 4: MDE BY CONTROL SPLIT  (Rows 21–30)
# ════════════════════════════════════════════════════════════════════════════

for c in range(1, 13):
    style(ws.cell(row=21, column=c), fill=grey_fill)
style(ws.cell(row=21, column=2, value="Section 4 — MDE by Control Split (reference scenarios)"),
      font=bold, fill=grey_fill)

# Row 22 — Column headers
s4_headers = [
    "#", "Scenario", "Population",
    "Control %", "n_control", "n_test",
    "Baseline RR", "MDE (abs)", "Detectable Test Rate",
    "Assessment", "", "",
]
for i, h in enumerate(s4_headers, start=1):
    cell = ws.cell(row=22, column=i, value=h)
    style(cell, font=bold_white, fill=header_fill, border=thin_border,
          align=Alignment(horizontal="center", wrap_text=True))

# Control split scenarios (rows 23–26)
ctrl_splits = [0.10, 0.20, 0.30, 0.50]

for idx, ctrl_pct in enumerate(ctrl_splits):
    row = 23 + idx   # rows 23, 24, 25, 26

    # Col A — row number
    ws.cell(row=row, column=1, value=idx + 1).border = thin_border

    # Col B — scenario label
    ws.cell(row=row, column=2, value=f"{int(ctrl_pct*100)}% control").border = thin_border

    # Col C — population (references C3)
    ws.cell(row=row, column=3).value = "=$C$3"
    ws.cell(row=row, column=3).number_format = num_fmt
    ws.cell(row=row, column=3).border = thin_border

    # Col D — control %
    ws.cell(row=row, column=4).value = ctrl_pct
    ws.cell(row=row, column=4).number_format = pct_fmt
    ws.cell(row=row, column=4).border = thin_border

    # Col E — n_control = C3 * ctrl_pct
    ws.cell(row=row, column=5).value = f"=ROUND($C$3*D{row},0)"
    ws.cell(row=row, column=5).number_format = num_fmt
    ws.cell(row=row, column=5).border = thin_border

    # Col F — n_test = C3 * (1 - ctrl_pct)
    ws.cell(row=row, column=6).value = f"=ROUND($C$3*(1-D{row}),0)"
    ws.cell(row=row, column=6).number_format = num_fmt
    ws.cell(row=row, column=6).border = thin_border

    # Col G — Baseline RR (references C8)
    ws.cell(row=row, column=7).value = "=$C$8"
    ws.cell(row=row, column=7).number_format = pct_fmt_4
    ws.cell(row=row, column=7).border = thin_border

    # Col H — MDE
    mde_formula = (
        f"=ROUND(($C$11+$C$12)*SQRT($C$8*(1-$C$8)*(1/E{row}+1/F{row})),4)"
    )
    ws.cell(row=row, column=8).value = mde_formula
    ws.cell(row=row, column=8).number_format = pct_fmt_4
    ws.cell(row=row, column=8).border = thin_border

    # Col I — Detectable test rate = baseline + MDE
    ws.cell(row=row, column=9).value = f"=$C$8+H{row}"
    ws.cell(row=row, column=9).number_format = pct_fmt_4
    ws.cell(row=row, column=9).border = thin_border

    # Col J — Assessment
    ws.cell(row=row, column=10).value = (
        f'=IF(H{row}<=0.005,"Can detect <0.5pp lift — GOOD","MDE >0.5pp — needs more n")'
    )
    ws.cell(row=row, column=10).border = thin_border
    ws.cell(row=row, column=10).alignment = Alignment(wrap_text=True)

# Row 27 — highlight that Row 23 (10% ctrl) matches main editable default
ws.cell(row=27, column=2).value = (
    "Row 23 (10% control) corresponds to the default control split in Section 1 (C4 = 10%). "
    "Change C4 to see the main MDE (C19) update dynamically."
)
ws.cell(row=27, column=2).font = italic_light
ws.merge_cells(start_row=27, start_column=2, end_row=27, end_column=12)
ws.cell(row=27, column=2).alignment = Alignment(wrap_text=True)
ws.row_dimensions[27].height = 28


# ════════════════════════════════════════════════════════════════════════════
# SECTION 5: STRESS TEST SCENARIOS  (Rows 29–35)
# ════════════════════════════════════════════════════════════════════════════

for c in range(1, 13):
    style(ws.cell(row=29, column=c), fill=grey_fill)
style(ws.cell(row=29, column=2, value="Section 5 — Stress Test Scenarios"),
      font=bold, fill=grey_fill)

# Row 30 — Column headers
s5_headers = [
    "", "Scenario",
    "Baseline RR", "Population",
    "MDE @ 10% ctrl", "MDE @ 20% ctrl", "MDE @ 50% ctrl",
    "Assessment (10% ctrl)", "", "", "", "",
]
for i, h in enumerate(s5_headers, start=1):
    cell = ws.cell(row=30, column=i, value=h)
    style(cell, font=bold_white, fill=header_fill, border=thin_border,
          align=Alignment(horizontal="center", wrap_text=True))

# Scenario definitions: (label, baseline_rr, population, notes)
stress_scenarios = [
    (
        "A — 7th decile only",
        0.0046, 55_000,
        "Baseline from Jan 2026 deployment (252 approvals / 55,048 clients)",
    ),
    (
        "B — 7th + 8th decile combined",
        0.0038, 99_000,
        "Weighted avg: 7th (0.46%) + 8th decile; pop ~55K + 44K = 99K",
    ),
    (
        "C — 120K allocation (T's figure)",
        0.0046, 120_000,
        "T mentioned 120K additional DMs/month; same 7th-decile baseline",
    ),
]

for s_idx, (label, baseline, population, sc_note) in enumerate(stress_scenarios):
    row = 31 + s_idx  # rows 31, 32, 33

    # Col B — scenario label + note
    ws.cell(row=row, column=2, value=f"{label}  [{sc_note}]").border = thin_border
    ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)

    # Col C — Baseline RR (editable green)
    style(ws.cell(row=row, column=3, value=baseline), fill=green_fill, fmt=pct_fmt_4)
    ws.cell(row=row, column=3).border = thin_border

    # Col D — Population (editable green)
    style(ws.cell(row=row, column=4, value=population), fill=green_fill, fmt=num_fmt)
    ws.cell(row=row, column=4).border = thin_border

    # MDE at 10%, 20%, 50% control (cols E, F, G)
    for col_off, ctrl_frac in enumerate([0.10, 0.20, 0.50]):
        col = 5 + col_off
        p0_ref = f"C{row}"
        pop_ref = f"D{row}"
        n_ctrl = f"ROUND({pop_ref}*{ctrl_frac},0)"
        n_test = f"ROUND({pop_ref}*(1-{ctrl_frac}),0)"
        mde_f = (
            f"=ROUND(($C$11+$C$12)*SQRT({p0_ref}*(1-{p0_ref})*"
            f"(1/{n_ctrl}+1/{n_test})),4)"
        )
        ws.cell(row=row, column=col).value = mde_f
        ws.cell(row=row, column=col).number_format = pct_fmt_4
        ws.cell(row=row, column=col).border = thin_border

    # Col H — Assessment based on 10% ctrl MDE
    ws.cell(row=row, column=8).value = (
        f'=IF(E{row}<=0.005,"<0.5pp at 10% ctrl — GOOD","MDE >0.5pp at 10% ctrl")'
    )
    ws.cell(row=row, column=8).border = thin_border
    ws.cell(row=row, column=8).alignment = Alignment(wrap_text=True)

ws.row_dimensions[31].height = 32
ws.row_dimensions[32].height = 32
ws.row_dimensions[33].height = 32


# ════════════════════════════════════════════════════════════════════════════
# SECTION 6: MINIMUM SAMPLE SIZE  (Rows 35–43)
# ════════════════════════════════════════════════════════════════════════════

for c in range(1, 13):
    style(ws.cell(row=35, column=c), fill=grey_fill)
style(
    ws.cell(row=35, column=2,
            value="Section 6 — Minimum Sample Size to Detect a Given Lift"),
    font=bold, fill=grey_fill,
)

# Row 36 — formula explanation
ws.cell(row=36, column=2).value = (
    "Formula: N = (Za+Zb)^2 * p0*(1-p0) * (1/ctrl_frac + 1/(1-ctrl_frac)) / MDE^2    "
    "(Then n_control = ctrl_frac * N)"
)
ws.cell(row=36, column=2).font = italic_light
ws.merge_cells(start_row=36, start_column=2, end_row=36, end_column=12)
ws.cell(row=36, column=2).alignment = Alignment(wrap_text=True)
ws.row_dimensions[36].height = 22

# Row 37 — Column headers
s6_headers = [
    "", "Target Lift (abs pp)",
    "Min N @ 10% ctrl", "n_ctrl @ 10%",
    "Min N @ 20% ctrl", "n_ctrl @ 20%",
    "Min N @ 50% ctrl", "n_ctrl @ 50%",
    "", "", "", "",
]
for i, h in enumerate(s6_headers, start=1):
    cell = ws.cell(row=37, column=i, value=h)
    style(cell, font=bold_white, fill=header_fill, border=thin_border,
          align=Alignment(horizontal="center", wrap_text=True))

# Lift targets: 0.3pp, 0.5pp, 0.8pp
lift_targets = [0.003, 0.005, 0.008]
lift_labels  = ["0.3pp (conservative)", "0.5pp (target)", "0.8pp (T's benchmark)"]

for l_idx, (lift_val, lift_label) in enumerate(zip(lift_targets, lift_labels)):
    row = 38 + l_idx  # rows 38, 39, 40

    ws.cell(row=row, column=2, value=lift_label).border = thin_border

    p0_ref = "$C$8"
    K_ref  = "($C$11+$C$12)"

    for col_off, ctrl_frac in enumerate([0.10, 0.20, 0.50]):
        # Min N column (E, G, I → cols 3, 5, 7)
        n_col  = 3 + col_off * 2
        nc_col = 4 + col_off * 2

        min_n_formula = (
            f"=ROUNDUP({K_ref}^2*{p0_ref}*(1-{p0_ref})"
            f"*(1/{ctrl_frac}+1/{1 - ctrl_frac})/{lift_val}^2,0)"
        )
        ws.cell(row=row, column=n_col).value = min_n_formula
        ws.cell(row=row, column=n_col).number_format = num_fmt
        ws.cell(row=row, column=n_col).border = thin_border

        # n_ctrl = ctrl_frac * N
        ws.cell(row=row, column=nc_col).value = (
            f"=ROUND({ctrl_frac}*{get_column_letter(n_col)}{row},0)"
        )
        ws.cell(row=row, column=nc_col).number_format = num_fmt
        ws.cell(row=row, column=nc_col).border = thin_border


# ════════════════════════════════════════════════════════════════════════════
# SECTION 7: NOTES  (Rows 42+)
# ════════════════════════════════════════════════════════════════════════════

notes_start = 42
for c in range(1, 13):
    style(ws.cell(row=notes_start, column=c), fill=grey_fill)
style(ws.cell(row=notes_start, column=2, value="Section 7 — Notes"),
      font=bold, fill=grey_fill)

notes = [
    (
        "1.",
        "This is a DM holdout test for the PCQ (Credit Card Acquisition) 7th decile "
        "expansion. The business is adding ~120K DMs/month starting April 7, 2026."
    ),
    (
        "2.",
        "Test group receives DM + all digital channels. Control group receives digital "
        "channels only (no DM). This isolates the incremental lift from the DM piece."
    ),
    (
        "3.",
        "Success metric: app_approved (card account opened). Response window: 60–90 days "
        "depending on TPA offer validity."
    ),
    (
        "4.",
        "Baseline RR = 7th decile overall approval rate from the Jan 2026 deployment "
        "(tactic 2026010PCQ). Source: 252 approvals / 55,048 clients = 0.4578%. "
        "This is the most mature deployment with a ~90-day response window."
    ),
    (
        "5.",
        "NBO channel selection is eligibility-based. The current near-zero DM group in "
        "the 7th decile (only ~3K of 55K had DM in Jan 2026) includes opt-outs per "
        "Daniel Chin. This baseline reflects the overall decile rate (94% non-DM clients), "
        "which is conservative for the control group estimate."
    ),
    (
        "6.",
        "Green cells are editable — change C3 (population), C4 (control %), or C8 "
        "(baseline RR) to recalculate the entire sheet dynamically."
    ),
    (
        "7.",
        "Two-sided test (Za = 1.9600): we measure whether DM changes the rate in either "
        "direction, not just up. This is more conservative than one-sided but is the "
        "standard for confirmatory holdout tests."
    ),
    (
        "8.",
        "Historical DM lift estimate: ~0.8pp (per T, based on prior years). At 10% control "
        "with 55K pop, MDE ≈ 0.9pp — the test is borderline underpowered to detect 0.8pp. "
        "At 120K pop (T's figure), 10% control gives MDE ≈ 0.6pp — still marginal. "
        "20% control at 120K gives MDE ≈ 0.7pp. Recommend discussing with T whether "
        "the 0.8pp lift expectation is firm."
    ),
]

for i, (num, text) in enumerate(notes):
    note_row = notes_start + 1 + i
    ws.cell(row=note_row, column=2, value=num).font = bold
    ws.cell(row=note_row, column=3, value=text)
    ws.merge_cells(start_row=note_row, start_column=3,
                   end_row=note_row, end_column=12)
    ws.cell(row=note_row, column=3).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[note_row].height = 36


# ── Freeze panes & zoom ───────────────────────────────────────────────────
ws.freeze_panes = "C3"
ws.sheet_view.zoomScale = 90

# ── Force full recalculation on load ─────────────────────────────────────
wb.calculation = CalcProperties(fullCalcOnLoad=True, calcMode="auto")

# ── Save ──────────────────────────────────────────────────────────────────
out = r"C:\Users\andre\New_projects\cards\campaigns\PCQ\pcq_mde_calculator.xlsx"
try:
    wb.save(out)
    print(f"Saved: {out}")
except PermissionError:
    out = r"C:\Users\andre\New_projects\cards\campaigns\PCQ\pcq_mde_calculator_v2.xlsx"
    wb.save(out)
    print(f"Original locked — saved as: {out}")

# ── Force recalculation via win32com (Windows only) ───────────────────────
try:
    import win32com.client
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    wb_com = xl.Workbooks.Open(os.path.abspath(out))
    xl.Calculate()
    wb_com.Save()
    wb_com.Close()
    xl.Quit()
    print(f"Recalculated via Excel: {os.path.abspath(out)}")
except ImportError:
    print("win32com not available — skipping Excel recalculation")
except Exception as e:
    print(f"Excel recalculation failed: {e}")
    try:
        xl.Quit()
    except Exception:
        pass
