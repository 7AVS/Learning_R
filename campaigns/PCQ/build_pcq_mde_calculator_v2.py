"""
Build PCQ MDE Calculator Excel — DM test framing.

Question: given the total population and a control holdout, what is the
smallest lift (in absolute pp) we can detect with 80% power at alpha=5%?

All calculated cells use Excel formulas so the user can change green input
cells and see everything recalculate.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties
import shutil
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "MDE Calculator"

# ── Styles ──────────────────────────────────────────────────────────────────
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
grey_fill  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
blue_fill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
bold       = Font(bold=True)
bold_white = Font(bold=True, color="FFFFFF")
italic_font = Font(italic=True, color="404040")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

pct_fmt   = "0.00%"
pct_fmt_4 = "0.0000%"
num_fmt   = "#,##0"
ratio_fmt = "0.00"

# ── Column widths ────────────────────────────────────────────────────────────
col_widths = {
    "A": 5,  "B": 50, "C": 20, "D": 40, "E": 16,
    "F": 16, "G": 16, "H": 16, "I": 18, "J": 16,
    "K": 16, "L": 50,
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w


# ── Helper: apply style to a cell ────────────────────────────────────────────
def style(cell, font=None, fill=None, fmt=None, align=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    return cell


# ════════════════════════════════════════════════════════════════════════════
# SECTION 1: STATISTICAL INPUTS  (Rows 2-10)
# ════════════════════════════════════════════════════════════════════════════

# Row 2 — section header
for c in range(1, 13):
    style(ws.cell(row=2, column=c), fill=grey_fill)
style(ws.cell(row=2, column=2, value="Statistical Inputs"), font=bold, fill=grey_fill)

# Row 3 — Total population
style(ws.cell(row=3, column=2, value="Total population"), font=bold)
style(ws.cell(row=3, column=3, value=55048), fill=green_fill, fmt=num_fmt)

# Row 4 — Control allocation
style(ws.cell(row=4, column=2, value="Control allocation"), font=bold)
style(ws.cell(row=4, column=3, value=0.20), fill=green_fill, fmt=pct_fmt)

# Row 5 — Test allocation (formula)
style(ws.cell(row=5, column=2, value="Test allocation"), font=bold)
ws.cell(row=5, column=3).value = "=1-C4"
ws.cell(row=5, column=3).number_format = pct_fmt

# Row 6 — Significance level (alpha)
style(ws.cell(row=6, column=2, value="Significance level (alpha)"), font=bold)
style(ws.cell(row=6, column=3, value=0.05), fill=green_fill, fmt=pct_fmt)

# Row 7 — Power level
style(ws.cell(row=7, column=2, value="Power level"), font=bold)
style(ws.cell(row=7, column=3, value=0.80), fill=green_fill, fmt=pct_fmt)

# Row 8 — Baseline RR
style(ws.cell(row=8, column=2, value="Baseline RR (Period-ASC, digital only)"), font=bold)
style(ws.cell(row=8, column=3, value=0.0030), fill=green_fill, fmt=pct_fmt_4)
ws.cell(row=8, column=4, value="7th decile Period-ASC rate, Jan 2026. Decile NOT targeted for DM.")
ws.cell(row=8, column=4).font = italic_font

# Row 9 — Minimum lift to act
style(ws.cell(row=9, column=2, value="Minimum lift to act"), font=bold)
style(ws.cell(row=9, column=3, value=0.005), fill=green_fill, fmt=pct_fmt)
ws.cell(row=9, column=4, value="Smallest lift the business would act on")
ws.cell(row=9, column=4).font = italic_font


# ════════════════════════════════════════════════════════════════════════════
# SECTION 2: DO NOT EDIT  (Rows 11-14)
# ════════════════════════════════════════════════════════════════════════════

for c in range(1, 13):
    style(ws.cell(row=11, column=c), fill=grey_fill)
style(ws.cell(row=11, column=2, value="Do Not Edit \u2014 Statistical Constants"), font=bold, fill=grey_fill)

# Row 12 — Za (two-sided)
style(ws.cell(row=12, column=2, value="Critical Value for Significance (Za)"), font=bold)
ws.cell(row=12, column=3).value = 1.9600
ws.cell(row=12, column=3).number_format = "0.0000"
ws.cell(row=12, column=4, value="Two-sided. Change to 1.6449 for one-sided test.")
ws.cell(row=12, column=4).font = italic_font

# Row 13 — Zb (80% power)
style(ws.cell(row=13, column=2, value="Critical Value for Power (Zb)"), font=bold)
ws.cell(row=13, column=3).value = 0.8416
ws.cell(row=13, column=3).number_format = "0.0000"
ws.cell(row=13, column=4, value="For 80% power.")
ws.cell(row=13, column=4).font = italic_font

# Row 14 — note
ws.cell(row=14, column=2).value = (
    "Za and Zb are hardcoded for alpha=5% (two-sided) and power=80%. "
    "If you change alpha or power, update Za/Zb manually using "
    "=NORM.S.INV(1-alpha/2) and =NORM.S.INV(power)."
)
ws.cell(row=14, column=2).font = Font(italic=True, color="808080", size=9)


# ════════════════════════════════════════════════════════════════════════════
# SECTION 3: DERIVED VALUES  (Rows 16-21)
# ════════════════════════════════════════════════════════════════════════════

for c in range(1, 13):
    style(ws.cell(row=16, column=c), fill=grey_fill)
style(ws.cell(row=16, column=2, value="Derived Values"), font=bold, fill=grey_fill)

# Row 17 — Control n
style(ws.cell(row=17, column=2, value="Control n"), font=bold)
ws.cell(row=17, column=3).value = "=C3*C4"
ws.cell(row=17, column=3).number_format = num_fmt

# Row 18 — Test n
style(ws.cell(row=18, column=2, value="Test n"), font=bold)
ws.cell(row=18, column=3).value = "=C3*C5"
ws.cell(row=18, column=3).number_format = num_fmt

# Row 19 — Allocation check
style(ws.cell(row=19, column=2, value="Allocation check"), font=bold)
ws.cell(row=19, column=3).value = "=C4+C5"
ws.cell(row=19, column=3).number_format = pct_fmt
ws.cell(row=19, column=4, value="Should show 100%")
ws.cell(row=19, column=4).font = italic_font

# Row 20 — MDE
style(ws.cell(row=20, column=2, value="MDE (minimum detectable effect)"), font=bold)
ws.cell(row=20, column=3).value = (
    "=ROUND(($C$12+$C$13)*SQRT(C8*(1-C8)*(1/C17+1/C18)),6)"
)
ws.cell(row=20, column=3).number_format = pct_fmt_4

# Row 21 — Detectable test rate
style(ws.cell(row=21, column=2, value="Detectable test rate"), font=bold)
ws.cell(row=21, column=3).value = "=C8+C20"
ws.cell(row=21, column=3).number_format = pct_fmt_4

# Row 22 — Powered for min lift?
style(ws.cell(row=22, column=2, value="Powered for minimum lift?"), font=bold)
ws.cell(row=22, column=3).value = '=IF(C20<=C9,"YES","NO")'
ws.cell(row=22, column=4, value='YES = MDE is at or below the minimum lift to act')
ws.cell(row=22, column=4).font = italic_font


# ════════════════════════════════════════════════════════════════════════════
# SECTION 4: EXPANSION SCENARIOS  (Rows 24-32)
# ════════════════════════════════════════════════════════════════════════════

for c in range(1, 13):
    style(ws.cell(row=24, column=c), fill=grey_fill)
style(ws.cell(row=24, column=2, value="Expansion Scenarios"), font=bold, fill=grey_fill)

# Row 25 — Column headers
headers_s4 = [
    "#", "Scenario", "Deciles", "Population",
    "Baseline (Period-ASC)", "Control %",
    "n_Control", "n_Test",
    "MDE", "Detectable Rate",
    "Min Lift", "Powered?",
]
for i, h in enumerate(headers_s4, start=1):
    cell = ws.cell(row=25, column=i, value=h)
    style(cell, font=bold_white, fill=blue_fill, border=thin_border,
          align=Alignment(horizontal="center", wrap_text=True))

# Scenario rows 26-31
scenarios_s4 = [
    (1, "7th only",          "7",   55048,  0.0030),
    (2, "7th + 8th",         "7-8", 98838,  0.0025),
    (3, "7th-9th",           "7-9", 167054, 0.0019),
    (4, "120K budget",       "7-9", 120000, 0.0019),
    (5, "All (7-10)",        "7-10",235642, 0.0016),
    (6, "Custom",            "\u2014", 120000, 0.0030),
]

for idx, (num, label, deciles, pop, base) in enumerate(scenarios_s4):
    row = 26 + idx  # rows 26-31

    # Col A — #
    ws.cell(row=row, column=1, value=num)
    ws.cell(row=row, column=1).border = thin_border

    # Col B — Scenario name
    ws.cell(row=row, column=2, value=label)
    ws.cell(row=row, column=2).border = thin_border

    # Col C — Deciles
    ws.cell(row=row, column=3, value=deciles)
    ws.cell(row=row, column=3).border = thin_border

    # Col D — Population (green, editable)
    style(ws.cell(row=row, column=4, value=pop),
          fill=green_fill, fmt=num_fmt, border=thin_border)

    # Col E — Baseline (green, editable)
    style(ws.cell(row=row, column=5, value=base),
          fill=green_fill, fmt=pct_fmt_4, border=thin_border)

    # Col F — Control % (references Section 1 input $C$4)
    ws.cell(row=row, column=6).value = "=$C$4"
    ws.cell(row=row, column=6).number_format = pct_fmt
    ws.cell(row=row, column=6).border = thin_border

    # Col G — n_Control
    ws.cell(row=row, column=7).value = f"=D{row}*F{row}"
    ws.cell(row=row, column=7).number_format = num_fmt
    ws.cell(row=row, column=7).border = thin_border

    # Col H — n_Test
    ws.cell(row=row, column=8).value = f"=D{row}*(1-F{row})"
    ws.cell(row=row, column=8).number_format = num_fmt
    ws.cell(row=row, column=8).border = thin_border

    # Col I — MDE
    mde_formula = (
        f"=ROUND(($C$12+$C$13)*SQRT(E{row}*(1-E{row})*(1/G{row}+1/H{row})),6)"
    )
    ws.cell(row=row, column=9).value = mde_formula
    ws.cell(row=row, column=9).number_format = pct_fmt_4
    ws.cell(row=row, column=9).border = thin_border

    # Col J — Detectable Rate
    ws.cell(row=row, column=10).value = f"=E{row}+I{row}"
    ws.cell(row=row, column=10).number_format = pct_fmt_4
    ws.cell(row=row, column=10).border = thin_border

    # Col K — Min Lift (references Section 1 $C$9)
    ws.cell(row=row, column=11).value = "=$C$9"
    ws.cell(row=row, column=11).number_format = pct_fmt
    ws.cell(row=row, column=11).border = thin_border

    # Col L — Powered?
    powered_formula = f'=IF(I{row}<=K{row},"YES","NO \u2014 MDE too large")'
    ws.cell(row=row, column=12).value = powered_formula
    ws.cell(row=row, column=12).border = thin_border
    ws.cell(row=row, column=12).alignment = Alignment(wrap_text=True)


# ════════════════════════════════════════════════════════════════════════════
# SECTION 5: STRESS TEST — CONTROL SPLIT SENSITIVITY  (Rows 34-43)
# ════════════════════════════════════════════════════════════════════════════

stress1_start = 34
for c in range(1, 13):
    style(ws.cell(row=stress1_start, column=c), fill=grey_fill)
style(
    ws.cell(row=stress1_start, column=2,
            value="Stress Test \u2014 Control Split Sensitivity (Scenario 1: 7th decile)"),
    font=bold, fill=grey_fill,
)

# Row 35 — editable baseline for this stress test
style(ws.cell(row=stress1_start + 1, column=2, value="Baseline (editable):"), font=bold)
style(ws.cell(row=stress1_start + 1, column=3, value=0.0030),
      fill=green_fill, fmt=pct_fmt_4)
ws.cell(row=stress1_start + 1, column=4,
        value="Defaults to Section 1 baseline — change to explore")
ws.cell(row=stress1_start + 1, column=4).font = italic_font

# Row 36 — column headers
s5_hdr_row = stress1_start + 2
s5_headers = [
    "", "Control %", "n_Control", "n_Test",
    "MDE", "Detectable Rate", "Clients Withheld",
]
for i, h in enumerate(s5_headers, start=1):
    cell = ws.cell(row=s5_hdr_row, column=i, value=h)
    style(cell, font=bold_white, fill=blue_fill, border=thin_border,
          align=Alignment(horizontal="center", wrap_text=True))

# Pop reference for scenario 1 (7th decile) = 55048 hardcoded in Section 4 row 26 (D26)
# We reference D26 for population so changes to Scenario 1 propagate
ctrl_splits = [0.05, 0.10, 0.15, 0.20, 0.25, 0.30, 0.50]
base_s5_ref = f"C{stress1_start + 1}"   # the editable baseline cell in this section

for i, ctrl in enumerate(ctrl_splits):
    row = s5_hdr_row + 1 + i  # rows 37-43

    # Col A — blank label
    ws.cell(row=row, column=1).border = thin_border

    # Col B — Control % (green, editable)
    style(ws.cell(row=row, column=2, value=ctrl),
          fill=green_fill, fmt=pct_fmt, border=thin_border)

    # Col C — n_Control = Scenario 1 pop * ctrl%
    ws.cell(row=row, column=3).value = f"=ROUND($D$26*B{row},0)"
    ws.cell(row=row, column=3).number_format = num_fmt
    ws.cell(row=row, column=3).border = thin_border

    # Col D — n_Test
    ws.cell(row=row, column=4).value = f"=ROUND($D$26*(1-B{row}),0)"
    ws.cell(row=row, column=4).number_format = num_fmt
    ws.cell(row=row, column=4).border = thin_border

    # Col E — MDE
    mde_s5 = (
        f"=ROUND(($C$12+$C$13)*SQRT(${base_s5_ref}*(1-${base_s5_ref})"
        f"*(1/C{row}+1/D{row})),6)"
    )
    ws.cell(row=row, column=5).value = mde_s5
    ws.cell(row=row, column=5).number_format = pct_fmt_4
    ws.cell(row=row, column=5).border = thin_border

    # Col F — Detectable Rate
    ws.cell(row=row, column=6).value = f"=${base_s5_ref}+E{row}"
    ws.cell(row=row, column=6).number_format = pct_fmt_4
    ws.cell(row=row, column=6).border = thin_border

    # Col G — Clients Withheld
    ws.cell(row=row, column=7).value = f'=TEXT(C{row},"#,##0")&" clients withheld from DM"'
    ws.cell(row=row, column=7).border = thin_border


# ════════════════════════════════════════════════════════════════════════════
# SECTION 6: STRESS TEST — BASELINE SENSITIVITY  (Rows 45-55)
# ════════════════════════════════════════════════════════════════════════════

stress2_start = 45
for c in range(1, 13):
    style(ws.cell(row=stress2_start, column=c), fill=grey_fill)
style(
    ws.cell(row=stress2_start, column=2,
            value="Stress Test \u2014 Baseline Sensitivity (fixed pop=120,000 | control=20%)"),
    font=bold, fill=grey_fill,
)

# Row 46 — column headers
s6_hdr_row = stress2_start + 1
s6_headers = ["", "Baseline Rate", "MDE", "Detectable Rate", "Notes"]
for i, h in enumerate(s6_headers, start=1):
    cell = ws.cell(row=s6_hdr_row, column=i, value=h)
    style(cell, font=bold_white, fill=blue_fill, border=thin_border,
          align=Alignment(horizontal="center", wrap_text=True))

baselines_s6 = [0.0010, 0.0015, 0.0020, 0.0025, 0.0030, 0.0040, 0.0050, 0.0100]
notes_s6 = [
    "Below observed 7th-decile range",
    "Below observed 7th-decile range",
    "~7th-9th decile blended",
    "~7th-9th decile blended",
    "7th decile (Jan 2026)",
    "",
    "",
    "Top deciles only",
]
# Fixed pop/control for this section
pop_s6    = 120000
ctrl_s6   = 0.20
n_ctrl_s6 = int(pop_s6 * ctrl_s6)     # 24000
n_test_s6 = int(pop_s6 * (1 - ctrl_s6))  # 96000

for i, (base, note) in enumerate(zip(baselines_s6, notes_s6)):
    row = s6_hdr_row + 1 + i  # rows 47-54

    # Col A — blank label
    ws.cell(row=row, column=1).border = thin_border

    # Col B — Baseline (green, editable)
    style(ws.cell(row=row, column=2, value=base),
          fill=green_fill, fmt=pct_fmt_4, border=thin_border)

    # Col C — MDE (hardcoded n_ctrl/n_test since pop/ctrl are fixed)
    mde_s6 = (
        f"=ROUND(($C$12+$C$13)*SQRT(B{row}*(1-B{row})"
        f"*(1/{n_ctrl_s6}+1/{n_test_s6})),6)"
    )
    ws.cell(row=row, column=3).value = mde_s6
    ws.cell(row=row, column=3).number_format = pct_fmt_4
    ws.cell(row=row, column=3).border = thin_border

    # Col D — Detectable Rate
    ws.cell(row=row, column=4).value = f"=B{row}+C{row}"
    ws.cell(row=row, column=4).number_format = pct_fmt_4
    ws.cell(row=row, column=4).border = thin_border

    # Col E — Notes
    ws.cell(row=row, column=5).value = note
    ws.cell(row=row, column=5).border = thin_border


# ════════════════════════════════════════════════════════════════════════════
# SECTION 7: REFERENCE DATA — JAN 2026 DEPLOYMENT  (Rows 56-72)
# ════════════════════════════════════════════════════════════════════════════

ref_start = 56
for c in range(1, 13):
    style(ws.cell(row=ref_start, column=c), fill=grey_fill)
style(
    ws.cell(row=ref_start, column=2,
            value="Reference Data \u2014 Jan 2026 Deployment (2026010PCQ) \u2014 Decile Breakdown"),
    font=bold, fill=grey_fill,
)

# Column headers — blue, NOT editable
s7_headers = [
    "Decile", "Total Clients", "DM Targeted", "DM Coverage",
    "Approved All", "Rate All",
    "Approved Period-ASC", "Rate Period-ASC",
]
for i, h in enumerate(s7_headers, start=2):
    cell = ws.cell(row=ref_start + 1, column=i, value=h)
    style(cell, font=bold_white, fill=blue_fill, border=thin_border,
          align=Alignment(horizontal="center", wrap_text=True))

ref_data = [
    ("1",     51780,  42539, 0.82, 4755, 0.0918, 3840, 0.0742),
    ("2",     48555,  41027, 0.85, 1793, 0.0369, 1445, 0.0298),
    ("3",     49293,  41873, 0.85, 1137, 0.0231,  925, 0.0188),
    ("4",     50270,  39502, 0.79,  730, 0.0145,  585, 0.0116),
    ("5",     47002,  35876, 0.76,  484, 0.0103,  389, 0.0083),
    ("6",     49254,  29641, 0.60,  342, 0.0069,  251, 0.0051),
    ("7",     55048,   3297, 0.06,  252, 0.0046,  166, 0.0030),
    ("8",     43790,    698, 0.02,  121, 0.0028,   78, 0.0018),
    ("9",     68216,    453, 0.01,  115, 0.0017,   76, 0.0011),
    ("10",    68588,    285, 0.00,   57, 0.0008,   39, 0.0006),
    ("1-6",  296154, 230458, 0.78, 9241, 0.0312, 7435, 0.0251),
    ("TOTAL",531796, 235191, None, 9786, 0.0184, 7794, 0.0147),
]
fmts_ref = [None, num_fmt, num_fmt, pct_fmt, num_fmt, pct_fmt_4, num_fmt, pct_fmt_4]

for i, row_data in enumerate(ref_data):
    r = ref_start + 2 + i
    decile, tot, dm_tgt, dm_cov, app_all, rate_all, app_asc, rate_asc = row_data
    vals = [decile, tot, dm_tgt, dm_cov, app_all, rate_all, app_asc, rate_asc]
    for j, (v, fmt) in enumerate(zip(vals, fmts_ref)):
        cell = ws.cell(row=r, column=2 + j)
        if v is None:
            cell.value = "\u2014"
        else:
            cell.value = v
            if fmt:
                cell.number_format = fmt
        cell.border = thin_border


# ════════════════════════════════════════════════════════════════════════════
# SECTION 8: NOTES  (Rows 70+)
# ════════════════════════════════════════════════════════════════════════════

notes_start = 70
for c in range(1, 13):
    style(ws.cell(row=notes_start, column=c), fill=grey_fill)
style(ws.cell(row=notes_start, column=2, value="Notes"), font=bold, fill=grey_fill)

notes = [
    "1. Period-ASC filter required for true campaign-attributed conversions (Daniel Chin, March 26, 2026).",
    "2. Baseline rates from 2026010PCQ (Jan 2026, most mature deployment).",
    "3. 7th decile NOT currently targeted for DM \u2014 volume is leakage (~6% coverage).",
    "4. DM is an awareness channel \u2014 clients apply via Online/Mobile/Branch, not directly via DM.",
    "5. Clients without DM may be Not Eligible, not opt-outs \u2014 invalid baseline comparison.",
    "6. Green cells are editable \u2014 change them to recalculate.",
    "7. MDE = minimum detectable effect in absolute percentage points.",
    "8. Two-sided test. Change Za to 1.6449 for one-sided.",
]
for i, note in enumerate(notes):
    ws.cell(row=notes_start + 1 + i, column=2, value=note)


# ── Zoom ────────────────────────────────────────────────────────────────────
ws.sheet_view.zoomScale = 90

# ── Force full recalculation on load ────────────────────────────────────────
wb.calculation = CalcProperties(fullCalcOnLoad=True, calcMode='auto')

# ── Save ────────────────────────────────────────────────────────────────────
output_path = r"C:\Users\andre\New_projects\cards\campaigns\PCQ\pcq_mde_calculator_v2.xlsx"
try:
    wb.save(output_path)
    print(f"Saved: {output_path}")
except PermissionError:
    output_path = r"C:\Users\andre\New_projects\cards\campaigns\PCQ\pcq_mde_calculator_v2b.xlsx"
    wb.save(output_path)
    print(f"Original locked \u2014 saved as: {output_path}")


# ── Force recalculation via win32com (Windows only) ─────────────────────────
def recalc_with_excel(filepath):
    """Open in Excel, force recalculate all formulas, save and close."""
    try:
        import win32com.client
        abs_path = os.path.abspath(filepath)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb_com = excel.Workbooks.Open(abs_path)
        excel.Calculate()
        wb_com.Save()
        wb_com.Close()
        excel.Quit()
        print(f"Recalculated via Excel: {abs_path}")
        return True
    except ImportError:
        print("win32com not available \u2014 skipping Excel recalculation")
        return False
    except Exception as e:
        print(f"Excel recalculation failed: {e}")
        try:
            excel.Quit()
        except:
            pass
        return False


recalc_with_excel(output_path)
